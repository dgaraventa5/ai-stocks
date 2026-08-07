"""One-time portfolio-construction-v2 migration (spec 2026-08-07 §A3).

Default run: print the before/after table (ticker, current weight, new
inverse-vol target, delta) plus the score→rank membership diff, for Dom's
explicit approval. WRITES NOTHING.

--apply (run ONLY on Dom's approval — the live portfolio must not change
without it, spec §A3): flip tracking/portfolio-config.json to
sizing.mode=inverse_vol + selection.mode=rank, set the Sizing Rules sheet's
Entry rank/Exit rank to 15/18 (the sheet stays authoritative, rule 26), then
run refresh_targets.refresh(resize=True, migration=True) so the book re-sizes
in one `sizing_migration_invvol` event — the seam in the history; nothing is
restated.
"""
from __future__ import annotations

import argparse
import json

from openpyxl import load_workbook

from portfolio_model import PCONFIG, current_weights, load_cfg, load_pcfg
from portfolio_sizing import rank_by_score, topn_membership
from position_sizing import inverse_vol_weights
from recalc_watchlist import recalc
import refresh_targets as rt

N, M = 15, 18   # spec B1 — continuity with the current ~15-name book


def build_table() -> None:
    cfg, pcfg = load_cfg(), load_pcfg()
    held = sorted(cfg['events'][-1]['allocations'])
    live = [x for x in recalc() if x['TOTAL'] is not None]
    live.sort(key=lambda x: -x['TOTAL'])
    info = {x['ticker']: x for x in live}
    layers = {x['ticker']: (x['layer'] or '')[:2] for x in live}

    ranked = rank_by_score(live, set(held))
    include, entered, exit_crossers = topn_membership(set(held), ranked, N, M)
    prices = rt._price_frame(include, int(pcfg['sizing']['lookback']))
    inv = inverse_vol_weights(prices, include, pcfg['sizing'],
                              layers={t: layers.get(t, '') for t in include})
    cur = current_weights(cfg) or {}
    rank = {t: i + 1 for i, t in enumerate(ranked)}

    print(f'Migration table (spec §A3) — N={N}, M={M}, '
          f'lookback {pcfg["sizing"]["lookback"]}d, cap '
          f'{pcfg["sizing"]["max_weight"]:.0%}, floor '
          f'{pcfg["sizing"]["min_weight"]:.0%}. NOTHING APPLIED.\n')
    print(f'{"Tkr":<7}{"Rank":>5}{"Score":>7}{"Now %":>7}{"New %":>7}{"Δ":>7}')
    for t in sorted(set(held) | set(include), key=lambda t: rank.get(t, 999)):
        now, new = cur.get(t, 0.0) * 100, inv.get(t, 0.0) * 100
        note = (' ENTER' if t in entered else
                ' EXIT-CROSS (2-run clock)' if t in exit_crossers else '')
        score = f'{info[t]["TOTAL"]:>7.1f}' if t in info else '      —'
        print(f'{t:<7}{rank.get(t, 0):>5}{score}'
              f'{now:>7.1f}{new:>7.1f}{new - now:>+7.1f}{note}')
    print(f'\nmembership diff (score→rank form): '
          f'+{", ".join(entered) if entered else "none"} / exit-crossers: '
          f'{", ".join(exit_crossers) if exit_crossers else "none"}')
    print('\nApprove with: python3 scripts/migrate_portfolio_v2.py --apply')


def apply() -> None:
    pcfg = json.loads(PCONFIG.read_text())
    pcfg.setdefault('sizing', {})['mode'] = 'inverse_vol'
    pcfg.setdefault('selection', {})['mode'] = 'rank'
    PCONFIG.write_text(json.dumps(pcfg, indent=2) + '\n')
    wb = load_workbook(rt.PORTFOLIO)
    ws = wb['Sizing Rules']
    rt.ensure_params(ws)
    for row in ws.iter_rows():
        if row[0].value == 'Entry rank':
            row[1].value = N
        if row[0].value == 'Exit rank':
            row[1].value = M
    wb.save(rt.PORTFOLIO)
    print(f'modes flipped: sizing=inverse_vol, selection=rank (N={N}, M={M})')
    rt.refresh(resize=True, migration=True)   # logs the one migration event


if __name__ == '__main__':
    ap = argparse.ArgumentParser(
        description='v2 migration: print the §A3 table (default) or apply it')
    ap.add_argument('--apply', action='store_true',
                    help='DOM-GATED: flip modes and re-size the live book in '
                         'one sizing_migration_invvol event')
    args = ap.parse_args()
    apply() if args.apply else build_table()
