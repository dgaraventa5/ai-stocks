"""Gate + staleness audit for SUBJECTIVE ratings (built 2026-06-10).

Why this exists: the Layer 1 re-rating (2026-06-10) found that names were
carrying high AI-Thesis ratings with thesis.md files that were the untouched
template -- i.e. ratings assigned by sub-layer pattern-match, never backed by
per-name research. Same failure that hit Layer 2 in May. Objective inputs have
an earnings-triggered refresh (CLAUDE.md rule 9); subjective ratings had NO
trigger and no gate. This closes both gaps. See CLAUDE.md rule 12.

Two violation classes:
  GATE  (hard) -- name carries subjective ratings but thesis.md is unpopulated
                  (substantially the template). These ratings are unbacked.
                  Exit code is nonzero if any exist, so this can gate scoring.
  STALE (soft) -- name is rated + thesis populated, but its most recent Rating
                  Audit entry is older than --stale-days (default 90).

A thesis is "populated" if it has >MIN_ADDED bytes of content beyond the
template baseline (read live from templates/per-stock-template.md so the
threshold tracks template edits). Brace count is reported for context.

Usage:
  python3 scripts/audit_rating_integrity.py                 # full report
  python3 scripts/audit_rating_integrity.py --layer 01      # one layer
  python3 scripts/audit_rating_integrity.py --summary       # one-line counts
  python3 scripts/audit_rating_integrity.py --stale-days 60
  python3 scripts/audit_rating_integrity.py --stalest 5     # N stalest rated names
"""
from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

from common import ROOT

XLSX = ROOT / '00-master' / 'ai_supply_chain_scoring.xlsx'
TEMPLATE = ROOT / 'templates' / 'per-stock-template.md'
AI_COLS = [20, 21, 22, 23, 24]   # D1..D5
MIN_ADDED = 400                  # bytes beyond template baseline = "worked on"


def template_baseline() -> int:
    return TEMPLATE.stat().st_size if TEMPLATE.exists() else 3789


def thesis_populated(ticker: str, baseline: int) -> bool:
    """A thesis counts as populated if it has >MIN_ADDED bytes beyond template."""
    f = ROOT / 'per-stock' / ticker / 'thesis.md'
    return f.exists() and (f.stat().st_size - baseline) >= MIN_ADDED


def newest_briefing_age(ticker: str, today: dt.date) -> int | None:
    """Days since the newest context-YYYY-MM-DD.md briefing, or None if none.

    /refresh-context writes per-name research to context-{date}.md (thesis.md
    only changes on material shifts, per CLAUDE.md rule 6), so a recent briefing
    is valid evidence the ratings are research-backed."""
    d = ROOT / 'per-stock' / ticker
    if not d.exists():
        return None
    best = None
    for f in d.glob('context-*.md'):
        try:
            day = dt.date.fromisoformat(f.stem.replace('context-', ''))
        except ValueError:
            continue
        age = (today - day).days
        best = age if best is None else min(best, age)
    return best


def last_audit_dates(wb) -> dict[str, str]:
    """ticker -> most recent Rating Audit date (ISO string)."""
    out: dict[str, str] = {}
    a = wb['Rating Audit']
    for row in a.iter_rows(min_row=2, values_only=True):
        d, tkr = row[0], row[1]
        if not tkr:
            continue
        ds = d.isoformat() if isinstance(d, (dt.date, dt.datetime)) else str(d)
        if ds and (tkr not in out or ds > out[tkr]):
            out[tkr] = ds
    return out


def audit(layer_filter=None, stale_days=90):
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['Watchlist']
    baseline = template_baseline()
    audits = last_audit_dates(wb)
    today = dt.date.today()

    rows = []
    for r in range(2, ws.max_row + 1):
        tkr = ws.cell(row=r, column=1).value
        if not tkr:
            continue
        layer = str(ws.cell(row=r, column=3).value or '')
        if layer_filter and not layer.startswith(layer_filter):
            continue
        rated = any(ws.cell(row=r, column=c).value not in (None, '') for c in AI_COLS)
        if not rated:
            continue
        has_thesis = thesis_populated(tkr, baseline)
        brief_age = newest_briefing_age(tkr, today)
        last = audits.get(tkr)
        audit_age = (today - dt.date.fromisoformat(last)).days if last else None
        # Backed = a durable thesis OR any per-name research briefing exists.
        backed = has_thesis or (brief_age is not None)
        # Review age = freshest evidence of an actual review pass.
        ages = [a for a in (brief_age, audit_age) if a is not None]
        review_age = min(ages) if ages else None
        gate = not backed
        stale = backed and (review_age is None or review_age > stale_days)
        backing = ('thesis' if has_thesis
                   else f'brief {brief_age}d' if brief_age is not None
                   else 'NONE')
        rows.append(dict(tkr=tkr, layer=layer[:2], backing=backing,
                         last=last, age=review_age, gate=gate, stale=stale))
    return rows


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--layer', default=None, help='filter, e.g. 01')
    ap.add_argument('--stale-days', type=int, default=90)
    ap.add_argument('--summary', action='store_true')
    ap.add_argument('--stalest', type=int, default=0,
                    help='print the N stalest rated names (for refresh rotation)')
    args = ap.parse_args()

    rows = audit(args.layer, args.stale_days)
    gates = [x for x in rows if x['gate']]
    stales = [x for x in rows if x['stale']]

    if args.stalest:
        # Priority for the refresh rotation: unbacked names (gate violations,
        # no research at all) are maximally stale and come first; then backed
        # names by descending review age. Refreshing a name resets its age, so
        # the rotation self-balances across runs. Excludes delisted tickers
        # (review age None AND a stale audit) only if --skip-dead given.
        INF = 10 ** 9
        ranked = sorted(
            rows,
            key=lambda x: -(INF if x['gate'] or x['age'] is None else x['age']))
        for x in ranked[:args.stalest]:
            print(x['tkr'])
        return 0

    scope = f"layer {args.layer}" if args.layer else "all layers"
    if args.summary:
        print(f"rating-integrity ({scope}): {len(rows)} rated names | "
              f"{len(gates)} UNGATED (no thesis) | {len(stales)} stale "
              f"(>{args.stale_days}d)")
        return 1 if gates else 0

    print(f"=== Subjective-rating integrity audit — {scope} "
          f"({len(rows)} rated names) ===\n")
    print(f"GATE violations — rated with NO thesis and NO research briefing "
          f"({len(gates)}):")
    if gates:
        print(f"  {'Tkr':<6}{'Lyr':<4}{'backing':<10}{'last rated':<12}")
        for x in sorted(gates, key=lambda x: x['layer']):
            print(f"  {x['tkr']:<6}{x['layer']:<4}{x['backing']:<10}"
                  f"{x['last'] or '—':<12}")
    else:
        print("  none")

    print(f"\nSTALE — populated thesis but rated >{args.stale_days}d ago "
          f"({len(stales)}):")
    if stales:
        print(f"  {'Tkr':<6}{'Lyr':<4}{'last rated':<12}{'age':>5}")
        for x in sorted(stales, key=lambda x: -(x['age'] or 0)):
            print(f"  {x['tkr']:<6}{x['layer']:<4}{x['last'] or '—':<12}{x['age']!s:>5}")
    else:
        print("  none")

    print(f"\nGATE blocks scoring. Run /refresh-context on UNGATED names before "
          f"trusting their AI ratings. Exit nonzero = {len(gates)} gate "
          f"violation(s).")
    return 1 if gates else 0


if __name__ == '__main__':
    raise SystemExit(main())
