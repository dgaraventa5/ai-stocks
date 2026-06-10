"""Cut D3 (Moat) for the four nat-gas E&Ps (2026-06-10 follow-up).

The 2026-06-10 directness re-rating set these names to D2=2 (commoditized
price-takers) but left D3=4 (strong moat) untouched, which is internally
contradictory — commodity gas producers are the textbook no-moat business.
Their "moat" is a modest low-cost-acreage / scale cost-advantage, not a durable
competitive barrier. EQT keeps a notch for integrated midstream (Equitrans +
Olympus) + largest-in-basin scale; RRC/AR/EXE drop to 2.
"""
from __future__ import annotations

from openpyxl import load_workbook

XLSX = '/Users/dom/Desktop/ai-stocks/00-master/ai_supply_chain_scoring.xlsx'
DATE = '2026-06-10'
D3_COL = 22

CHANGES = {
    'EQT': (3, "Commodity gas producer has no durable moat; D2=2 (commoditized) is incompatible with D3=4. Keeps a notch above peers for integrated midstream (Equitrans+Olympus) + largest in-basin scale = modest cost advantage, not a barrier.",
            'EQT 10-Q Q1 2026; directness re-rate 2026-06-10', 'High'),
    'RRC': (2, "Commodity gas producer, weak moat; low-cost Appalachian acreage is a modest cost advantage, not durable defensibility. Aligns D3 with D2=2 (price-taker, sells below Henry Hub).",
            'RRC 8-K 2026-04-21; directness re-rate 2026-06-10', 'High'),
    'AR':  (2, "Commodity gas producer, weak moat; superior firm-transport/LNG logistics is market access, not a defensible barrier. Aligns D3 with D2=2.",
            'AR 10-Q Q1 2026; directness re-rate 2026-06-10', 'High'),
    'EXE': (2, "Commodity gas producer, weak moat; #1-scale is a cost advantage within a commodity game, not pricing-power defensibility. Aligns D3 with D2=2 (sells below NYMEX).",
            'EXE 10-Q Q1 2026; directness re-rate 2026-06-10', 'High'),
}


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)
    ws = wb['Watchlist']
    audit = wb['Rating Audit']
    rows = {ws.cell(row=r, column=1).value: r
            for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}

    for tkr, (new, rationale, source, conf) in CHANGES.items():
        r = rows[tkr]
        old = ws.cell(row=r, column=D3_COL).value
        ws.cell(row=r, column=D3_COL).value = new
        audit.append([DATE, tkr, 'D3', new, f'(was {old}) {rationale}',
                      source, conf, 'Update'])
        ws.cell(row=r, column=4).value = DATE
        print(f'{tkr} D3: {old} -> {new}')

    wb.save(XLSX)
    print('done')


if __name__ == '__main__':
    main()
