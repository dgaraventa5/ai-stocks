"""Refresh EV/FCF multiples for the reverse-DCF (P2) metric into
00-master/reverse-dcf.json.

recalc_watchlist reads this file plus each name's Rev-3y-CAGR (from the sheet)
to compute the reverse-DCF mispricing score — a 6th Value sub-metric (rule 21):
implied FCF growth solved from EV/FCF vs. the grounded growth estimate; a bigger
positive gap = cheaper-vs-expectations = higher score.

EV/FCF is kept currency-consistent (rule 19): mapped ADRs (TSM/UMC) use their
local listing; other foreign filers FX-convert the market cap; USD names use it
directly. Blank when FCF is non-positive/unavailable, so the name simply drops
the reverse-DCF term from its Value average.

Usage:  python3 scripts/refresh_reverse_dcf.py            # all watchlist names
        python3 scripts/refresh_reverse_dcf.py NVDA TSM   # a subset
Run alongside objective refreshes (rule 9).
"""
from __future__ import annotations

import json
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
JSON_PATH = ROOT / "00-master" / "reverse-dcf.json"
SCORING = ROOT / "00-master" / "ai_supply_chain_scoring.xlsx"


def _watchlist_tickers():
    ws = load_workbook(SCORING, data_only=False)["Watchlist"]
    return [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value]


def _ev_over_fcf(ticker):
    """EV/FCF for one name in a single, consistent currency. None if unavailable."""
    import yfinance as yf
    from batch_score import statement_fcf
    import adr_currency

    def ev_from(info, mcap):
        ev = info.get("enterpriseValue")
        if ev is None and mcap is not None:
            ev = mcap + (info.get("totalDebt") or 0) - (info.get("totalCash") or 0)
        return ev

    t = yf.Ticker(ticker)
    info = t.info or {}
    trading, financial = info.get("currency"), info.get("financialCurrency")

    if adr_currency.has_currency_mismatch(trading, financial):
        local = adr_currency.ADR_LOCAL.get(ticker)
        if local:                                   # mapped US ADR -> local listing
            lt = yf.Ticker(local)
            li = lt.info or {}
            ev, fcf = ev_from(li, li.get("marketCap")), statement_fcf(lt)
        else:                                       # foreign primary listing -> FX
            rate = adr_currency.fx_rate(trading, financial)
            mcap = info.get("marketCap")
            ev = ((mcap * rate) + (info.get("totalDebt") or 0)
                  - (info.get("totalCash") or 0)) if (mcap and rate) else None
            fcf = statement_fcf(t)
    else:
        ev, fcf = ev_from(info, info.get("marketCap")), statement_fcf(t)

    if ev is not None and fcf is not None and fcf > 0:
        return round(ev / fcf, 2)
    return None


def main(argv):
    # Timestamp passed in (scripts can't call Date.now equivalents in some envs);
    # here we're a normal CLI so use today.
    import datetime as dt
    today = dt.date.today().isoformat()
    targets = [a.upper() for a in argv[1:]] or _watchlist_tickers()

    data = {}
    if JSON_PATH.exists():
        data = json.loads(JSON_PATH.read_text())

    ok = blank = 0
    for i, tk in enumerate(targets, 1):
        try:
            v = _ev_over_fcf(tk)
        except Exception as e:
            v = None
            print(f"  {tk}: error {e}")
        data[str(tk)] = {"ev_fcf": v, "asof": today}
        ok += v is not None
        blank += v is None
        if i % 20 == 0:
            print(f"  ... {i}/{len(targets)}")
        time.sleep(0.2)

    JSON_PATH.write_text(json.dumps(data, indent=1, sort_keys=True) + "\n")
    print(f"wrote {JSON_PATH} — {ok} with EV/FCF, {blank} blank, {len(targets)} processed")


if __name__ == "__main__":
    main(sys.argv)
