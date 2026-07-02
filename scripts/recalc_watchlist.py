"""Recalculate Watchlist TOTAL scores in Python using canonical methodology.

Mirrors the Excel formulas in Methodology tab so we can read TOTALs without
needing Excel/LibreOffice to evaluate cached formulas. Read-only — does not
modify the workbook.
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

import cohort_percentile
import reverse_dcf

_REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = str(_REPO_ROOT / '00-master/ai_supply_chain_scoring.xlsx')


def _num(v):
    if v is None or v == "": return None
    if isinstance(v, str):
        try: return float(v)
        except ValueError: return None
    return v


def _band(value, bands):
    """bands: list of (predicate_value, score) ordered from best to worst.
    For lower-is-better, predicate is "value <= threshold".
    For higher-is-better, predicate is "value >= threshold".
    Last entry is the catch-all (e.g., (None, 5)).
    """
    if value is None or value == "":
        return None
    for thresh, score in bands:
        if thresh is None:
            return score
        if isinstance(thresh, tuple):
            op, t = thresh
            if op == '<=' and value <= t: return score
            if op == '>=' and value >= t: return score
        else:
            if value >= thresh: return score
    return None


def score_fwd_pe(v):
    v = _num(v)
    if v is None: return None
    if v < 0: return 5
    for t, s in [(15, 90), (20, 75), (25, 60), (30, 45), (40, 30), (60, 15)]:
        if v <= t: return s
    return 5

def score_ev_ebitda(v):
    v = _num(v)
    if v is None: return None
    if v < 0: return 5
    for t, s in [(10, 90), (15, 75), (20, 60), (25, 45), (35, 30), (50, 15)]:
        if v <= t: return s
    return 5

def score_ev_fcf_saas(v):
    # Rule 10 (2026-05-26): Layer 10 col F holds EV/FCF, SaaS-calibrated bands.
    # Bands were documented in CLAUDE.md but not implemented in formulas/recalc
    # until 2026-06-12.
    v = _num(v)
    if v is None: return None
    if v < 0: return 5
    for t, s in [(20, 100), (30, 90), (40, 75), (55, 60), (75, 45), (100, 30)]:
        if v <= t: return s
    return 15

def score_ev_mw(v):
    # Rule 13 (2026-06-12): Layer 9 capacity cohort col F holds EV per secured
    # gross MW ($M/MW). Anchor: ~$9-10M/gross-MW replacement cost at the 45/60
    # boundary. MW data: 00-master/capacity-mw.json.
    v = _num(v)
    if v is None: return None
    for t, s in [(2, 100), (4, 90), (6, 75), (9, 60), (12, 45), (18, 30)]:
        if v <= t: return s
    return 15

def is_layer10(layer):
    return bool(layer) and str(layer).strip().startswith('10')

def is_layer9_capacity(layer):
    if not layer or not str(layer).strip().startswith('09'):
        return False
    l = str(layer).lower()
    return 'bitcoin' in l or 'neocloud' in l

def score_ev_col(v, layer):
    """Layer-conditional col F: EV/EBITDA standard, EV/FCF for Layer 10,
    EV/MW for the Layer 9 capacity cohort."""
    if is_layer10(layer):
        return score_ev_fcf_saas(v)
    if is_layer9_capacity(layer):
        return score_ev_mw(v)
    return score_ev_ebitda(v)

def score_fcf_yield(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(8, 100), (6, 90), (4, 75), (2, 60), (1, 45), (0, 30)]:
        if v >= t: return s
    return 15

def score_ps(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(2, 90), (4, 75), (7, 60), (10, 45), (15, 30), (25, 15)]:
        if v <= t: return s
    return 5

def score_peg(v):
    # PEG added to Value 2026-05-25. Mirrors Watchlist col I band.
    v = _num(v)
    if v is None: return None
    for t, s in [(0.5, 100), (1, 90), (1.5, 75), (2, 60), (2.5, 45), (3, 30)]:
        if v <= t: return s
    return 15

def score_roic(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(25, 100), (20, 90), (15, 75), (10, 60), (5, 45), (0, 30)]:
        if v >= t: return s
    return 15

def score_gm(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(60, 100), (50, 90), (40, 75), (30, 60), (20, 45), (10, 30)]:
        if v >= t: return s
    return 15

def score_fcf_mgn(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(25, 100), (20, 90), (15, 75), (10, 60), (5, 45), (0, 30)]:
        if v >= t: return s
    return 15

def score_nd_ebitda(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(0, 100), (1, 90), (2, 75), (3, 60), (4, 45), (5, 30)]:
        if v <= t: return s
    return 15

def score_rev_cagr(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(40, 100), (30, 90), (20, 75), (15, 60), (10, 45), (5, 30)]:
        if v >= t: return s
    return 15

def score_rev_yoy(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(40, 100), (30, 90), (20, 75), (10, 60), (5, 45), (0, 30)]:
        if v >= t: return s
    return 15

def score_eps_yoy(v):
    v = _num(v)
    if v is None: return None
    for t, s in [(40, 100), (30, 90), (20, 75), (10, 60), (0, 45), (-10, 30)]:
        if v >= t: return s
    return 15

def score_50dma(v):
    # % of last 120 trading days with close > 50-day SMA. Added 2026-06-09.
    # Mirrors Watchlist col AC band inside the Momentum Score formula.
    v = _num(v)
    if v is None: return None
    for t, s in [(85, 100), (70, 90), (55, 75), (40, 60), (25, 40)]:
        if v >= t: return s
    return 20


def avg_nonnull(vals):
    xs = [v for v in vals if v is not None]
    return sum(xs) / len(xs) if xs else None


def tier(t):
    if t is None: return None
    if t >= 85: return '✓✓✓'
    if t >= 70: return '✓✓'
    if t >= 55: return '✓'
    if t >= 40: return '?'
    return '✗'


def _read_raw(ws, r):
    """Raw inputs for one row (PEG recomputed like the Watchlist formula)."""
    fwd_pe = ws.cell(row=r, column=5).value
    E, Rv = _num(fwd_pe), _num(ws.cell(row=r, column=18).value)
    peg = (E / Rv) if (E is not None and Rv is not None and Rv > 0 and E >= 0) else None
    layer = ws.cell(row=r, column=3).value
    return dict(
        ticker=ws.cell(row=r, column=1).value, layer=layer, row=r,
        tl=cohort_percentile.top_level_layer(layer),
        fwd_pe=fwd_pe, ev_ebitda=ws.cell(row=r, column=6).value,
        fcf_yield=ws.cell(row=r, column=7).value, ps=ws.cell(row=r, column=8).value,
        peg=peg, roic=ws.cell(row=r, column=11).value, gm=ws.cell(row=r, column=12).value,
        fcf_mgn=ws.cell(row=r, column=13).value, nd_eb=ws.cell(row=r, column=14).value,
        rev_cagr=ws.cell(row=r, column=16).value, rev_yoy=ws.cell(row=r, column=17).value,
        eps_yoy=ws.cell(row=r, column=18).value,
        ai_inputs=[ws.cell(row=r, column=c).value for c in (20, 21, 22, 23, 24)],
        mom_inputs=[ws.cell(row=r, column=c).value for c in (26, 27, 28)],
        dma_pct=ws.cell(row=r, column=29).value,
        risk_inputs=[ws.cell(row=r, column=c).value for c in (31, 32, 33, 34, 38)],
    )


def _metric_scores_absolute(row):
    """The six style-biased metric scores via the live absolute bands."""
    return dict(
        ev=score_ev_col(row['ev_ebitda'], row['layer']),
        fcf_yield=score_fcf_yield(row['fcf_yield']), ps=score_ps(row['ps']),
        roic=score_roic(row['roic']), gm=score_gm(row['gm']),
        fcf_mgn=score_fcf_mgn(row['fcf_mgn']))


def _metric_scores_percentile(rows):
    """The six style-biased metric scores as within-cohort percentiles (P1),
    keyed by row number. Cohort = top-level layer; min size 8 else absolute
    fallback (cohort_percentile.cohort_metric_scores). Col F (EV/*) stays
    absolute for Layer 9 (mixed EV/EBITDA + EV/MW), EV/FCF for Layer 10."""
    by_layer = {}
    for row in rows:
        by_layer.setdefault(row['tl'], []).append(row)
    out = {row['row']: {} for row in rows}
    uniform = [('fcf_yield', True, score_fcf_yield), ('ps', False, score_ps),
               ('roic', True, score_roic), ('gm', True, score_gm),
               ('fcf_mgn', True, score_fcf_mgn)]
    for tl, cohort in by_layer.items():
        for key, hib, absfn in uniform:
            vals = [_num(r[key]) for r in cohort]
            for r, s in zip(cohort,
                            cohort_percentile.cohort_metric_scores(vals, hib, absfn)):
                out[r['row']][key] = s
        vals = [_num(r['ev_ebitda']) for r in cohort]
        if tl == '09':
            ev = [score_ev_col(r['ev_ebitda'], r['layer']) for r in cohort]
        elif tl == '10':
            ev = cohort_percentile.cohort_metric_scores(vals, False, score_ev_fcf_saas)
        else:
            ev = cohort_percentile.cohort_metric_scores(vals, False, score_ev_ebitda)
        for r, s in zip(cohort, ev):
            out[r['row']]['ev'] = s
    return out


def _assemble(row, ms, w):
    """Category subscores + TOTAL, using the cohort/absolute metric scores `ms`
    for the six style-biased metrics and live absolute bands for the rest."""
    value = avg_nonnull([score_fwd_pe(row['fwd_pe']), ms['ev'], ms['fcf_yield'],
                         ms['ps'], score_peg(row['peg']),
                         reverse_dcf.reverse_dcf_score(row.get('ev_fcf'),
                                                       _num(row['rev_cagr']))])
    quality = avg_nonnull([ms['roic'], ms['gm'], ms['fcf_mgn'],
                           score_nd_ebitda(row['nd_eb'])])
    growth = avg_nonnull([score_rev_cagr(row['rev_cagr']), score_rev_yoy(row['rev_yoy']),
                          score_eps_yoy(row['eps_yoy'])])
    # AI Thesis / Risk map rating -> score as (mean-1)*25 (rule 22, P6): 1->0,
    # 3->50, 5->100, so a genuine "zero" dimension can drag the score to 0 rather
    # than flooring at 20. Momentum keeps rating*20 (it blends with the objective
    # 50DMA band, out of P6 scope).
    ai_present = [x for x in row['ai_inputs'] if x is not None]
    ai = ((sum(ai_present) / len(ai_present) - 1) * 25) if ai_present else None
    momentum = avg_nonnull([x * 20 if x is not None else None for x in row['mom_inputs']]
                           + [score_50dma(row['dma_pct'])])
    risk_present = [x for x in row['risk_inputs'] if x is not None]
    risk = ((sum(risk_present) / len(risk_present) - 1) * 25) if risk_present else None
    parts = []
    for sub, key in [(value, 'Value'), (quality, 'Quality'), (growth, 'Growth'),
                     (ai, 'AI'), (momentum, 'Momentum'), (risk, 'Risk')]:
        if sub is not None:
            parts.append(sub * w[key])
    total = sum(parts) if parts else None
    return {'ticker': row['ticker'], 'layer': row['layer'], 'row': row['row'],
            'Value': value, 'Quality': quality, 'Growth': growth, 'AI': ai,
            'Momentum': momentum, 'Risk': risk, 'TOTAL': total, 'Tier': tier(total)}


def _load_ev_fcf_map():
    """{ticker: ev_fcf} from 00-master/reverse-dcf.json (the P2 reverse-DCF input,
    refreshed by refresh_reverse_dcf.py). Empty if the file is absent, so the
    reverse-DCF term simply drops out of the Value average."""
    import json
    p = _REPO_ROOT / '00-master' / 'reverse-dcf.json'
    if not p.exists():
        return {}
    try:
        data = json.loads(p.read_text())
    except Exception:
        return {}
    return {k: (v.get('ev_fcf') if isinstance(v, dict) else v)
            for k, v in data.items()}


def recalc(xlsx=XLSX, mode='percentile'):
    """Recompute Watchlist scores.

    mode='percentile' (LIVE default, P1): the six style-biased metrics (EV/EBITDA,
    FCF-yield, P/S, ROIC, gross margin, FCF margin) are ranked within each
    top-level-layer cohort; a cohort with <8 non-null values for a metric falls
    back to that metric's absolute band. Everything else (ND/EBITDA, Fwd P/E, PEG,
    Growth, and all subjective dims) is scored on absolute bands exactly as before.

    mode='absolute' reproduces the pre-P1 fixed-band scores (for before/after).
    """
    wb = load_workbook(xlsx, data_only=False)
    ws = wb['Watchlist']
    weights = {}
    if 'Weights' in wb.sheetnames:
        for row in wb['Weights'].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                weights[row[0]] = row[1]
    w = {
        'Value': weights.get('Value', 0.20), 'Quality': weights.get('Quality', 0.20),
        'Growth': weights.get('Growth', 0.15),
        'AI': weights.get('AI Thesis', weights.get('AI', 0.20)),
        'Momentum': weights.get('Momentum', 0.10), 'Risk': weights.get('Risk', 0.15)}

    rows = [_read_raw(ws, r) for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value]
    ev_fcf_map = _load_ev_fcf_map()
    for row in rows:
        row['ev_fcf'] = ev_fcf_map.get(row['ticker'])
    if mode == 'percentile':
        ms = _metric_scores_percentile(rows)
    else:
        ms = {row['row']: _metric_scores_absolute(row) for row in rows}
    return [_assemble(row, ms[row['row']], w) for row in rows]


def sync_scores(xlsx=XLSX):
    """Write the P1 percentile Value Score (col J) and Quality Score (col O) to
    the Watchlist as VALUES.

    These two subscores are cohort-relative (they rank a name against its peers),
    which cannot be expressed as a per-row Excel formula. The TOTAL and Tier
    cells stay live formulas that reference J and O, so on open Excel computes the
    correct percentile TOTAL/Tier while Growth/AI/Momentum/Risk remain live
    formulas that auto-update from their inputs. recalc() is the authoritative
    engine for every consumer (site, portfolio) regardless; this only keeps the
    raw sheet honest when opened directly.

    Re-run after rebuild_watchlist_formulas (which restores the absolute-band
    formulas here) and after any objective-input refresh. Returns rows written.
    """
    results = recalc(xlsx, mode='percentile')
    wb = load_workbook(xlsx, data_only=False)
    ws = wb['Watchlist']
    n = 0
    for x in results:
        if x['Value'] is not None:
            ws.cell(row=x['row'], column=10).value = round(x['Value'], 1)
        if x['Quality'] is not None:
            ws.cell(row=x['row'], column=15).value = round(x['Quality'], 1)
        n += 1
    wb.save(xlsx)
    return n


if __name__ == '__main__':
    import sys
    if '--sync' in sys.argv:
        n = sync_scores()
        print(f'synced P1 percentile Value/Quality Score cells for {n} rows '
              f'(TOTAL/Tier formulas reference them)')
        sys.exit(0)
    rs = recalc()
    # Sort by TOTAL desc (None last)
    rs.sort(key=lambda x: (x['TOTAL'] is None, -(x['TOTAL'] or 0)))
    print(f"{'Tkr':<7}{'Layer':<35}{'Val':>6}{'Qual':>6}{'Grw':>6}{'AI':>6}{'Mom':>6}{'Risk':>6}{'TOT':>7}{'Tier':>5}")
    for x in rs:
        def f(v): return f"{v:5.1f}" if isinstance(v,(int,float)) else "  -- "
        print(f"{x['ticker']:<7}{(x['layer'] or '')[:34]:<35}{f(x['Value'])}{f(x['Quality'])}{f(x['Growth'])}{f(x['AI'])}{f(x['Momentum'])}{f(x['Risk'])}{f(x['TOTAL']):>7}{(x['Tier'] or ''):>5}")
