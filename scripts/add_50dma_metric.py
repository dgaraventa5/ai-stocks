"""One-off migration: add '50DMA %' as a 4th Momentum metric (2026-06-09).

Approved by Dom 2026-06-09. Adds an objective momentum metric — the share of
the last 120 trading days where close > 50-day SMA — to distinguish consistent
uptrends from one-gap bounces. Suggested by external feedback, reviewed against
the existing model before adoption.

Changes:
  - Watchlist: insert col AC ('50DMA %') between Insider (AB) and Momentum
    Score. Momentum Score (now AD) blends the three subjective 1-5 ratings
    (x20) with the banded 50DMA score. Risk Score / TOTAL / Tier formulas are
    rewritten because openpyxl shifts cells without updating references.
  - Methodology: band row under MOMENTUM (>=85->100, >=70->90, >=55->75,
    >=40->60, >=25->40, <25->20).
  - README: refresh the category-weights block (was still pre-2026-05-25).
  - Rating Audit: append a Methodology-change entry.

Mirrors the precedent of reweight_and_add_peg.py (2026-05-25).
"""
from __future__ import annotations

from copy import copy

from openpyxl import load_workbook

XLSX = '/Users/dom/Desktop/ai-stocks/00-master/ai_supply_chain_scoring.xlsx'

# Banded conversion, same nesting style as the other objective metrics.
DMA_BAND = ('IF({c}{r}>=85,100,IF({c}{r}>=70,90,IF({c}{r}>=55,75,'
            'IF({c}{r}>=40,60,IF({c}{r}>=25,40,20)))))')


def migrate() -> None:
    wb = load_workbook(XLSX, data_only=False)

    # ---------- Watchlist ----------
    ws = wb['Watchlist']
    assert ws.cell(row=1, column=28).value == 'Insider', 'schema drift — abort'
    assert ws.cell(row=1, column=29).value == 'Momentum Score', 'schema drift — abort'

    ws.insert_cols(29)
    hdr = ws.cell(row=1, column=29, value='50DMA %')
    src_hdr = ws.cell(row=1, column=28)
    for attr in ('font', 'fill', 'border', 'alignment'):
        setattr(hdr, attr, copy(getattr(src_hdr, attr)))
    ws.column_dimensions['AC'].width = ws.column_dimensions['AB'].width or 10

    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=1).value:
            continue
        # Input styling: 50DMA % is an input cell like Insider next door.
        dma = ws.cell(row=r, column=29)
        src = ws.cell(row=r, column=28)
        for attr in ('font', 'fill', 'border', 'alignment'):
            setattr(dma, attr, copy(getattr(src, attr)))
        dma.number_format = '0.0'

        band = DMA_BAND.format(c='AC', r=r)
        # Momentum Score (AD): 3 subjective ratings x20 + banded 50DMA, same
        # skip-blanks idiom as the Value/Quality/Growth score formulas.
        ws.cell(row=r, column=30).value = (
            f'=IFERROR(AVERAGE('
            f'IFERROR(IF(Z{r}="","",Z{r}*20),""),'
            f'IFERROR(IF(AA{r}="","",AA{r}*20),""),'
            f'IFERROR(IF(AB{r}="","",AB{r}*20),""),'
            f'IFERROR(IF(AC{r}="","",{band}),"")'
            f'),"")'
        )
        # Risk Score (AI): inputs shifted AD:AG -> AE:AH.
        ws.cell(row=r, column=35).value = (
            f'=IFERROR(AVERAGE(AE{r},AF{r},AG{r},AH{r})*20,"")'
        )
        # TOTAL (AJ) and Tier (AK).
        ws.cell(row=r, column=36).value = (
            f'=IFERROR(J{r}*Weights!$B$4 + O{r}*Weights!$B$5 + S{r}*Weights!$B$6'
            f' + Y{r}*Weights!$B$7 + AD{r}*Weights!$B$8 + AI{r}*Weights!$B$9,"")'
        )
        ws.cell(row=r, column=37).value = (
            f'=IF(AJ{r}="","",IF(AJ{r}>=85,"✓✓✓",IF(AJ{r}>=70,"✓✓",'
            f'IF(AJ{r}>=55,"✓",IF(AJ{r}>=40,"?","✗")))))'
        )

    # ---------- Methodology ----------
    m = wb['Methodology']
    assert m.cell(row=32, column=1).value == 'MOMENTUM (subjective 1-5)'
    assert 'Insider Activity' in str(m.cell(row=35, column=1).value)
    m.cell(row=32, column=1).value = (
        'MOMENTUM (subjective 1-5 except 50DMA %, which is objective and banded)')
    m.insert_rows(36)
    # insert_rows does NOT shift merged ranges, and openpyxl silently drops
    # writes to merged non-anchor cells. The blank separator rows are merged
    # A:H, so the stale A36:H36 merge must be removed before writing the band
    # row, then restored one row down where the separator now lives.
    # (drop the range declaration directly — unmerge_cells() KeyErrors here
    # because insert_rows moved the MergedCell placeholders to row 37 while
    # the declared range stayed at 36)
    for rng in list(m.merged_cells.ranges):
        if str(rng) == 'A36:H36':
            m.merged_cells.ranges.remove(rng)
    new = ['% days above 50-DMA, last 120 trading days (higher = better)',
           'Threshold:', '≥85 → 100', '≥70 → 90', '≥55 → 75', '≥40 → 60',
           '≥25 → 40', '<25 → 20', None,
           'Added 2026-06-09. Share of last 120 trading days with close > '
           '50-day SMA, from yfinance price history '
           '(scripts/momentum_50dma.py). Separates consistent uptrends from '
           'one-bounce moves. Names with <60 days of valid history are left '
           'blank and flagged.']
    for c, v in enumerate(new, start=1):
        cell = m.cell(row=36, column=c, value=v)
        src = m.cell(row=10, column=c)  # FCF Yield row — canonical band styling
        for attr in ('font', 'fill', 'border', 'alignment'):
            setattr(cell, attr, copy(getattr(src, attr)))
    m.merge_cells('A37:H37')  # restore the separator band before RISK

    # ---------- README: weights block was still pre-2026-05-25 ----------
    rd = wb['README']
    fixes = {
        12: 'Value (20%)       - Is the price reasonable vs. fundamentals?',
        13: 'Quality (20%)     - Is the business well-run and financially sound?',
        14: 'Growth (15%)      - Is the company growing meaningfully?',
        15: 'AI Thesis (20%)   - How strong is the AI infrastructure exposure?',
        16: 'Momentum (10%)    - Are estimates rising and price acting well?',
        17: 'Risk (15%)        - What concentration / geographic / regulatory risk exists?',
    }
    for r, text in fixes.items():
        rd.cell(row=r, column=1).value = text

    # ---------- Rating Audit ----------
    audit = wb['Rating Audit']
    audit.append([
        '2026-06-09', 'ALL', 'Methodology', '—',
        'Added 4th Momentum metric: % of last 120 trading days with close > '
        '50-day SMA (objective, from yfinance price history). Bands: '
        '≥85→100, ≥70→90, ≥55→75, ≥40→60, ≥25→40, <25→20. Momentum Score now '
        'averages the 3 subjective ratings (×20) with the banded 50DMA score. '
        'Rationale: objectifies trend consistency; separates durable uptrends '
        'from one-gap bounces. Origin: external feedback reviewed 2026-06-09; '
        '5 sibling suggestions declined (P/B, op margin, volume, inst. '
        'ownership, analyst PTs) — see session notes. Approved by Dom.',
        'scripts/momentum_50dma.py; Methodology tab row 36', 'High',
        'Methodology change', None,
    ])

    wb.save(XLSX)
    print('migration complete')


if __name__ == '__main__':
    migrate()
