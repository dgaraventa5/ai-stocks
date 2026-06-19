"""Weekly scan runner for 2026-06-19.

Covers filings from 2026-06-12 through 2026-06-19.
Steps:
  1. Read watchlist tickers + tiers
  2. Query EDGAR for 8-Ks (and 6-Ks for foreign filers)
  3. Print material events + routine filings
  4. Check for 13F-HR filings from tracked funds
  5. List any earnings 8-Ks (Item 2.02) for Rule-9 refresh
"""

from __future__ import annotations

import json
import sys
import time
from datetime import date, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCORING_PATH = ROOT / "00-master" / "ai_supply_chain_scoring.xlsx"
SEC_HEADERS_DATA = {
    "User-Agent": "Dom Researcher dgaraventa5@gmail.com",
    "Accept-Encoding": "gzip, deflate",
    "Host": "data.sec.gov",
}
SEC_HEADERS_WWW = {
    "User-Agent": "Dom Researcher dgaraventa5@gmail.com",
    "Accept-Encoding": "gzip, deflate",
    "Host": "www.sec.gov",
}

SCAN_FROM = date(2026, 6, 12)
SCAN_TO   = date(2026, 6, 19)

# Tracked 13F funds and their CIKs
TRACKED_FUNDS = {
    "Berkshire Hathaway": "0001067983",
    "Baillie Gifford":    "0001048268",
    "Tiger Global":       "0001167483",
    "Coatue Management":  "0001336528",
    "Whale Rock Capital": "0001485922",
    "Lone Pine Capital":  "0001061165",
}

# Known foreign ADRs that don't file with EDGAR (no 8-K; covered by home-market disclosure)
KNOWN_FOREIGN_NO_EDGAR = {"SBGSY", "TOELY", "BESIY"}

_last_req = 0.0


def _sec_get(url: str, host_data: bool = False) -> requests.Response:
    global _last_req
    now = time.monotonic()
    if now - _last_req < 0.11:
        time.sleep(0.11 - (now - _last_req))
    headers = SEC_HEADERS_DATA if host_data else SEC_HEADERS_WWW
    r = requests.get(url, headers=headers, timeout=30)
    _last_req = time.monotonic()
    return r


def load_cik_map() -> dict[str, str]:
    cache = ROOT / ".cache" / "company_tickers.json"
    if cache.exists():
        return json.loads(cache.read_text())
    resp = _sec_get("https://www.sec.gov/files/company_tickers.json")
    raw = resp.json()
    mapping = {v["ticker"].upper(): str(v["cik_str"]).zfill(10) for v in raw.values()}
    cache.parent.mkdir(parents=True, exist_ok=True)
    cache.write_text(json.dumps(mapping))
    return mapping


def load_watchlist() -> list[dict]:
    """Return list of {ticker, company, layer, tier, score} dicts."""
    wb = load_workbook(SCORING_PATH, data_only=True)
    ws = wb["Watchlist"]
    rows = []
    for r in range(2, ws.max_row + 1):
        tkr = ws.cell(row=r, column=1).value
        if not tkr:
            break
        company  = ws.cell(row=r, column=2).value or ""
        layer    = ws.cell(row=r, column=3).value or ""
        score    = ws.cell(row=r, column=35).value  # TOTAL col
        tier     = ws.cell(row=r, column=36).value  # Tier col
        rows.append(dict(ticker=str(tkr).strip().upper(),
                         company=str(company).strip(),
                         layer=str(layer).strip(),
                         score=score,
                         tier=str(tier).strip() if tier else ""))
    return rows


def fetch_recent_filings(cik: str, forms_wanted: set[str]) -> list[dict]:
    """Return filings for the given forms within SCAN_FROM..SCAN_TO."""
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    r = _sec_get(url, host_data=True)
    if r.status_code != 200:
        return []
    data = r.json()
    recent = data.get("filings", {}).get("recent", {})
    form_list  = recent.get("form", [])
    acc_list   = recent.get("accessionNumber", [])
    date_list  = recent.get("filingDate", [])
    prim_list  = recent.get("primaryDocument", [])
    items_list = recent.get("items", [])

    results = []
    for form, acc, fdate, prim, items in zip(
            form_list, acc_list, date_list, prim_list, items_list):
        if form not in forms_wanted:
            continue
        fd = date.fromisoformat(fdate)
        if fd < SCAN_FROM or fd > SCAN_TO:
            continue
        results.append(dict(form=form, accession=acc, date=fdate,
                            primary=prim, items=str(items or "")))
    return results


def check_13f_funds(cik_map: dict[str, str]) -> list[dict]:
    hits = []
    for fund, cik in TRACKED_FUNDS.items():
        filings = fetch_recent_filings(cik, {"13F-HR", "13F-HR/A"})
        for f in filings:
            hits.append(dict(fund=fund, **f))
    return hits


def main():
    print("Loading watchlist...")
    watchlist = load_watchlist()
    print(f"  {len(watchlist)} tickers loaded.")

    print("Loading CIK map...")
    cik_map = load_cik_map()
    print(f"  {len(cik_map)} tickers in CIK map.")

    material = []
    routine  = []
    earnings_8k = []
    no_cik   = []

    total = len(watchlist)
    for i, row in enumerate(watchlist, 1):
        tkr = row["ticker"]
        if tkr in KNOWN_FOREIGN_NO_EDGAR:
            no_cik.append(tkr)
            print(f"  [{i:>3}/{total}] {tkr:<6} — foreign ADR, no EDGAR (known)")
            continue
        cik = cik_map.get(tkr)
        if not cik:
            no_cik.append(tkr)
            print(f"  [{i:>3}/{total}] {tkr:<6} — [FLAG] no CIK")
            continue

        filings = fetch_recent_filings(cik, {"8-K", "8-K/A", "6-K", "6-K/A"})
        if not filings:
            print(f"  [{i:>3}/{total}] {tkr:<6} — no filings this week")
            continue

        for f in filings:
            print(f"  [{i:>3}/{total}] {tkr:<6} — {f['form']} {f['date']} items={f['items']!r}")
            entry = dict(ticker=tkr, company=row["company"], tier=row["tier"],
                         score=row["score"], layer=row["layer"], **f)
            # Earnings 8-Ks (Item 2.02)
            if "2.02" in f["items"]:
                earnings_8k.append(entry)

            # Classify as material vs routine based on items
            items_str = f["items"]
            material_items = {
                "1.01",  # Agreements
                "1.02",  # Termination
                "1.03",  # Bankruptcy
                "2.01",  # Asset sale/acquisition
                "2.02",  # Earnings
                "2.05",  # Departure of directors
                "2.06",  # Material impairment
                "3.01",  # Delisting
                "4.01",  # Auditor change
                "4.02",  # Restatement
                "5.01",  # Control change
                "5.02",  # Leadership departure/appointment (CEO, CFO, CTO)
                "7.01",  # Reg-FD (often used for major announcements)
                "8.01",  # Other events (large financings, spin-offs)
            }
            # 7.01 and 8.01 can be routine (conference calls) -- we collect all and sort later
            item_tags = {it.strip() for it in items_str.split(",")} if items_str else set()
            if item_tags & material_items:
                material.append(entry)
            else:
                routine.append(entry)

    print("\nChecking 13F filings from tracked funds...")
    fund_13f = check_13f_funds(cik_map)
    print(f"  Found {len(fund_13f)} 13F filings.")

    # Dump raw results as JSON for the next stage
    out = {
        "scan_from": SCAN_FROM.isoformat(),
        "scan_to": SCAN_TO.isoformat(),
        "material": material,
        "routine": routine,
        "earnings_8k": earnings_8k,
        "no_cik": no_cik,
        "fund_13f": fund_13f,
    }
    out_path = ROOT / ".cache" / "weekly_scan_raw.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2, default=str))
    print(f"\nRaw results written to {out_path.relative_to(ROOT)}")
    print(f"Material: {len(material)} | Routine: {len(routine)} | "
          f"Earnings 8-Ks: {len(earnings_8k)} | No CIK: {len(no_cik)}")


if __name__ == "__main__":
    main()
