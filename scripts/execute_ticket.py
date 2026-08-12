"""Deterministic ticket executor — §C of the agentic-execution spec (2026-08-09).

Dom-run, no LLM anywhere. Reads a ticket, validates (C2 gates — ANY failure
refuses the WHOLE ticket), transmits marketable-limit orders over the
Robinhood MCP HTTP endpoint, writes a receipt. Default is dry-run: print the
order table and every validation result, send nothing. `--confirm` sends.

This file is the ONLY order writer in the repo (CLAUDE.md rule 29): the MCP
order tool name appears nowhere else, and Claude never calls it. Caps live in
tracking/live/executor-config.json (C3) — a deliberate edit, not a CLI flag.

Usage:
  python3 scripts/execute_ticket.py --ticket tracking/live/tickets/T.json
  python3 scripts/execute_ticket.py --ticket ... --confirm
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import subprocess
import sys
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
LIVE_DIR = _REPO_ROOT / 'tracking' / 'live'
MCP_URL = 'https://agent.robinhood.com/mcp/trading'
QUOTE_TOL = 0.03          # C2.5: refuse limits >3% from the live quote
REQUIRED_CAPS = ('ACCOUNT_CAP', 'MAX_ORDER_NOTIONAL', 'MAX_TURNOVER_PCT')


# ---------------------------------------------------------------- validation

def _parse_iso(ts: str) -> dt.datetime:
    return dt.datetime.fromisoformat(ts.replace('Z', '+00:00'))


def _in_regular_hours(now_iso: str) -> bool:
    """9:30–16:00 ET, Mon–Fri (holidays not modeled — the broker's own
    rejection covers those)."""
    from zoneinfo import ZoneInfo
    t = _parse_iso(now_iso).astimezone(ZoneInfo('America/New_York'))
    return t.weekday() < 5 and (9, 30) <= (t.hour, t.minute) and t.hour < 16


def validate(ticket: dict, *, cfg: dict | None, roster: set[str],
             cash: float, equity: float, quotes: dict[str, float],
             receipt_path: Path, halt_path: Path, now: str,
             allow_full_turnover: bool = False) -> list[str]:
    """C2 gates. Returns ALL failures (empty list = ticket may transmit)."""
    fails: list[str] = []
    orders = ticket.get('orders', [])

    # C3: caps must exist before anything else is trusted.
    if cfg is None:
        return ['executor-config.json missing — caps are required (C3); '
                'create tracking/live/executor-config.json']
    for key in REQUIRED_CAPS:
        if key not in cfg:
            fails.append(f'{key} missing from executor-config.json (C3)')
    if fails:
        return fails

    # C2.1 — integrity / expiry / executed-once
    import trade_ticket as tt
    if tt.ticket_checksum(orders) != ticket.get('checksum'):
        fails.append('checksum mismatch — ticket was modified after generation')
    if _parse_iso(now) > _parse_iso(ticket['expires_at']):
        fails.append(f'ticket expired at {ticket["expires_at"]} '
                     f'(stale market context — regenerate)')
    if receipt_path.exists():
        # Executed-once guard — but a receipt where EVERY order died at
        # transmit (schema rejection, transport down) means nothing reached
        # the broker, so the re-run is legitimate (2026-08-11: all 10 orders
        # bounced on MCP param typing and the receipt then blocked the fix).
        # Any order without a transmit_error state may exist broker-side:
        # refuse — double-execution risk beats convenience.
        try:
            prior_orders = json.loads(receipt_path.read_text()).get('orders')
        except ValueError:
            prior_orders = None
        all_failed = bool(prior_orders) and all(
            o.get('state') == 'transmit_error' for o in prior_orders)
        if all_failed:
            print(f're-run allowed: prior receipt {receipt_path.name} shows '
                  f'all orders transmit_error (nothing was placed)')
        else:   # placed, partial, empty, or unreadable: fail closed
            fails.append(f'already executed — receipt exists: '
                         f'{receipt_path.name}')

    # Fractional orders transmit as market-type (no fractional limit orders
    # on Robinhood) and are accepted only in regular hours — refuse rather
    # than let the broker bounce them one by one.
    if (any(o['shares'] % 1 != 0 for o in orders)
            and not _in_regular_hours(now)):
        fails.append('fractional orders are market-type and accepted only '
                     'during regular market hours (9:30–16:00 ET Mon–Fri) — '
                     're-run while the market is open')

    # C2.4 — kill switch
    if halt_path.exists():
        why = halt_path.read_text().strip()
        print(f'TRADING HALT ACTIVE: {why}')
        fails.append(f'kill switch present ({halt_path.name}) — clear it '
                     f'deliberately after investigating')

    # C2.2 — roster allowlist
    for o in orders:
        if o['ticker'] not in roster:
            fails.append(f"{o['ticker']}: not in current Targets roster — "
                         f'executor cannot be pointed at arbitrary symbols')

    # C2.3 — notional / cash / turnover / account cap
    max_order = float(cfg['MAX_ORDER_NOTIONAL'])
    for o in orders:
        if o['notional_est'] > max_order + 1e-6:
            fails.append(f"{o['ticker']}: notional {o['notional_est']:.2f} > "
                         f'MAX_ORDER_NOTIONAL {max_order:.2f}')
    buys = sum(o['notional_est'] for o in orders if o['side'] == 'buy')
    if buys > cash + 1e-6:
        fails.append(f'total buy notional {buys:.2f} > available cash {cash:.2f}')
    if buys > 0 and equity > float(cfg['ACCOUNT_CAP']) + 1e-6:
        fails.append(f'account equity {equity:.2f} > ACCOUNT_CAP '
                     f"{float(cfg['ACCOUNT_CAP']):.2f} — refusing buys "
                     f'(this system may not manage more than the cap)')
    turnover = sum(o['notional_est'] for o in orders)
    max_turn = float(cfg['MAX_TURNOVER_PCT'])
    if equity > 0 and turnover / equity > max_turn and not allow_full_turnover:
        fails.append(f'turnover {turnover / equity:.0%} > cap {max_turn:.0%} '
                     f'— a full redeploy needs --allow-full-turnover')

    # C2.5 — live-price sanity
    for o in orders:
        q = quotes.get(o['ticker'])
        if q is None or q <= 0:
            fails.append(f"{o['ticker']}: no live quote — cannot sanity-check "
                         f'limit price')
        elif abs(o['limit_price'] - q) / q > QUOTE_TOL:
            fails.append(f"{o['ticker']}: limit {o['limit_price']} is "
                         f'{abs(o["limit_price"] - q) / q:.1%} from live quote '
                         f'{q} — stale ticket in a moving market, regenerate')
    return fails


# ---------------------------------------------------------------- execution

def run(ticket_path, *, live_dir: Path, roster: set[str], transport,
        confirm: bool = False, allow_full_turnover: bool = False,
        now: str | None = None) -> dict:
    live_dir = Path(live_dir)
    ticket = json.loads(Path(ticket_path).read_text())
    now = now or dt.datetime.now(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')

    cfg_path = live_dir / 'executor-config.json'
    cfg = json.loads(cfg_path.read_text()) if cfg_path.exists() else None
    receipt_path = (live_dir / 'receipts'
                    / f"receipt-{ticket['ticket_id']}.json")
    halt_path = live_dir / 'trading-halt.flag'

    pf = transport.portfolio()
    quotes = transport.quotes([o['ticker'] for o in ticket.get('orders', [])])

    print(f"ticket {ticket['ticket_id']}  "
          f"({len(ticket.get('orders', []))} orders, "
          f"equity at gen {ticket.get('account_equity_at_gen')})")
    for o in ticket.get('orders', []):
        print(f"  {o['side']:<5} {o['ticker']:<7} {o['shares']:>10.4f} sh "
              f"@ limit {o['limit_price']:>9.2f}  (~{o['notional_est']:.2f})")

    fails = validate(ticket, cfg=cfg, roster=roster, cash=pf['cash'],
                     equity=pf['equity'], quotes=quotes,
                     receipt_path=receipt_path, halt_path=halt_path, now=now,
                     allow_full_turnover=allow_full_turnover)
    for f in fails:
        print(f'  REFUSE: {f}')
    if fails:
        print('ticket REFUSED — nothing sent')
        return {'failures': fails, 'sent': False, 'receipt': None}
    print('all validation gates passed')
    if not confirm:
        print('dry run (no --confirm) — nothing sent')
        return {'failures': [], 'sent': False, 'receipt': None}

    results = []
    for o in ticket['orders']:
        o = {**o, 'ref_id': order_ref_id(ticket['ticket_id'], o)}
        try:
            r = transport.place_equity_order(o)
            results.append({**o, 'order_id': r.get('order_id'),
                            'state': r.get('state', 'submitted'),
                            **({'raw_response': r['raw_response']}
                               if 'raw_response' in r else {})})
        except Exception as e:   # record, never silently drop (rule 3)
            results.append({**o, 'order_id': None, 'state': 'transmit_error',
                            'error': str(e)})
            print(f"  TRANSMIT ERROR {o['ticker']}: {e}")
    receipt = {'ticket_id': ticket['ticket_id'], 'sent_at': now,
               'checksum': ticket['checksum'], 'orders': results}
    receipt_path.parent.mkdir(parents=True, exist_ok=True)
    if receipt_path.exists():   # all-transmit_error re-run: archive, don't erase
        stamp = now.replace(':', '').replace('-', '')[:15]
        receipt_path.rename(receipt_path.with_name(
            f'{receipt_path.stem}.superseded-{stamp}.json'))
    receipt_path.write_text(json.dumps(receipt, indent=2) + '\n')
    print(f'receipt written: {receipt_path}')
    return {'failures': [], 'sent': True, 'receipt': receipt_path}


def extract_order_ack(payload) -> dict:
    """Find the order acknowledgement in a place_equity_order response.

    The 2026-08-12 execution proved the ack is nested deeper than a top-level
    d['id'] (all 10 receipts recorded order_id=None while the orders filled).
    Walk the payload for the first dict that looks order-shaped — has a state
    plus an id/order_id — wherever the server nests it. Returns {} when
    nothing matches; the caller then preserves the raw response instead."""
    def walk(node):
        if isinstance(node, dict):
            if 'state' in node and ('id' in node or 'order_id' in node):
                return node
            for v in node.values():
                hit = walk(v)
                if hit is not None:
                    return hit
        elif isinstance(node, list):
            for v in node:
                hit = walk(v)
                if hit is not None:
                    return hit
        return None
    return walk(payload) or {}


def order_ref_id(ticket_id: str, order: dict) -> str:
    """Deterministic idempotency key: same ticket + same logical order ->
    same UUID, so a retry after a transport failure can never double-place
    (the upstream deduplicates by ref_id)."""
    import uuid
    return str(uuid.uuid5(uuid.NAMESPACE_URL,
                          f'ai-stocks:{ticket_id}:{order["ticker"]}:'
                          f'{order["side"]}:{order["shares"]:.4f}'))


def mcp_order_params(account_number: str, order: dict) -> dict:
    """Robinhood MCP place_equity_order arguments. Schema constraints
    (discovered 2026-08-11/12, both bounced the first live ticket):
    - quantity/limit_price are STRINGS; JSON numbers fail -32602.
    - Fractional shares exist ONLY as type=market + regular_hours; there is
      no fractional limit order. The ticket's limit price stays behind as
      the quote-sanity reference; whole-share orders keep real limits."""
    fractional = order['shares'] % 1 != 0
    params = {
        'account_number': account_number,
        'symbol': order['ticker'], 'side': order['side'],
        'quantity': f"{order['shares']:.4f}",
        'type': 'market' if fractional else 'limit',
        'time_in_force': 'gfd',
        'market_hours': 'regular_hours',
    }
    if not fractional:
        params['limit_price'] = f"{order['limit_price']:.2f}"
    if order.get('ref_id'):
        params['ref_id'] = order['ref_id']
    return params


# ------------------------------------------------- live transport (Phase 2)

class RobinhoodTransport:
    """Plain JSON-RPC-over-HTTP MCP client. stdlib only, no LLM.

    The order tool name below is the only occurrence in the repo — Claude
    never calls it (rule 29). Live transmission is a Phase-2 concern; Phase 1
    runs are dry-run + fixture-transport tests. Token comes from
    ROBINHOOD_MCP_TOKEN or the Claude Code OAuth store (best-effort; if
    neither yields one, the executor refuses with this message rather than
    guessing)."""

    def __init__(self, url: str = MCP_URL, token: str | None = None):
        self.url = url
        self.token = token or self._find_token()
        self._id = 0
        self._session_id: str | None = None
        self._account: str | None = None

    @staticmethod
    def _find_token() -> str:
        tok = os.environ.get('ROBINHOOD_MCP_TOKEN')
        if tok:
            return tok
        try:   # Claude Code stores MCP OAuth tokens in the macOS Keychain
            out = subprocess.run(
                ['security', 'find-generic-password', '-s',
                 'Claude Code-credentials', '-w'],
                capture_output=True, text=True, timeout=10)
            if out.returncode == 0 and out.stdout.strip():
                creds = json.loads(out.stdout)
                for key, val in creds.get('mcpOAuth', {}).items():
                    if 'robinhood' in key.lower():
                        return val.get('accessToken', '')
        except Exception:
            pass
        raise SystemExit(
            'No Robinhood MCP token found. Set ROBINHOOD_MCP_TOKEN or ensure '
            'the Claude Code OAuth session exists (claude mcp list). '
            'Refusing to proceed without credentials.')

    def _post(self, message: dict) -> tuple[dict | None, dict]:
        """One JSON-RPC POST. Handles both application/json and SSE response
        bodies (streamable-HTTP MCP servers answer POSTs with a short SSE
        stream). Returns (parsed JSON-RPC payload or None, response headers)."""
        import urllib.request
        headers = {'Content-Type': 'application/json',
                   'Accept': 'application/json, text/event-stream',
                   'Authorization': f'Bearer {self.token}'}
        if self._session_id:
            headers['Mcp-Session-Id'] = self._session_id
        req = urllib.request.Request(self.url, data=json.dumps(message).encode(),
                                     headers=headers)
        with urllib.request.urlopen(req, timeout=30) as resp:
            ctype = resp.headers.get('Content-Type', '')
            raw = resp.read().decode()
            resp_headers = dict(resp.headers)
        if not raw.strip():
            return None, resp_headers
        if 'text/event-stream' in ctype:
            for line in raw.splitlines():        # last data: frame wins
                if line.startswith('data:'):
                    payload = json.loads(line[5:].strip())
            return payload, resp_headers
        return json.loads(raw), resp_headers

    def _ensure_session(self) -> None:
        """MCP streamable-HTTP handshake: initialize → session id →
        notifications/initialized."""
        if self._session_id is not None:
            return
        self._id += 1
        payload, headers = self._post({
            'jsonrpc': '2.0', 'id': self._id, 'method': 'initialize',
            'params': {'protocolVersion': '2025-03-26',
                       'capabilities': {},
                       'clientInfo': {'name': 'execute_ticket',
                                      'version': '1.0'}}})
        if payload and 'error' in payload:
            raise RuntimeError(f'MCP initialize failed: {payload["error"]}')
        self._session_id = headers.get('Mcp-Session-Id', '')
        self._post({'jsonrpc': '2.0',
                    'method': 'notifications/initialized', 'params': {}})

    def _call(self, tool: str, args: dict) -> dict:
        self._ensure_session()
        self._id += 1
        payload, _ = self._post({'jsonrpc': '2.0', 'id': self._id,
                                 'method': 'tools/call',
                                 'params': {'name': tool, 'arguments': args}})
        if payload is None:
            raise RuntimeError(f'{tool}: empty MCP response')
        if 'error' in payload:
            raise RuntimeError(payload['error'])
        content = payload['result'].get('content', [])
        text = next((c['text'] for c in content if c.get('type') == 'text'), '{}')
        try:
            return json.loads(text)
        except ValueError:
            # Surface WHAT came back — the 2026-08-12 re-run failed with a
            # bare 'Expecting value: char 0' that hid the server's response.
            raise RuntimeError(
                f'{tool}: unparseable result text {text[:300]!r} '
                f'(full payload keys: {sorted(payload["result"])})')

    def account_number(self) -> str:
        if self._account is None:
            accts = self._call('get_accounts', {})['data']['accounts']
            agentic = [a for a in accts if a.get('agentic_allowed')]
            if len(agentic) != 1:
                raise SystemExit(f'expected exactly 1 agentic account, '
                                 f'found {len(agentic)} — refusing')
            self._account = agentic[0]['account_number']
        return self._account

    def portfolio(self) -> dict:
        d = self._call('get_portfolio',
                       {'account_number': self.account_number()})['data']
        return {'cash': float(d['cash']), 'equity': float(d['total_value'])}

    def quotes(self, tickers: list[str]) -> dict[str, float]:
        if not tickers:
            return {}
        d = self._call('get_equity_quotes', {'symbols': tickers})['data']
        out = {}
        for r in d.get('results', []):
            q = r.get('quote', {})
            px = q.get('last_trade_price') or q.get('last_non_reg_trade_price')
            if q.get('symbol') and px:
                out[q['symbol']] = float(px)
        return out

    def positions(self) -> dict[str, float]:
        """READ: {ticker: shares} for the agentic account."""
        d = self._call('get_equity_positions',
                       {'account_number': self.account_number()})['data']
        return {p['symbol']: float(p['quantity'])
                for p in d.get('positions', []) if p.get('symbol')}

    def orders(self) -> list[dict]:
        """READ: recent order states, for reconciliation fill-verification."""
        d = self._call('get_equity_orders',
                       {'account_number': self.account_number()})['data']
        return [{'order_id': o.get('id') or o.get('order_id'),
                 'state': o.get('state')} for o in d.get('orders', [])]

    def place_equity_order(self, order: dict) -> dict:
        r = self._call('place_equity_order',     # THE only order-tool call site
                       mcp_order_params(self.account_number(), order))
        ack = extract_order_ack(r)
        out = {'order_id': ack.get('id') or ack.get('order_id'),
               'state': ack.get('state', 'submitted')}
        if not ack:   # unknown shape: keep the evidence in the receipt
            out['raw_response'] = r
        return out


def _load_roster() -> set[str]:
    """Targets allowlist: Include=Y rows of the committed workbook (C2.2)."""
    from openpyxl import load_workbook   # lazy (minimal-CI convention)
    wb = load_workbook(_REPO_ROOT / '00-master' / 'portfolio.xlsx',
                       data_only=False)
    ws = wb['Targets']
    hdr = [c.value for c in ws[2]]
    col = {name: i for i, name in enumerate(hdr)}
    roster = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[0] != 'TOTAL' and row[col['Include?']] == 'Y':
            roster.add(row[0])
    return roster


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='Validate + transmit a trade ticket')
    ap.add_argument('--ticket', required=True)
    ap.add_argument('--confirm', action='store_true',
                    help='actually transmit (default: dry-run, send nothing)')
    ap.add_argument('--allow-full-turnover', action='store_true')
    args = ap.parse_args()
    res = run(args.ticket, live_dir=LIVE_DIR, roster=_load_roster(),
              transport=RobinhoodTransport(), confirm=args.confirm,
              allow_full_turnover=args.allow_full_turnover)
    sys.exit(0 if not res['failures'] else 1)
