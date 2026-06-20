"""One-off: add 'Balance sheet (quarterly)' + formula-driven 'FY Summary' to
AVGO financials.xlsx. Raw line-items reference the statement sheets via
cross-sheet formulas; every ratio is an Excel formula (CLAUDE.md rule 4).
GAAP basis throughout (deep-dive pitfall: use GAAP, not non-GAAP, here).
"""
from pathlib import Path
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

P = Path("per-stock/AVGO/financials.xlsx")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)

wb = openpyxl.load_workbook(P)


def row_of(ws, label):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    raise KeyError(f"{label!r} not in {ws.title}")


# ---- add quarterly balance sheet (script doesn't pull it; TTM needs MRQ) ----
if "Balance sheet (quarterly)" not in wb.sheetnames:
    qb = yf.Ticker("AVGO").quarterly_balance_sheet
    ws = wb.create_sheet("Balance sheet (quarterly)")
    ws.append(["Line item"] + [str(c.date()) for c in qb.columns])
    for cell in ws[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for idx, row in qb.iterrows():
        ws.append([str(idx)] + [None if (v != v or v is None) else float(v)
                                for v in row.tolist()])
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            ws.cell(row=r, column=c).number_format = "#,##0"
    ws.column_dimensions["A"].width = 30
    for c in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

ai, qi = wb["Income stmt (annual)"], wb["Income stmt (quarterly)"]
ca, cq = wb["Cash flow (annual)"], wb["Cash flow (quarterly)"]
ba, qbs = wb["Balance sheet (annual)"], wb["Balance sheet (quarterly)"]

# annual cols: B=FY2025, C=FY2024, D=FY2023  -> FY-3=D, FY-2=C, FY-1=B
# quarterly flow cols B,C,D,E = last 4 quarters -> TTM = SUM(B:E)
AI = "'Income stmt (annual)'"
QI = "'Income stmt (quarterly)'"
CA = "'Cash flow (annual)'"
CQ = "'Cash flow (quarterly)'"
BA = "'Balance sheet (annual)'"
QB = "'Balance sheet (quarterly)'"

r_rev_a, r_rev_q = row_of(ai, "Total Revenue"), row_of(qi, "Total Revenue")
r_gp_a, r_gp_q = row_of(ai, "Gross Profit"), row_of(qi, "Gross Profit")
r_oi_a, r_oi_q = row_of(ai, "Operating Income"), row_of(qi, "Operating Income")
r_eb_a, r_eb_q = row_of(ai, "EBITDA"), row_of(qi, "EBITDA")
r_fcf_a, r_fcf_q = row_of(ca, "Free Cash Flow"), row_of(cq, "Free Cash Flow")
r_cap_a, r_cap_q = row_of(ca, "Capital Expenditure"), row_of(cq, "Capital Expenditure")
r_nd_a, r_nd_q = row_of(ba, "Net Debt"), row_of(qbs, "Net Debt")
r_sh_a, r_sh_q = row_of(ba, "Ordinary Shares Number"), row_of(qbs, "Ordinary Shares Number")


def annual(sheet, row, col):  # col: 'D'=FY-3,'C'=FY-2,'B'=FY-1
    return f"={sheet}!{col}{row}"


def ttm_sum(sheet, row):
    return f"=SUM({sheet}!B{row}:E{row})"


# ---- FY Summary sheet (insert right after Summary) ----
if "FY Summary" in wb.sheetnames:
    del wb["FY Summary"]
fy = wb.create_sheet("FY Summary", 1)
fy.append(["AVGO — fiscal-year summary (GAAP, $ except where noted)",
           "FY-3", "FY-2", "FY-1", "TTM", "Source / formula"])
for cell in fy[1]:
    cell.fill, cell.font = HEADER_FILL, HEADER_FONT
fy.append(["Fiscal period end", "2023-10-29", "2024-11-03", "2025-11-02",
           "Q2 FY26 (2026-05-03)", "AVGO 10-K FY2025; 10-Q Q2 FY26 (filed 2026-06-09)"])

# B=FY-3(2023→annual col D), C=FY-2(2024→C), D=FY-1(2025→B), E=TTM
def line(label, a_sheet, a_row, q_sheet, q_row, src, neg=False):
    sign = "-" if neg else ""
    fy.append([
        label,
        f"={sign}{a_sheet}!D{a_row}",
        f"={sign}{a_sheet}!C{a_row}",
        f"={sign}{a_sheet}!B{a_row}",
        (f"=-SUM({q_sheet}!B{q_row}:E{q_row})" if neg else ttm_sum(q_sheet, q_row)),
        src,
    ])

line("Revenue", AI, r_rev_a, QI, r_rev_q, "Total Revenue (annual sheet; TTM=Σ last 4 qtrs)")
line("Gross profit", AI, r_gp_a, QI, r_gp_q, "Gross Profit (GAAP)")
fy.append(["Gross margin %", "=B4/B3", "=C4/C3", "=D4/D3", "=E4/E3",
           "GAAP = Gross profit / Revenue. Non-GAAP ~77% (excl. acq. intangible amort.) — see Data flags"])
line("Operating income", AI, r_oi_a, QI, r_oi_q, "Operating Income (yfinance line)")
fy.append(["Operating margin %", "=B6/B3", "=C6/C3", "=D6/D3", "=E6/E3", "= Operating income / Revenue"])
line("EBITDA", AI, r_eb_a, QI, r_eb_q, "EBITDA (yfinance)")
line("Free cash flow", CA, r_fcf_a, CQ, r_fcf_q, "FCF = OCF − capex (statement-based)")
fy.append(["FCF margin %", "=B9/B3", "=C9/C3", "=D9/D3", "=E9/E3", "= FCF / Revenue"])
line("Capex", CA, r_cap_a, CQ, r_cap_q, "Capital Expenditure (shown positive)", neg=True)
fy.append(["Capex / revenue %", "=B11/B3", "=C11/C3", "=D11/D3", "=E11/E3", "= Capex / Revenue"])
# Net debt: annual stock from annual BS; TTM = MRQ (quarterly BS col B)
fy.append(["Net debt", f"={BA}!D{r_nd_a}", f"={BA}!C{r_nd_a}", f"={BA}!B{r_nd_a}",
           f"={QB}!B{r_nd_q}", "Total debt − cash; TTM = MRQ (Q2 FY26)"])
fy.append(["Net debt / EBITDA", "=B13/B8", "=C13/C8", "=D13/D8", "=E13/E8", "= Net debt / EBITDA"])
fy.append(["Share count (period-end, M)", f"={BA}!D{r_sh_a}/1e6", f"={BA}!C{r_sh_a}/1e6",
           f"={BA}!B{r_sh_a}/1e6", f"={QB}!B{r_sh_q}/1e6",
           "Ordinary Shares Number ÷ 1e6; TTM = MRQ"])
fy.append(["Share count YoY %", "", "=C15/B15-1", "=D15/C15-1", "=E15/D15-1", "vs prior period"])

# number formats: 5/7/10/12=margins(%), 16=share YoY(%), 14=net debt/EBITDA(x), 15=shares(M)
pct_rows = {5, 7, 10, 12, 16}
for r in range(3, fy.max_row + 1):
    for c in range(2, 6):
        cell = fy.cell(row=r, column=c)
        cell.number_format = "0.0%" if r in pct_rows else "#,##0"
for c in range(2, 6):
    fy.cell(row=14, column=c).number_format = '0.00"x"'   # net debt / EBITDA
    fy.cell(row=15, column=c).number_format = "#,##0.0"    # share count (M)
fy.column_dimensions["A"].width = 28
for c in "BCDE":
    fy.column_dimensions[c].width = 15
fy.column_dimensions["F"].width = 60
for c in range(1, 7):
    fy.cell(row=2, column=c).font = Font(italic=True)

# ---- Data flag: GAAP vs non-GAAP gross margin ----
fl = wb["Data flags"]
if fl.cell(row=2, column=1).value == "—":
    fl.delete_rows(2)
fl.append(["Gross margin basis",
           "FY Summary uses GAAP gross margin (~68%). Non-GAAP ~77% (Q2 FY26 77.1%) "
           "excludes amortization of acquired (VMware) intangibles in COGS. "
           "Watchlist 'Gross Mgn %' 76.3% is the non-GAAP/info.grossMargins basis."])
fl.append(["Operating income line",
           "Uses yfinance 'Operating Income'; 'Total Operating Income As Reported' "
           "is ~$0.6B lower (FY25 25,484 vs 26,075) — item classification."])

wb.save(P)
print("OK — added 'Balance sheet (quarterly)' + 'FY Summary' to", P)
