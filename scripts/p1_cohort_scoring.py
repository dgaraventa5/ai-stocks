"""P1 cohort-relative scoring — rank each name against its layer peers for the
six style-biased metrics, falling back to the live absolute bands when a peer
group is too small to rank meaningfully.

READ-ONLY: this computes a before (live absolute) vs after (cohort-relative)
comparison. It does NOT modify the workbook, recalc_watchlist, or any live
score. Turning P1 on is a separate, reviewed step — the tier bands must be
recalibrated to the new distribution first (critique §5-4).

Hybrid (§5-2): only these six metrics go cohort-relative —
  Value:   EV/EBITDA (col F), FCF Yield, P/S
  Quality: ROIC, Gross Margin, FCF Margin
Everything else (Net Debt/EBITDA, Fwd P/E, PEG, Growth, and all subjective
dims) stays exactly as the live model scores it.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

import recalc_watchlist as rc          # live band functions + weights + avg
import cohort_percentile

_REPO = Path(__file__).resolve().parent.parent
XLSX = str(_REPO / "00-master" / "ai_supply_chain_scoring.xlsx")

MIN_COHORT_SIZE = 8
# Live portfolio membership thresholds (refresh_targets.py) — for the roster diff.
ENTRY_SCORE = 74.5
EXIT_SCORE = 73.0


def top_level_layer(layer_str):
    """'06 AI Compute Silicon / GPUs' -> '06'. None/blank/no-code -> None."""
    if not layer_str:
        return None
    m = re.match(r"\s*(\d{2})", str(layer_str))
    return m.group(1) if m else None


def tier(t):
    """Same tier scale as the live model (recalc_watchlist)."""
    if t is None:
        return None
    if t >= 85:
        return "✓✓✓"
    if t >= 70:
        return "✓✓"
    if t >= 55:
        return "✓"
    if t >= 40:
        return "?"
    return "✗"


def cohort_metric_scores(values, higher_is_better, absolute_fn,
                         min_size=MIN_COHORT_SIZE):
    """Percentile-rank `values` within their cohort when >= min_size are
    non-null; otherwise score each via `absolute_fn` (the live band). Blanks
    stay blank in both cases."""
    non_null = [v for v in values if v is not None]
    if len(non_null) >= min_size:
        return cohort_percentile.rank_cohort(values, higher_is_better)
    return [absolute_fn(v) if v is not None else None for v in values]


def _num(v):
    return rc._num(v)


# --- workbook read -------------------------------------------------------

def _read_rows(xlsx=XLSX):
    wb = load_workbook(xlsx, data_only=False)
    ws = wb["Watchlist"]
    rows = []
    for r in range(2, ws.max_row + 1):
        tk = ws.cell(r, 1).value
        if not tk:
            continue
        layer = ws.cell(r, 3).value
        E = _num(ws.cell(r, 5).value)       # fwd P/E
        R = _num(ws.cell(r, 18).value)      # eps yoy (for PEG)
        peg = (E / R) if (E is not None and R is not None and R > 0 and E >= 0) else None
        rows.append(dict(
            ticker=tk, layer=layer, tl=top_level_layer(layer),
            fwd_pe=ws.cell(r, 5).value, ev=ws.cell(r, 6).value,
            fcf_yield=ws.cell(r, 7).value, ps=ws.cell(r, 8).value, peg=peg,
            roic=ws.cell(r, 11).value, gm=ws.cell(r, 12).value,
            fcf_mgn=ws.cell(r, 13).value, nd=ws.cell(r, 14).value,
            rev_cagr=ws.cell(r, 16).value, rev_yoy=ws.cell(r, 17).value,
            eps_yoy=ws.cell(r, 18).value,
            ai=[ws.cell(r, c).value for c in (20, 21, 22, 23, 24)],
            mom=[ws.cell(r, c).value for c in (26, 27, 28)],
            dma=ws.cell(r, 29).value,
            risk=[ws.cell(r, c).value for c in (31, 32, 33, 34, 38)],
        ))
    return rows


# --- subscore assembly (shared by both modes) ----------------------------

def _subj_avg(inputs):
    xs = [x for x in inputs if x is not None]
    return (sum(xs) / len(xs) * 20) if xs else None


def _assemble(row, mscores):
    """Given a row and its six metric SCORES (mscores dict: ev, fcf_yield, ps,
    roic, gm, fcf_mgn), build all category subscores + TOTAL. The non-cohort
    metrics use the live band functions directly."""
    value = rc.avg_nonnull([
        rc.score_fwd_pe(row["fwd_pe"]), mscores["ev"],
        mscores["fcf_yield"], mscores["ps"], rc.score_peg(row["peg"])])
    quality = rc.avg_nonnull([
        mscores["roic"], mscores["gm"], mscores["fcf_mgn"],
        rc.score_nd_ebitda(row["nd"])])
    growth = rc.avg_nonnull([
        rc.score_rev_cagr(row["rev_cagr"]), rc.score_rev_yoy(row["rev_yoy"]),
        rc.score_eps_yoy(row["eps_yoy"])])
    ai = _subj_avg(row["ai"])
    momentum = rc.avg_nonnull(
        [x * 20 if x is not None else None for x in row["mom"]]
        + [rc.score_50dma(row["dma"])])
    risk = _subj_avg(row["risk"])
    w = rc_weights()
    parts = []
    for sub, key in [(value, "Value"), (quality, "Quality"), (growth, "Growth"),
                     (ai, "AI"), (momentum, "Momentum"), (risk, "Risk")]:
        if sub is not None:
            parts.append(sub * w[key])
    total = sum(parts) if parts else None
    return dict(Value=value, Quality=quality, Growth=growth, AI=ai,
                Momentum=momentum, Risk=risk, TOTAL=total, Tier=tier(total))


_WEIGHTS_CACHE = {}


def rc_weights(xlsx=XLSX):
    if not _WEIGHTS_CACHE:
        wb = load_workbook(xlsx, data_only=False)
        weights = {}
        if "Weights" in wb.sheetnames:
            for r in wb["Weights"].iter_rows(min_row=2, values_only=True):
                if r[0] and r[1] is not None:
                    weights[r[0]] = r[1]
        _WEIGHTS_CACHE.update({
            "Value": weights.get("Value", 0.20),
            "Quality": weights.get("Quality", 0.20),
            "Growth": weights.get("Growth", 0.15),
            "AI": weights.get("AI Thesis", 0.20),
            "Momentum": weights.get("Momentum", 0.10),
            "Risk": weights.get("Risk", 0.15)})
    return _WEIGHTS_CACHE


# --- the two scoring modes ------------------------------------------------

def _absolute_mscores(row):
    return dict(
        ev=rc.score_ev_col(row["ev"], row["layer"]),
        fcf_yield=rc.score_fcf_yield(row["fcf_yield"]),
        ps=rc.score_ps(row["ps"]),
        roic=rc.score_roic(row["roic"]),
        gm=rc.score_gm(row["gm"]),
        fcf_mgn=rc.score_fcf_mgn(row["fcf_mgn"]))


def _percentile_mscores(rows):
    """Return {ticker: mscores} using cohort percentiles for the six metrics,
    per top-level-layer cohort, with the min-size fallback to absolute bands."""
    by_layer = {}
    for row in rows:
        by_layer.setdefault(row["tl"], []).append(row)

    out = {r["ticker"]: {} for r in rows}
    # Metrics uniform across every layer -> straightforward per-cohort percentile.
    UNIFORM = [
        ("fcf_yield", True, rc.score_fcf_yield),
        ("ps", False, rc.score_ps),
        ("roic", True, rc.score_roic),
        ("gm", True, rc.score_gm),
        ("fcf_mgn", True, rc.score_fcf_mgn),
    ]
    for tl, cohort in by_layer.items():
        for key, hib, absfn in UNIFORM:
            vals = [_num(r[key]) for r in cohort]
            scores = cohort_metric_scores(vals, hib, absfn)
            for r, s in zip(cohort, scores):
                out[r["ticker"]][key] = s

        # col F (EV/*) — meaning is uniform *within* a layer EXCEPT Layer 9,
        # where hyperscalers hold EV/EBITDA and capacity names hold EV/MW. Never
        # percentile a mixed column: Layer 9 stays absolute (per-row).
        vals = [_num(r["ev"]) for r in cohort]
        if tl == "09":
            ev_scores = [rc.score_ev_col(r["ev"], r["layer"]) for r in cohort]
        elif tl == "10":
            ev_scores = cohort_metric_scores(vals, False, rc.score_ev_fcf_saas)
        else:
            ev_scores = cohort_metric_scores(vals, False, rc.score_ev_ebitda)
        for r, s in zip(cohort, ev_scores):
            out[r["ticker"]]["ev"] = s
    return out


def score_watchlist(xlsx=XLSX):
    """Compute live-absolute and cohort-relative scores for every name.
    Returns a list of per-name dicts with both totals + tiers + subscores."""
    rows = _read_rows(xlsx)
    pct = _percentile_mscores(rows)
    results = []
    for row in rows:
        a = _assemble(row, _absolute_mscores(row))
        p = _assemble(row, pct[row["ticker"]])
        results.append(dict(
            ticker=row["ticker"], layer=row["layer"], tl=row["tl"],
            abs_total=a["TOTAL"], abs_tier=a["Tier"],
            pct_total=p["TOTAL"], pct_tier=p["Tier"],
            abs_value=a["Value"], pct_value=p["Value"],
            abs_quality=a["Quality"], pct_quality=p["Quality"]))
    return results
