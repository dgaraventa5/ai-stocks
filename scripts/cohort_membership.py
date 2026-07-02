"""Cohort-membership governance for P1 percentile scoring (CLAUDE.md rule 24).

P1 scores the six style-biased metrics as percentile ranks WITHIN a top-level-layer
cohort, so a name's score depends on who else is in its cohort. That makes cohort
membership score-determining — and it was previously only implicit in the Watchlist
Layer column (a binary .xlsx, not a reviewable diff). This module makes it governable:

  --update      regenerate 00-master/cohort-membership.md — the committed, reviewable
                map: each cohort's members + which metrics are percentile-eligible
                (n>=8) vs absolute-fallback, with the non-null count.
  (default)     --check: regenerate in memory, diff against the committed file, exit 1
                on drift. This is the GATE (also run as a test) — you cannot change
                cohort membership/eligibility without the committed map (and its diff)
                moving with it.
  --impact TKR  show how TKR's presence shifts every incumbent's percentile-metric
                scores in its cohort (add-a-weak-peer-inflates-the-strong-names,
                quantified) — run this BEFORE adding/removing a cohort member.

Workflow (rule 24): after any Watchlist cohort change (add/remove a name, edit a
Layer), run `--impact <ticker>` to see the effect on incumbents, then `--update`,
and commit the new cohort-membership.md in the same change. The git diff of that
file is the logged before/after; the drift test keeps it honest.
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

import cohort_percentile
from recalc_watchlist import _num

_ROOT = Path(__file__).resolve().parent.parent
XLSX = str(_ROOT / "00-master" / "ai_supply_chain_scoring.xlsx")
OUT = _ROOT / "00-master" / "cohort-membership.md"

# The six style-biased metrics that go cohort-relative, with their Watchlist column.
METRIC_COLS = [("ev_ebitda", 6), ("fcf_yield", 7), ("ps", 8),
               ("roic", 11), ("gm", 12), ("fcf_mgn", 13)]
MIN_SIZE = cohort_percentile.MIN_COHORT_SIZE


def _cohorts(ws):
    coh = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, 1).value
        if not t:
            continue
        tl = cohort_percentile.top_level_layer(ws.cell(r, 3).value)
        coh[tl].append((str(t), r))
    return coh


def _metric_values(ws, members, col):
    vals = []
    for _t, r in members:
        v = ws.cell(r, col).value
        if isinstance(v, str) and v.startswith("="):
            v = None
        vals.append(_num(v))
    return vals


def generate(xlsx=XLSX) -> str:
    """The cohort-membership map as deterministic markdown."""
    ws = load_workbook(xlsx, data_only=False)["Watchlist"]
    coh = _cohorts(ws)
    out = [
        "# Cohort membership — P1 percentile scoring (GENERATED, do not hand-edit)",
        "",
        "Source of truth: the Watchlist `Layer` column. Regenerate with",
        "`python3 scripts/cohort_membership.py --update`. A diff here means the way",
        "names are peer-ranked changed (rule 24). `pct` = percentile-ranked (cohort",
        f"has >= {MIN_SIZE} non-null); `abs` = absolute-band fallback. Layer 09 col-F",
        "(EV/*) is always `abs` (mixed EV/EBITDA + EV/MW).",
        "",
    ]
    for tl in sorted(coh):
        if tl is None:
            continue
        members = coh[tl]
        elig = []
        for name, col in METRIC_COLS:
            n = sum(1 for v in _metric_values(ws, members, col) if v is not None)
            pct = n >= MIN_SIZE and not (name == "ev_ebitda" and tl == "09")
            elig.append(f"{name}={'pct' if pct else 'abs'}({n})")
        out.append(f"## Layer {tl} — {len(members)} names")
        out.append("metrics: " + "  ".join(elig))
        out.append("members: " + ", ".join(sorted(t for t, _r in members)))
        out.append("")
    return "\n".join(out) + "\n"


def check(xlsx=XLSX) -> bool:
    """True if the committed map matches the live Watchlist-derived cohorts."""
    return OUT.exists() and OUT.read_text() == generate(xlsx)


def impact(ticker, xlsx=XLSX) -> str:
    """How `ticker`'s presence shifts each incumbent's percentile-metric scores in
    its cohort — the peer-inflation effect, quantified. Compares rank_cohort WITH
    vs WITHOUT the ticker, per metric, for cohorts big enough to be percentile-ranked."""
    ticker = ticker.upper()
    ws = load_workbook(xlsx, data_only=False)["Watchlist"]
    coh = _cohorts(ws)
    tl = None
    for k, members in coh.items():
        if any(t == ticker for t, _r in members):
            tl = k
            break
    if tl is None:
        return f"{ticker}: not on the Watchlist."
    members = coh[tl]
    lines = [f"Impact of {ticker} on Layer {tl} ({len(members)} names) — "
             f"incumbent percentile shifts if {ticker} were removed:"]
    hib = {"ev_ebitda": False, "fcf_yield": True, "ps": False,
           "roic": True, "gm": True, "fcf_mgn": True}
    for name, col in METRIC_COLS:
        vals = _metric_values(ws, members, col)
        idx = [i for i, (t, _r) in enumerate(members) if t == ticker][0]
        with_vals = vals
        without = vals[:idx] + [None] + vals[idx + 1:]  # drop ticker, keep positions
        n_with = sum(1 for v in with_vals if v is not None)
        if n_with < MIN_SIZE or (name == "ev_ebitda" and tl == "09"):
            continue  # absolute-fallback metric — no percentile to shift
        sw = cohort_percentile.rank_cohort(with_vals, hib[name])
        so = cohort_percentile.rank_cohort(without, hib[name])
        moved = [(members[i][0], round(so[i] - sw[i], 1))
                 for i in range(len(members))
                 if i != idx and sw[i] is not None and so[i] is not None
                 and abs(so[i] - sw[i]) >= 0.05]
        if moved:
            lines.append(f"  {name}: " + ", ".join(f"{t}{d:+.1f}" for t, d in
                                                    sorted(moved, key=lambda x: -abs(x[1]))))
    if len(lines) == 1:
        lines.append("  (no percentile-ranked metric affected — thin cohort on absolute fallback)")
    return "\n".join(lines)


def main(argv):
    if "--update" in argv:
        OUT.write_text(generate())
        print(f"wrote {OUT.relative_to(_ROOT)}")
    elif "--impact" in argv:
        i = argv.index("--impact")
        print(impact(argv[i + 1]))
    else:
        if check():
            print("cohort-membership.md is current ✓")
        else:
            print("DRIFT: cohort-membership.md is stale — run --update and review the diff")
            sys.exit(1)


if __name__ == "__main__":
    main(sys.argv)
