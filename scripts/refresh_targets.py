"""Regenerate portfolio.xlsx Targets + Reconciliation from live Watchlist scores.

This is the score→portfolio pipeline step that was missing: the Targets sheet
was hand-built 2026-05-18/25 and froze while live scores moved (the portfolio
README referenced this script before it existed). Built 2026-06-09.

Pipeline rules (parameters live on the Sizing Rules sheet — edit there):
  - Hysteresis: non-holdings ENTER at rank <= entry_rank; holdings HOLD until
    rank > exit_rank OR score < 70 (tier drop below ✓✓), and an exit must be
    confirmed on 2 consecutive runs (EXIT PENDING -> EXIT) so a single noisy
    refresh can't trade the portfolio. Entrants admitted in rank order only
    while position count < max_positions. The pending clock is persisted in
    tracking/performance-config.json (`exit_pending`, rule 26), NOT the Targets
    sheet: freeze-snapshot gating rewrites the sheet only when a rebalance
    fires, so a sheet-persisted clock could never start (2026-08-02 fix).
    Frozen runs save only the config; dry runs never mutate it.
  - Freshness gate: candidates whose last yfinance trade is >7 days old are
    treated as dead (halted/delisted), excluded, and flagged — a delisted
    name (CTRA, 2026-05) once sat in the live top-20 on frozen data.
  - Manual overrides: Targets rows whose Override col says EXCLUDE stay out
    regardless of rank (kept across refreshes, flagged when their rank would
    otherwise include them). Set INCLUDE to force a name in.
  - Sizing: tier-band linear interpolation on score (Sizing Rules table),
    normalized to (1 - cash buffer), then max-single-position cap, then layer
    cap (default 1.0 = off, per Dom's pure-score-driven preference — layer
    exposure is reported every run so that choice stays visible).

Per CLAUDE.md rule 5: the Reconciliation sheet is a mechanical target-vs-
current diff, not a trade recommendation. Decisions are Dom's.

Usage:
  python3 scripts/refresh_targets.py            # refresh + write
  python3 scripts/refresh_targets.py --dry-run  # print plan, write nothing
"""
from __future__ import annotations

import argparse
import datetime as dt
import time
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from common import flag
from portfolio_model import (current_weights, load_cfg, load_pcfg,
                             log_rebalance, save_cfg)
from portfolio_sizing import (build_reason, is_tradable, rank_by_score,
                              tier_changes, topn_membership,
                              weights_score_monotonic)
from position_sizing import drift_band_filter, inverse_vol_weights
from recalc_watchlist import recalc

_REPO_ROOT = Path(__file__).resolve().parent.parent
PORTFOLIO = str(_REPO_ROOT / '00-master/portfolio.xlsx')

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF')

# Methodology-seam damping (rule 32-C, added 2026-08-31). All three of the
# model's costliest exits (CRM 2026-06-10, ZS 2026-06-18, PLTR 2026-07-02 —
# +53%/+50%/+44% after exit vs QQQ ~flat) were triggered on methodology-deploy
# days (50DMA launch, threshold experiment, P1 go-live), not on company news,
# and the 2-run clock confirmed within ~a day. When a methodology change
# deploys, the deploying session stamps a seam (`--seam "reason"`); exit clocks
# STARTED inside the window cannot confirm until it closes, so the market gets
# SEAM_DAYS to disagree with the new methodology before the model trades on it.
# Deliberately narrow: clocks that predate the seam are data-driven and confirm
# on schedule; --resize and the dead/untradable branches bypass as before.
SEAM_DAYS = 7


def seam_blocks_confirm(clock_start: str, today: str, seam: dict | None,
                        seam_days: int = SEAM_DAYS) -> bool:
    """True while the seam window forbids confirming an exit whose pending
    clock started inside it (ISO-date strings; seam = {'date','reason'})."""
    if not seam or not seam.get('date'):
        return False
    start = dt.date.fromisoformat(seam['date'])
    end = start + dt.timedelta(days=seam_days)
    return (start <= dt.date.fromisoformat(clock_start) < end
            and dt.date.fromisoformat(today) < end)


def stamp_seam(reason: str) -> dict:
    """Record a methodology seam starting today in the performance config."""
    cfg = load_cfg()
    cfg['methodology_seam'] = {'date': dt.date.today().isoformat(),
                               'reason': reason}
    save_cfg(cfg)
    return cfg['methodology_seam']

# Defaults for parameters added 2026-06-09; written to Sizing Rules if absent
# so they are visible and editable in the workbook, not buried here.
# THE SHEET IS AUTHORITATIVE for live values — ensure_params only seeds missing
# rows, never overwrites. Keep these seeds in sync with the sheet: drift here
# misled CLAUDE.md rule 20 into quoting 74.5/73.0 while the sheet ran the
# 2026-06-18 76/74.5 concentration experiment (reverted 2026-08-02, Dom).
NEW_PARAMS = [
    # Membership is SCORE-threshold-anchored (2026-06-10, Dom): hold the genuine
    # ✓✓ cluster, let the count flex with scores rather than fix it. Hysteresis
    # band entry>=74.5 / exit<73.0 yields ~15 names today. (Rank params retained
    # below only as an inert backstop / historical reference.)
    ('Entry score', 74.5, 'Non-holding ENTERs when score >= this'),
    ('Exit score', 73.0, 'Holding EXITs when score < this (2-run confirm, or immediate on --resize)'),
    ('Exit confirm runs', 2, 'Consecutive refreshes below exit score before EXIT'),
    ('Max positions', 20, 'Hard ceiling backstop; threshold normally binds first'),
    ('Max stale price days', 7, 'Last trade older than this = dead, excluded'),
    ('Layer cap %', 1.0, 'Max weight per layer. 1.0 = off (Dom 2026-06-10: caps '
                         'manage label- not factor-concentration; use single-name '
                         'cap + capex tripwire instead).'),
    ('Entry rank', 15, 'selection.mode=rank: non-holding ENTERs at rank <= N '
                       '(v2 spec B1); inert under mode=score'),
    ('Exit rank', 18, 'selection.mode=rank: holding EXITs below rank M; '
                      'N+1..M is the hysteresis dead-band; inert under '
                      'mode=score'),
]


def load_params(ws) -> dict:
    """Read scalar params + tier band table from the Sizing Rules sheet."""
    params, tiers = {}, {}
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        if row[0] in ('Cash buffer %', 'Max single position %',
                      'Target position count', 'Portfolio Value ($)',
                      'Entry score', 'Exit score',
                      'Entry rank', 'Exit rank', 'Exit confirm runs',
                      'Max positions', 'Max stale price days', 'Layer cap %'):
            params[row[0]] = row[1]
        if row[0] == 'Tier' and row[1] == 'Score Floor':
            for trow in rows[i + 1:]:
                if trow[0] in ('✓✓✓', '✓✓', '✓', '?', '✗'):
                    tiers[trow[0]] = dict(floor=trow[1], w_floor=trow[2],
                                          w_ceil=trow[3], ceil=trow[4])
    return params | {'tiers': tiers}


def ensure_params(ws) -> None:
    """Append any missing pipeline params to Sizing Rules (idempotent)."""
    present = {row[0].value for row in ws.iter_rows()}
    for name, default, note in NEW_PARAMS:
        if name not in present:
            ws.append([name, default, note + ' (added 2026-06-09)'])


def base_weight(score: float, tiers: dict) -> float:
    """Linear interpolation of weight within the score's tier band."""
    for t in tiers.values():
        if t['floor'] <= score <= t['ceil']:
            span = t['ceil'] - t['floor']
            frac = (score - t['floor']) / span if span else 0
            return t['w_floor'] + frac * (t['w_ceil'] - t['w_floor'])
    return 0.0


def cap_and_normalize(weights: dict[str, float], layers: dict[str, str],
                      budget: float, max_single: float,
                      layer_cap: float) -> dict[str, float]:
    """Scale to budget, then iteratively enforce single + layer caps."""
    w = dict(weights)
    for _ in range(20):
        total = sum(w.values())
        if not total:
            return w
        w = {t: v / total * budget for t, v in w.items()}
        over = {t: max_single for t, v in w.items() if v > max_single + 1e-9}
        # Layer cap: shave members of an over-cap layer proportionally.
        by_layer: dict[str, float] = {}
        for t, v in w.items():
            by_layer[layers[t]] = by_layer.get(layers[t], 0) + v
        for lay, lv in by_layer.items():
            if lv > layer_cap * budget + 1e-9:
                scale = layer_cap * budget / lv
                for t in w:
                    if layers[t] == lay:
                        over[t] = min(over.get(t, w[t]), w[t] * scale)
        if not over:
            return w
        # Freeze capped names, renormalize the rest into the leftover budget.
        fixed = sum(over.values())
        rest = {t: v for t, v in w.items() if t not in over}
        rest_total = sum(rest.values())
        if not rest_total:
            return over
        w = over | {t: v / rest_total * (budget - fixed) for t, v in rest.items()}
    return w


def last_trade_age_days(ticker: str) -> int | None:
    import yfinance as yf  # lazy: the fire/dry-run path is offline (site-deploy CI has no yfinance)
    hist = yf.Ticker(ticker).history(period='1mo', auto_adjust=True)
    if hist is None or hist.empty:
        return None
    return (dt.date.today() - hist.index[-1].date()).days


def current_price(ticker: str) -> float | None:
    import yfinance as yf  # lazy (see last_trade_age_days)
    hist = yf.Ticker(ticker).history(period='5d', auto_adjust=False)
    if hist is None or hist.empty:
        return None
    return round(float(hist['Close'].iloc[-1]), 2)


def _price_frame(tickers: list[str], lookback: int):
    """Wide adjusted-close frame for the inverse-vol sizing window, via the
    tracker's cached _series (dividend-adjusted, serialized fetches). Lazy
    imports keep the module importable in the openpyxl-only deploy CI."""
    import pandas as pd

    from portfolio_model import _series
    earliest = (dt.date.today()
                - dt.timedelta(days=int(lookback * 2.2) + 10)).isoformat()
    cols = {}
    for t in tickers:
        s = _series(t, earliest)
        if s is not None:
            # Normalize tz-aware exchange timestamps to plain dates so mixed
            # calendars (Tokyo/US/EU listings) align instead of interleaving.
            s = s.copy()
            s.index = pd.Index([ts.date() for ts in s.index])
            cols[t] = s[~s.index.duplicated(keep='last')]
    return pd.DataFrame(cols)


def _update_band_shadows(cfg: dict, live: list, pcfg: dict, today: str) -> bool:
    """Refresh BAND_TOP/NEXT/TAIL rosters from today's ranks (v2 spec C1).
    Pure-score ranking (no incumbency tie-break): the bands are unsized
    scouts, not the book. Appends a shadow event only when a roster actually
    changed; same-day re-runs replace, prior days are never rewritten;
    forward-only (first event = first real run after deploy). Returns True
    when anything changed (caller persists cfg)."""
    sh = pcfg['shadows']
    ranked = [x['ticker'] for x in
              sorted(live, key=lambda x: (-x['TOTAL'], x['ticker']))]
    rosters = {'BAND_TOP': ranked[:sh['top']],
               'BAND_NEXT': ranked[sh['top']:sh['next']],
               'BAND_TAIL': ranked[sh['next']:sh['tail']]}
    events = cfg.setdefault('shadow_events', {})
    changed = False
    for name, roster in rosters.items():
        evs = events.setdefault(name, [])
        if not roster and not evs:
            continue          # watchlist smaller than the band: nothing to log
        if evs and sorted(evs[-1]['roster']) == sorted(roster):
            continue
        if evs and evs[-1]['date'] == today:
            evs[-1] = {'date': today, 'roster': roster}   # same-day idempotence
        else:
            evs.append({'date': today, 'roster': roster})
        changed = True
        flag(f'shadow {name}: roster refreshed ({len(roster)} names)')
    return changed


def refresh(dry_run: bool = False, resize: bool = False,
            portfolio: str | None = None, check_freshness: bool = True,
            migration: bool = False) -> None:
    today = dt.date.today().isoformat()
    path = portfolio or PORTFOLIO
    wb = load_workbook(path, data_only=False)
    sizing, targets = wb['Sizing Rules'], wb['Targets']
    # Reconciliation + Positions were removed in the notional-only privacy pass
    # (PR #9); both are optional. The site reads only Targets, and Positions is
    # unused (the model is the portfolio of record — portfolio_model.py). Tolerate
    # their absence so refresh runs on the slimmed workbook instead of crashing —
    # that KeyError crash is what forced the ad-hoc Targets hand-edits this change
    # replaces. Absent sheets are not recreated (Dom 2026-06-29: keep it slim).
    recon = wb['Reconciliation'] if 'Reconciliation' in wb.sheetnames else None
    positions = wb['Positions'] if 'Positions' in wb.sheetnames else None

    ensure_params(sizing)
    p = load_params(sizing)
    # Score-threshold membership (2026-06-10). resize=True drops below-exit names
    # immediately (intentional re-sizing), bypassing the 2-run exit hysteresis.
    entry_score = float(p.get('Entry score', 74.5))
    exit_score = float(p.get('Exit score', 73.0))
    confirm_runs = int(p.get('Exit confirm runs', 2))
    max_positions = int(p.get('Max positions', 20))
    stale_days = int(p.get('Max stale price days', 7))
    layer_cap = float(p.get('Layer cap %', 1.0))
    max_single = float(p['Max single position %'])
    cash = float(p['Cash buffer %'])
    capital = float(p['Portfolio Value ($)'])

    # ---- prior state: model config (pending clock) + Targets sheet ----
    cfg = load_cfg()
    # Exit-pending clock: ticker -> 'since' date, persisted in the config, NOT
    # the sheet — freeze-snapshot gating rewrites the Status column only when a
    # rebalance fires, so a sheet-persisted clock could never start (the
    # 2026-08-02 bug: GOOGL/META/AMZN restarted their clocks every run).
    exit_pending: dict[str, str] = dict(cfg.get('exit_pending', {}))
    hdr = [c.value for c in targets[2]]
    col = {name: i for i, name in enumerate(hdr)}
    prior_include: set[str] = set()
    overrides: dict[str, str] = {}
    notes: dict[str, str] = {}
    for row in targets.iter_rows(min_row=3, values_only=True):
        tkr = row[0]
        if not tkr or tkr == 'TOTAL':
            continue
        if row[col.get('Include?', 5)] == 'Y':
            prior_include.add(tkr)
        note = str(row[col.get('Notes', 8)] or '')
        if note:
            notes[tkr] = note
        ov = str(row[col['Override']] or '') if 'Override' in col else ''
        # Migration: the pre-2026-06-09 sheet encoded manual excludes in Notes.
        # Only applies when the Override column doesn't exist yet — once it
        # does, a blank Override means "no override", even if historical note
        # text ('Manually excluded ...') is retained for the audit trail.
        if 'Override' not in col and ('Manually excluded' in note
                                      or 'DELISTED' in note
                                      or 'investability override' in note):
            ov = 'EXCLUDE'
        if ov:
            overrides[tkr] = ov

    # ---- live scores ----
    live = [x for x in recalc() if x['TOTAL'] is not None]
    live.sort(key=lambda x: -x['TOTAL'])
    info = {x['ticker']: x for x in live}
    layers = {x['ticker']: (x['layer'] or '')[:2] for x in live}
    if not dry_run and portfolio is None:
        # Point-in-time score panel (v2 spec C2); date-deduped, so the
        # sync_scores hook and this one never double-log a pass. Gated to the
        # real workbook (portfolio=None) so test runs on temp workbooks never
        # touch the committed panel.
        from score_history import append_snapshot
        append_snapshot(live)

    # ---- tradability filter (spec 2026-08-11): selection ranks over the
    # buyable universe only; info/layers/score-panel stay full-universe.
    pcfg = load_pcfg()
    tradable_only = bool(pcfg['selection'].get('tradable_only'))
    sel_live = ([x for x in live if is_tradable(x['ticker'])]
                if tradable_only else live)
    untradable = ({t for t in info if not is_tradable(t)}
                  if tradable_only else set())
    for _t, _v in overrides.items():
        if _v == 'INCLUDE' and _t in untradable:
            flag(f'{_t}: Override=INCLUDE but untradable — filter wins, stays out')

    # ---- selection form (v2 spec B1): score thresholds or top-N ranks ----
    sel_mode = pcfg['selection']['mode']
    if sel_mode == 'rank':
        # Top-N with hysteresis: outsiders ENTER at rank <= N, incumbents HOLD
        # through rank <= M; N+1..M is the dead-band. The rule-26 2-run exit
        # confirm clock still wraps every exit crossing (continuity decision,
        # plan Task 5): below-M starts the clock, the next run confirms.
        entry_rank = int(p.get('Entry rank', 15))
        exit_rank = int(p.get('Exit rank', 18))
        ranked = rank_by_score(sel_live, prior_include)
        rank = {t: i + 1 for i, t in enumerate(ranked)}   # incumbency tie-break
        order = ranked
        keeps = lambda t: t in rank and rank[t] <= exit_rank
        enters = lambda t: rank[t] <= entry_rank
        why = lambda t: f'rank {rank[t]}, score {info[t]["TOTAL"]:.1f}'
        below = lambda t: f'rank {rank.get(t, "?")} > exit rank {exit_rank}'
        in_play = {t for t in rank if rank[t] <= exit_rank}
    else:
        rank = {x['ticker']: i + 1 for i, x in enumerate(sel_live)}
        order = [x['ticker'] for x in sel_live]
        keeps = lambda t: t in info and info[t]['TOTAL'] >= exit_score
        enters = lambda t: info[t]['TOTAL'] >= entry_score
        why = lambda t: f'score {info[t]["TOTAL"]:.1f}'
        below = lambda t: f'score {info[t]["TOTAL"]:.1f} < {exit_score}'
        in_play = {x['ticker'] for x in sel_live
                   if x['TOTAL'] >= exit_score}

    # ---- freshness gate on every name that could end up included ----
    candidates = sorted(
        in_play | prior_include
        | {t for t, v in overrides.items() if v == 'INCLUDE'})
    dead: set[str] = set()
    if check_freshness:   # network; skipped by the offline pending_rebalance() gate
        for t in candidates:
            age = last_trade_age_days(t)
            time.sleep(0.25)
            if age is None or age > stale_days:
                dead.add(t)
                flag(f'{t}: last trade {"none" if age is None else f"{age}d ago"} '
                     f'— excluded as halted/delisted')

    # ---- membership: hysteresis ----
    include: list[str] = []
    statuses: dict[str, str] = {}
    new_pending: dict[str, str] = {}    # next state of the exit-pending clock
    for t in prior_include:
        if t in dead or overrides.get(t) == 'EXCLUDE' or t in untradable:
            # Untradable is a constraint, not a signal — it can't mean-revert
            # next refresh, so it skips the 2-run confirm clock (CTRA precedent).
            statuses[t] = ('EXIT (untradable)' if t in untradable
                           else 'EXIT (dead/override)')
            if t in untradable:
                flag(f'{t}: held but untradable on US brokerages — EXIT')
            continue
        if keeps(t):
            include.append(t)
            statuses[t] = 'HOLD'
        elif resize:
            # Intentional re-sizing: drop below-threshold names immediately.
            statuses[t] = f'EXIT (resize, {below(t)})'
            flag(f'{t}: EXIT (resize — {below(t)})')
        elif (t in exit_pending and exit_pending[t] != today
              and seam_blocks_confirm(exit_pending[t], today,
                                      cfg.get('methodology_seam'))):
            # Rule 32-C: the clock started inside an open methodology-seam
            # window — hold the name and keep the ORIGINAL clock date until
            # the window closes, so a methodology deploy alone can't trade
            # the book before the market gets SEAM_DAYS to disagree.
            include.append(t)
            new_pending[t] = exit_pending[t]
            seam = cfg['methodology_seam']
            damp_end = (dt.date.fromisoformat(seam['date'])
                        + dt.timedelta(days=SEAM_DAYS)).isoformat()
            statuses[t] = (f'EXIT PENDING (seam-damped until {damp_end}, '
                           f'{below(t)})')
            flag(f'{t}: {below(t)} — confirm seam-damped until {damp_end} '
                 f'({seam.get("reason", "methodology change")})')
        elif t in exit_pending and exit_pending[t] != today:
            # EXIT PENDING on a PRIOR refresh date → confirmed now. Same-day
            # re-runs don't count as the second look — the confirm exists to
            # require two independent data points, not two script invocations.
            statuses[t] = f'EXIT (pending since {exit_pending[t]}, confirmed)'
            flag(f'{t}: exit confirmed ({below(t)}, '
                 f'pending since {exit_pending[t]})')
        else:
            include.append(t)   # still held while pending
            new_pending[t] = exit_pending.get(t, today)   # start/keep the clock
            statuses[t] = f'EXIT PENDING ({below(t)}, since {new_pending[t]})'
            flag(f'{t}: {below(t)} — EXIT PENDING, confirms next refresh')
    for t in order:
        if t in include or t in statuses or t in dead or t in untradable:
            continue
        if overrides.get(t) == 'EXCLUDE':
            if enters(t):
                flag(f'{t}: would enter ({why(t)}) but Override=EXCLUDE '
                     f'({notes.get(t, "no note")[:80]})')
            continue
        forced = overrides.get(t) == 'INCLUDE'
        if enters(t) or forced:
            if len(include) < max_positions:
                include.append(t)
                statuses[t] = ('ENTER (forced)' if forced
                               else f'ENTER ({why(t)})')
                flag(f'{t}: ENTER ({why(t)})')
            else:
                statuses[t] = f'BLOCKED ({why(t)}, max positions full)'
                flag(f'{t}: qualifies ({why(t)}) but max positions '
                     f'({max_positions}) full')

    # ---- sizing (v2 spec A1): tier bands or inverse trailing volatility ----
    siz = pcfg['sizing']
    if siz['mode'] == 'inverse_vol':
        if dry_run and not check_freshness:
            # Offline probe (pending_rebalance / rule-25 gate): the gate reads
            # only membership + tier state, never weights — keep it networkless
            # with an equal-weight placeholder rather than fetching prices.
            weights = {t: (1.0 - cash) / len(include) for t in include} \
                if include else {}
        else:
            # The score chooses the names; trailing vol chooses the sizes.
            prices = _price_frame(include, int(siz['lookback']))
            inv = inverse_vol_weights(prices, include, siz,
                                      layers={t: layers.get(t, '')
                                              for t in include})
            weights = {t: w * (1.0 - cash) for t, w in inv.items()}
    else:
        base = {t: base_weight(info[t]['TOTAL'], p['tiers']) for t in include}
        weights = cap_and_normalize(base, layers, 1.0 - cash, max_single,
                                    layer_cap)

    # ---- rebalance gate: membership change OR a held name crossed a tier band ----
    # (--resize forces it). Crossing baseline is the LAST MODEL EVENT's stored
    # tiers, NOT the Targets sheet: the sheet can carry a fresh tier from an
    # out-of-band score edit (the bug this fixes), so it is not a trustworthy
    # baseline. Within-tier drift changes nothing — hysteresis preserved.
    # Under inverse-vol sizing a tier crossing carries no weight information
    # (v2 spec A0: stop using the score as sizer), so it does not fire.
    last_ev = cfg['events'][-1]
    prior_model = set(last_ev['allocations'])
    last_tiers = last_ev.get('tiers', {})
    tier_chg = tier_changes(include, info, last_tiers)
    tier_fire = bool(tier_chg) and siz['mode'] != 'inverse_vol'
    entered = sorted(set(include) - prior_model)
    exited = sorted(prior_model - set(include))
    fire = bool(entered or exited) or tier_fire or resize
    kind = ('sizing_migration_invvol' if migration else
            'membership' if (entered or exited) else
            'tier' if tier_fire else 'manual_resize')

    # ---- monthly drift-band re-size (v2 spec A2) ----
    # First real online run of a new calendar month; trades ONLY names outside
    # the drift band. Dry runs and the offline rule-25 gate never see or
    # advance it, so pending_rebalance() stays offline-pure; a vol spike alone
    # never trades intra-month.
    monthly_traded: list[str] = []
    state_changed = False
    if (siz['mode'] == 'inverse_vol' and not fire and not dry_run
            and check_freshness):
        state = cfg.setdefault('sizing_state', {})
        if today[:7] > state.get('last_resize_check', ''):
            cur = current_weights(cfg)
            state['last_resize_check'] = today[:7]
            state_changed = True
            if cur is None:
                flag('monthly re-size skipped: no prices to mark the book')
            else:
                new_w, monthly_traded = drift_band_filter(
                    cur, weights, float(siz['drift_band']))
                if monthly_traded:
                    weights = new_w
                    fire, kind = True, 'resize_monthly'
                    flag(f'monthly re-size: {", ".join(monthly_traded)} '
                         f'outside ±{float(siz["drift_band"]):.0%} band — '
                         f'trading to target')
                else:
                    flag('monthly re-size: all names inside the drift band — '
                         'no trades')

    # ---- tier-change reporting (allocation delta vs prior model weight) ----
    denom = sum(last_ev['allocations'].values()) + last_ev.get('cash', 0)
    prior_w = {t: a / denom for t, a in last_ev['allocations'].items()} if denom else {}
    for t, old, new in tier_chg:
        flag(f'{t}: tier {old} → {new} | allocation '
             f'{prior_w.get(t, 0) * 100:.1f}% → {weights.get(t, 0) * 100:.1f}%')

    by_layer: dict[str, float] = {}
    for t, v in weights.items():
        by_layer[layers[t]] = by_layer.get(layers[t], 0) + v
    for lay, v in sorted(by_layer.items(), key=lambda kv: -kv[1]):
        if v > 0.25:
            flag(f'layer {lay}: {v:.0%} of portfolio (no layer cap active)')

    # ---- sanity: tier-sized weights must be monotonic in score (rule 18).
    # Scoped to tier mode: inverse-vol weights are deliberately NOT
    # score-ordered (v2 spec A0) — the EW_ROSTER shadow audits sizing there.
    if siz['mode'] != 'inverse_vol':
        viol = weights_score_monotonic([(info[t]['TOTAL'], weights.get(t, 0))
                                        for t in include])
        if viol:
            flag(f'non-monotonic weights {viol} — investigate base_weight/caps')

    # ---- band shadow roster upkeep (v2 spec C1): every real run, frozen or
    # not — the scouts track today's ranks regardless of the book.
    shadows_changed = False if dry_run else _update_band_shadows(
        cfg, sel_live, pcfg, today)

    if dry_run:
        print(f'\n{"Tkr":<7}{"Rank":>5}{"Score":>7}{"Status":<34}{"Wt %":>6}')
        for t in sorted(include, key=lambda t: -weights.get(t, 0)):
            print(f'{t:<7}{rank.get(t, 0):>5}{info[t]["TOTAL"]:>7.1f}'
                  f'{statuses.get(t, ""):<34}{weights.get(t, 0) * 100:>6.1f}')
        verdict = ('REBALANCE: ' + build_reason(entered, exited, tier_chg, resize)
                   if fire else 'FREEZE (no membership/tier change)')
        print(f'(dry run — would {verdict}; nothing written)')
        return {'fire': fire, 'entered': entered, 'exited': exited,
                'tier_chg': tier_chg, 'wrote': False, 'pending': new_pending}

    if not fire:
        # The workbook snapshot stays frozen, but the exit-pending clock,
        # band-shadow rosters and monthly-resize stamp are cross-run state and
        # MUST survive this run — persist them to the config (rule 26).
        if (new_pending != dict(cfg.get('exit_pending', {}))
                or shadows_changed or state_changed):
            cfg['exit_pending'] = new_pending
            save_cfg(cfg)
            print(f'cross-run state persisted to config (pending clock'
                  f'{"/shadows" if shadows_changed else ""}'
                  f'{"/resize stamp" if state_changed else ""})')
        print('membership & tiers unchanged since last rebalance — '
              'snapshot frozen, nothing written')
        return {'fire': False, 'entered': [], 'exited': [], 'tier_chg': [],
                'wrote': False, 'pending': new_pending}

    # ---- prices for the Reconciliation trade plan (only if that sheet exists) ----
    # Prices feed the Reconciliation sheet only; skip the network entirely when it
    # was slimmed away (the common case on the notional workbook).
    prices: dict[str, float] = {}
    if recon is not None:
        for t in include:
            prices[t] = current_price(t)
            time.sleep(0.25)
            if prices[t] is None:
                flag(f'{t}: no current price — Reconciliation row left without shares')

    # ---- rewrite Targets ----
    targets.delete_rows(1, targets.max_row)
    targets.append([f'Target Portfolio (refreshed {today}, live Watchlist scores)'])
    targets.append(['Ticker', 'Layer', 'TOTAL', 'Tier', 'Rank', 'Status',
                    'Include?', 'Override', 'Target %', 'Notes'])
    for c in targets[2]:
        c.fill, c.font = copy(HEADER_FILL), copy(HEADER_FONT)
    for x in live:
        t = x['ticker']
        if x['TOTAL'] < 55 and t not in include and t not in overrides:
            continue   # keep the sheet to ✓-and-better plus anything tracked
        targets.append([
            t, x['layer'], round(x['TOTAL'], 2), x['Tier'], rank.get(t),
            statuses.get(t, ''), 'Y' if t in include else 'N',
            overrides.get(t, ''), round(weights.get(t, 0) * 100, 2) or None,
            notes.get(t, ''),
        ])

    # ---- rewrite Reconciliation (skipped if the sheet was slimmed away) ----
    if recon is not None:
        has_positions = positions is not None and any(
            positions.cell(row=r, column=1).value not in (None, 'TOTAL')
            for r in range(2, positions.max_row + 1))
        recon.delete_rows(1, recon.max_row)
        mode = ('target vs current positions'
                if has_positions else 'fresh deployment — Positions sheet is empty')
        recon.append([f'Mechanical target diff ({mode}) — not a trade '
                      f'recommendation; decisions are Dom\'s (CLAUDE.md rule 5)'])
        recon.append([f'Portfolio: ${capital:,.0f} | Prices as of {today}'])
        recon.append(['Ticker', 'Score', 'Target %', '$ Amount', 'Price', 'Shares',
                      'Round $', 'Delta'])
        for c in recon[3]:
            c.fill, c.font = copy(HEADER_FILL), copy(HEADER_FONT)
        held = {}
        if has_positions:
            for r in range(2, positions.max_row + 1):
                t = positions.cell(row=r, column=1).value
                if t and t != 'TOTAL':
                    held[t] = positions.cell(row=r, column=3).value or 0
        cap_row = next(r for r in range(1, sizing.max_row + 1)
                       if sizing.cell(row=r, column=1).value == 'Portfolio Value ($)')
        cap_cell = f"'Sizing Rules'!$B${cap_row}"
        for t in sorted(include, key=lambda t: -weights.get(t, 0)):
            r = recon.max_row + 1
            # Formulas, not pasted values (CLAUDE.md rule 4).
            recon.append([
                t, round(info[t]['TOTAL'], 1), round(weights[t] * 100, 2),
                f'=C{r}/100*{cap_cell}', prices.get(t),
                f'=IF(E{r}="","",ROUND(D{r}/E{r},4))',
                f'=IF(E{r}="","",F{r}*E{r})',
                (f'={held.get(t, 0)}-F{r}' if has_positions else 'NEW'),
            ])
        r = recon.max_row + 1
        recon.append(['TOTAL', None, f'=SUM(C4:C{r - 1})', f'=SUM(D4:D{r - 1})',
                      None, None, f'=SUM(G4:G{r - 1})', None])

    # ---- README breadcrumbs ----
    rm = wb['README']
    for row in rm.iter_rows():
        if row[0].value == 'Last built':
            row[1].value = f'{today} (refresh_targets.py — hysteresis pipeline)'

    wb.save(path)
    print(f'wrote {len(include)} target positions to {path}')

    # ---- log the rebalance event (membership and/or tier change, or resize) ----
    cfg['exit_pending'] = new_pending   # persisted by log_rebalance's save_cfg
    if kind == 'resize_monthly':
        reason = (f'resize_monthly: {", ".join(monthly_traded)} outside '
                  f'drift band')
    elif kind == 'sizing_migration_invvol':
        reason = ('sizing_migration_invvol: one-time switch to inverse-vol '
                  'sizing + rank selection (v2 spec A3)')
    else:
        reason = build_reason(entered, exited, tier_chg, resize)
    tiers_now = {t: info[t]['Tier'] for t in include}
    if siz['mode'] == 'inverse_vol':
        # Any fired full re-size IS that month's scheduled re-size (spec A2).
        cfg.setdefault('sizing_state', {})['last_resize_check'] = today[:7]
    ev = log_rebalance(cfg, weights, reason, tiers_now, kind=kind)
    print(f'model rebalanced: {reason} — value at rebalance '
          f'${sum(ev["allocations"].values()) + ev["cash"]:,.0f}')

    # ---- trade ticket hook (agentic-execution spec B1): every real fired
    # event generates an executable ticket from ACTUAL account state. Gated to
    # the real workbook (portfolio=None) like the score panel; a failure is
    # flagged, never fatal — the model event already logged.
    if portfolio is None:
        try:
            from generate_trade_ticket import on_model_event
            on_model_event({'date': today, 'kind': kind, 'reason': reason},
                           weights)
        except Exception as e:
            flag(f'trade ticket not generated ({e}) — run '
                 f'scripts/generate_trade_ticket.py manually')
    return {'fire': True, 'entered': entered, 'exited': exited,
            'tier_chg': tier_chg, 'wrote': True, 'pending': new_pending}


def pending_rebalance(portfolio: str | None = None) -> bool:
    """True when the committed Targets are STALE vs the live scores — a rebalance
    is pending because membership or a held name's tier has changed since the last
    logged model event, OR an exit-pending clock from a prior date is due to
    confirm (the name is still below the exit score, so the next real run exits
    it). Offline (dry-run path uses no yfinance) and read-only: never starts or
    advances pending clocks. This is the guarantee behind rule 25: True here means
    the weights no longer reflect the scores/rules, and the fix is a real run."""
    import contextlib
    import io
    with contextlib.redirect_stdout(io.StringIO()):   # swallow the dry-run print
        rep = refresh(dry_run=True, portfolio=portfolio, check_freshness=False)
    return bool(rep and rep.get('fire'))


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='Refresh portfolio targets from live scores')
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--resize', action='store_true',
                    help='intentional re-sizing: drop below-exit-score names '
                         'immediately, bypassing the 2-run exit hysteresis')
    ap.add_argument('--portfolio-path', default=None,
                    help='override path to portfolio.xlsx (default: auto-derived '
                         'from repo root)')
    ap.add_argument('--check', action='store_true',
                    help='exit 1 if the Targets are stale vs live scores (a '
                         'rebalance is pending) — the CI/commit gate; writes nothing')
    ap.add_argument('--seam', metavar='REASON',
                    help='stamp a methodology seam starting today (rule 32-C): '
                         'exit clocks started inside the next SEAM_DAYS cannot '
                         'confirm until the window closes. Run this BEFORE the '
                         'recalc --sync that deploys a methodology change.')
    args = ap.parse_args()
    if args.seam:
        if args.dry_run or args.check:
            ap.error('--seam stamps the config — cannot combine with '
                     '--dry-run/--check')
        seam = stamp_seam(args.seam)
        print(f'methodology seam stamped: {seam["date"]} + {SEAM_DAYS}d '
              f'— {seam["reason"]}')
    if args.check:
        import sys
        if pending_rebalance(portfolio=args.portfolio_path):
            print('STALE: Targets do not reflect current scores — a rebalance is '
                  'pending. Run `python3 scripts/refresh_targets.py` (--resize to '
                  'force immediate exits).')
            sys.exit(1)
        print('Targets reflect current scores ✓')
        sys.exit(0)
    refresh(dry_run=args.dry_run, resize=args.resize, portfolio=args.portfolio_path)
