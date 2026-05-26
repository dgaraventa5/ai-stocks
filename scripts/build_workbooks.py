"""One-shot script to create the three tracking workbooks with live formulas.

Run once at project setup. Re-running overwrites — don't run after you've
entered data unless you intend to wipe.

Per CLAUDE.md §4: calculated cells are Excel formulas, not pasted values.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_portfolio():
    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"

    headers = [
        "Ticker", "Layer", "Shares", "Cost Basis ($)", "Current Price ($)",
        "Market Value ($)", "Weight %", "Unrealized P/L ($)", "Unrealized %",
        "Thesis Date", "Notes",
    ]
    ws.append(headers)
    style_header(ws)

    # 20 empty data rows with formulas pre-wired so adding a position only
    # requires filling Ticker/Layer/Shares/Cost Basis/Current Price.
    for r in range(2, 22):
        ws.cell(row=r, column=6, value=f"=IFERROR(C{r}*E{r},\"\")")        # Mkt Value
        ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/SUM($F$2:$F$21),\"\")")  # Weight %
        ws.cell(row=r, column=8, value=f"=IFERROR((E{r}-D{r})*C{r},\"\")") # Unrealized $
        ws.cell(row=r, column=9, value=f"=IFERROR((E{r}-D{r})/D{r},\"\")") # Unrealized %

    # Totals row
    ws.cell(row=23, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=23, column=6, value="=SUM(F2:F21)").font = Font(bold=True)
    ws.cell(row=23, column=7, value="=SUM(G2:G21)").font = Font(bold=True)
    ws.cell(row=23, column=8, value="=SUM(H2:H21)").font = Font(bold=True)

    # Number formats
    for r in range(2, 24):
        ws.cell(row=r, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6).number_format = '"$"#,##0'
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=8).number_format = '"$"#,##0;[Red]-"$"#,##0'
        ws.cell(row=r, column=9).number_format = "0.0%;[Red]-0.0%"

    autosize(ws, [10, 22, 10, 14, 16, 16, 10, 16, 14, 12, 40])

    # Sizing tab — capital allocation guard rails
    sz = wb.create_sheet("Sizing rules")
    sz.append(["Tier", "Score band", "Target weight", "Max weight", "Notes"])
    style_header(sz)
    sz.append(["✓✓✓", "85-100", 0.08, 0.12, "Top-tier conviction"])
    sz.append(["✓✓",  "70-84",  0.05, 0.08, "Strong"])
    sz.append(["✓",   "55-69",  0.03, 0.05, "Average — small starter only"])
    sz.append(["?",   "40-54",  0.00, 0.02, "Below average — research more"])
    sz.append(["✗",   "0-39",   0.00, 0.00, "Avoid"])
    for r in range(2, 7):
        sz.cell(row=r, column=3).number_format = "0.0%"
        sz.cell(row=r, column=4).number_format = "0.0%"
    autosize(sz, [8, 12, 16, 14, 40])

    out = ROOT / "00-master" / "portfolio.xlsx"
    wb.save(out)
    return out


def build_13f_tracking():
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    headers = [
        "Filer", "CIK", "Period (YYYY-Qn)", "Filing date", "Ticker",
        "Shares", "Value ($000s)", "Shares Δ vs prior Q", "% of portfolio",
        "Source URL", "Notes",
    ]
    ws.append(headers)
    style_header(ws)

    # No formulas in this sheet — it's a log. But pre-format columns.
    for r in range(2, 200):
        ws.cell(row=r, column=6).number_format = "#,##0"
        ws.cell(row=r, column=7).number_format = "#,##0"
        ws.cell(row=r, column=8).number_format = "#,##0;[Red]-#,##0"
        ws.cell(row=r, column=9).number_format = "0.00%"

    autosize(ws, [28, 12, 16, 12, 8, 14, 16, 18, 14, 40, 30])

    # Filers sheet — the funds we're tracking
    f = wb.create_sheet("Filers to watch")
    f.append(["Fund", "CIK", "Why we track", "Last 13F pulled"])
    style_header(f)
    starter = [
        ("Berkshire Hathaway", "0001067983", "Buffett — owns AAPL stake worth tracking", ""),
        ("Baillie Gifford",    "0001097278", "Long-duration tech conviction",            ""),
        ("Tiger Global",       "0001167483", "Concentrated tech bets",                   ""),
        ("Coatue Mgmt",        "0001135730", "AI thematic concentration",                ""),
        ("Whale Rock",         "0001394730", "AI infra heavy",                           ""),
        ("Lone Pine",          "0001061165", "Long-only tech",                           ""),
    ]
    for row in starter:
        f.append(list(row))
    autosize(f, [26, 14, 50, 16])

    out = ROOT / "tracking" / "13f-tracking.xlsx"
    wb.save(out)
    return out


def build_hyperscaler_capex():
    wb = Workbook()
    ws = wb.active
    ws.title = "Capex by quarter"

    hyperscalers = ["MSFT", "GOOGL", "META", "AMZN", "ORCL"]
    quarters = []
    for y in (2024, 2025, 2026):
        for q in (1, 2, 3, 4):
            quarters.append(f"{y}-Q{q}")

    # Header: ticker + N quarter columns
    ws.cell(row=1, column=1, value="Ticker")
    for i, q in enumerate(quarters, start=2):
        ws.cell(row=1, column=i, value=q)
    style_header(ws)

    for i, t in enumerate(hyperscalers, start=2):
        ws.cell(row=i, column=1, value=t).font = Font(bold=True)

    # Totals + YoY rows
    total_row = len(hyperscalers) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    yoy_row = total_row + 1
    ws.cell(row=yoy_row, column=1, value="YoY %").font = Font(bold=True)

    last_col = len(quarters) + 1
    for col in range(2, last_col + 1):
        col_letter = get_column_letter(col)
        # Sum across hyperscalers (rows 2 through total_row-1)
        ws.cell(row=total_row, column=col,
                value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})").font = Font(bold=True)
        # YoY: compare to same quarter 4 cols back
        if col >= 6:  # 2025-Q1 onward has a YoY comp
            prior_col = get_column_letter(col - 4)
            ws.cell(row=yoy_row, column=col,
                    value=f"=IFERROR({col_letter}{total_row}/{prior_col}{total_row}-1,\"\")")
            ws.cell(row=yoy_row, column=col).number_format = "0.0%;[Red]-0.0%"

    # Number formats for capex cells ($ in billions, entered as raw $)
    for r in range(2, total_row + 1):
        for col in range(2, last_col + 1):
            ws.cell(row=r, column=col).number_format = '"$"#,##0'

    autosize(ws, [10] + [11] * len(quarters))

    # Notes sheet — capture guidance, commentary, links
    n = wb.create_sheet("Guidance & commentary")
    n.append(["Date", "Ticker", "Source", "Quote / data point", "Implication"])
    style_header(n)
    autosize(n, [12, 8, 30, 60, 40])

    out = ROOT / "tracking" / "hyperscaler-capex.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    for path in (build_portfolio(), build_13f_tracking(), build_hyperscaler_capex()):
        print(f"wrote {path.relative_to(ROOT)}")
