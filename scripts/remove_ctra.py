"""Remove CTRA from the Watchlist — Coterra merged into Devon (2026-06-10 cleanup).

Coterra Energy (CTRA) completed an all-stock merger into Devon Energy (DVN) at
0.70 DVN/share, consummated 2026-05-07, delisted from NYSE same day. The entity
no longer exists, so its frozen pre-merger ratings (rank ~13 on stale data, the
last Layer 1 gate violation) are removed rather than re-scored. The Rating Audit
history for CTRA is retained as a record; a tombstone audit row notes the merger.
Successor DVN is a separate new-name decision (different company, oil-weighted).
"""
from __future__ import annotations

from openpyxl import load_workbook

XLSX = '/Users/dom/Desktop/ai-stocks/00-master/ai_supply_chain_scoring.xlsx'


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)
    ws = wb['Watchlist']
    audit = wb['Rating Audit']

    target = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'CTRA':
            target = r
            break
    if target is None:
        print('CTRA not on Watchlist — nothing to remove')
        return

    audit.append(['2026-06-10', 'CTRA', '—', '—',
                  'Removed from Watchlist: Coterra merged into Devon Energy '
                  '(DVN) at 0.70 DVN/sh, consummated 2026-05-07, delisted NYSE '
                  'same day. Pre-merger ratings retired (not re-scored — entity '
                  'gone). Successor DVN is a separate new-name decision.',
                  'Globe and Mail / StockTitan 8-K 2026-05-07', 'High', 'Removed'])
    ws.delete_rows(target, 1)
    wb.save(XLSX)
    print(f'removed CTRA (was row {target}); merger tombstone logged to Rating Audit')


if __name__ == '__main__':
    main()
