"""Refresh objective Watchlist inputs (portfolio / all / subset) — rule-9 helper.

Subjective rating cells are NEVER touched. Deliberate-blank conventions
(foreign-ADR currency, EPS-YoY one-offs, negative-EBITDA, ROIC) are guarded so
a mechanical refresh cannot silently regress them. See
docs/superpowers/specs/2026-06-24-refresh-objective-inputs-design.md.
"""
from __future__ import annotations

import sys
import json
import time
import argparse
import subprocess
import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

try:                                 # works as a sibling (scripts/ on path) AND
    import adr_currency              # as scripts.refresh_objective_inputs (package);
except ModuleNotFoundError:          # pure + yfinance-free either way
    from scripts import adr_currency

# Sibling scripts (batch_score, momentum_50dma) import yfinance at module load.
# The site deploy runs `pytest tests` with ONLY openpyxl+pytest installed (no
# yfinance), so those are imported LAZILY inside the functions that need them —
# keeping this module and its tests importable without yfinance.
sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(__file__).resolve().parent.parent
SCORING_PATH = ROOT / "00-master" / "ai_supply_chain_scoring.xlsx"
PORTFOLIO_PATH = ROOT / "00-master" / "portfolio.xlsx"

# Mirrored from batch_score so mw_staleness_flag doesn't import batch_score
# (which pulls yfinance, absent in the site-deploy pytest env). Keep in sync
# with batch_score._is_layer9_capacity / _CAPACITY_JSON.
_CAPACITY_JSON = ROOT / "00-master" / "capacity-mw.json"


def _is_layer9_capacity(layer):
    if not layer or not str(layer).strip().startswith("09"):
        return False
    l = str(layer).lower()
    return "bitcoin" in l or "neocloud" in l


def _watchlist_tickers(scoring_path) -> list[str]:
    ws = load_workbook(scoring_path, read_only=True)["Watchlist"]
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            out.append(str(v).strip().upper())
    return out


def _portfolio_hold_tickers(portfolio_path) -> list[str]:
    ws = load_workbook(portfolio_path, read_only=True)["Targets"]
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Ticker":
            header_row = r
            break
    status_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == "Status":
            status_col = c
            break
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        s = ws.cell(row=r, column=status_col).value
        if t and s and str(s).strip().upper() == "HOLD":
            out.append(str(t).strip().upper())
    return out


def resolve_targets(arg, scoring_path=SCORING_PATH, portfolio_path=PORTFOLIO_PATH) -> list[str]:
    watchlist = _watchlist_tickers(scoring_path)
    if arg == "all":
        return watchlist
    if arg == "portfolio":
        hold = _portfolio_hold_tickers(portfolio_path)
        wlset = set(watchlist)
        return [t for t in hold if t in wlset]
    # explicit subset
    wlset = set(watchlist)
    return [t.strip().upper() for t in arg if t.strip().upper() in wlset]


OBJ_COLS = {"fwd_pe": 5, "ev_ebitda": 6, "fcf_yield": 7, "ps": 8,
            "roic": 11, "gross_mgn": 12, "fcf_mgn": 13, "nd_ebitda": 14,
            "rev_3y_cagr": 16, "rev_yoy": 17, "eps_yoy": 18}
DMA_COL = 29
LAST_UPDATED_COL = 4
EPS_YOY_EXTREME = 300.0
_ADR_BLANK_KEYS = ("ps", "ev_ebitda", "fcf_yield")
# Keys where a fresh None must blank the cell (guard-2/3/4 deliberate blanks must
# win over guard-5's keep-prior behaviour).  For all other keys, guard-5 protects
# existing values from a transient yfinance "no data" return.
_BLANK_THROUGH_ON_NONE = frozenset({"ev_ebitda", "nd_ebitda"})


def _blank(v):
    return v is None or v == ""


def apply_guards(info, fresh, existing):
    """Apply smart blank-handling guards to produce safe write set.

    Args:
        info: yfinance info dict (needs 'financialCurrency').
        fresh: compute_inputs output dict (input-key → value).
        existing: current Watchlist cell values (input-key → value or None/"" if blank).

    Returns:
        (writes, flags): writes maps input-key → value for keys to write (absent = leave
        untouched; present-None = write blank). flags is a list of human-readable strings.
    """
    writes, flags = {}, []

    # Guard 1: foreign-ADR currency trap. A trading-vs-financial currency
    # mismatch corrupts the market-cap-based ratios (see adr_currency). Detection
    # is trading != financial — NOT "financial != USD", which also wrongly blanks
    # a clean same-currency foreign name (Vanguard 5347.TWO, TWD/TWD).
    # compute_inputs is authoritative: it sources the 3 from the local listing
    # (fresh carries real values) or leaves them None (unmapped). Write those
    # through here; never force-blank a clean same-currency name.
    trading = info.get("currency")
    financial = info.get("financialCurrency")
    trap = adr_currency.has_currency_mismatch(trading, financial)
    if trap:
        for k in _ADR_BLANK_KEYS:
            writes[k] = fresh.get(k)
        fixed = [k for k in _ADR_BLANK_KEYS if fresh.get(k) is not None]
        if fixed:
            flags.append(
                f"Foreign filer ({trading}/{financial}): P/S, EV/EBITDA, FCF-Yield "
                f"converted to a common currency where available ({', '.join(fixed)})."
            )
        else:
            flags.append(
                f"Foreign filer ({trading}/{financial}): P/S, EV/EBITDA, FCF-Yield "
                f"could not be put in a common currency — blanked."
            )

    # Per-key loop
    for key in OBJ_COLS:
        # trap keys handled above (written or blanked); skip in the per-key loop
        if trap and key in _ADR_BLANK_KEYS:
            continue

        if key == "eps_yoy":
            # (a) existing blank → preserve, do not write
            if _blank(existing.get(key)):
                flags.append(
                    "EPS YoY: cell currently blank (likely a documented one-off) "
                    "— preserved blank, not refilled."
                )
                continue
            # (b) fresh value extreme → withhold
            v = fresh.get(key)
            if v is not None and abs(v) >= EPS_YOY_EXTREME:
                flags.append(
                    f"EPS YoY: fresh {v:+.0f}% ≥ {EPS_YOY_EXTREME:.0f}% "
                    f"→ withheld as possible one-off (rule 15). Confirm blank or write."
                )
                continue
            # (c) fresh None over a non-blank existing → keep existing (don't clobber).
            # yfinance earningsQuarterlyGrowth is frequently None; same protection
            # the generic-key path applies (guard 4).
            if v is None and not _blank(existing.get(key)):
                flags.append(
                    f"eps_yoy: fetch returned no data — kept prior value {existing.get(key)}."
                )
                continue
            # (d) normal — write (may be None when existing is also blank)
            writes[key] = v
            continue

        # Guard 4/5: every other key.
        # For most keys: if fresh is None and existing is non-blank, keep existing
        # (guard 5 — protects curated values like ROIC from a transient fetch gap).
        # Exception: ev_ebitda and nd_ebitda blank-through on a fresh None so that
        # guard-2 (non-positive EBITDA) and guard-3/4 (L10 EV/FCF / L9 EV/MW
        # missing denominator) blanks win over keep-prior (_BLANK_THROUGH_ON_NONE).
        v = fresh.get(key)
        if v is None and key not in _BLANK_THROUGH_ON_NONE and not _blank(existing.get(key)):
            if key == "roic":
                # ROIC is never fetched (yfinance.info has no ROIC — batch_score
                # leaves it blank), so this branch fires on EVERY refresh. Say so:
                # the earlier wording read as a transient failure and got logged
                # as one twice (sentinel 2026-09-02/03). It is a curated input.
                flags.append(
                    f"roic: curated input, not fetched (yfinance has no ROIC) — kept "
                    f"{existing.get(key)}; hand-refresh from the filing after earnings."
                )
            else:
                flags.append(
                    f"{key}: fetch returned no data — kept prior value {existing.get(key)}."
                )
            continue
        if v is None and key in _BLANK_THROUGH_ON_NONE and not _blank(existing.get(key)):
            flags.append(
                f"{key}: no computable value (e.g. non-positive EBITDA / missing denominator)"
                f" — blanked (was {existing.get(key)})."
            )
        writes[key] = v

    return writes, flags


def mw_staleness_flag(ticker, layer, today, capacity_json=None):
    """Check Layer-9 capacity cohort MW data staleness (rule 13).

    Returns flag string if layer is Layer-9 capacity AND entry is missing OR as_of >90 days old;
    else None.

    Args:
        ticker: Stock ticker symbol.
        layer: Layer string from Watchlist.
        today: date object for age calculation.
        capacity_json: Override path to capacity-mw.json (for testing).

    Returns:
        str | None: Flag message or None.
    """
    if not _is_layer9_capacity(layer):
        return None
    path = capacity_json if capacity_json is not None else _CAPACITY_JSON
    try:
        data = json.loads(Path(path).read_text())
    except Exception:
        data = {}
    rec = data.get(ticker)
    if not rec or rec.get("secured_gross_mw") is None:
        return "EV/MW: missing from capacity-mw.json — add it (rule 13)."
    # Real capacity-mw.json uses the key "asof" (no underscore); accept the
    # underscored spelling too for resilience against legacy entries.
    as_of = rec.get("asof") or rec.get("as_of")
    try:
        age = (today - dt.date.fromisoformat(as_of)).days
    except Exception:
        return f"EV/MW: capacity-mw.json asof unparseable ({as_of!r}) — stale, refresh MW."
    if age > 90:
        return f"EV/MW: capacity-mw.json asof {as_of} → {age} days old (>90) → stale, refresh MW."
    return None


def read_existing(ws, row):
    """Read current cell values for all objective input keys.

    Args:
        ws: openpyxl worksheet.
        row: Row number to read from.

    Returns:
        dict: maps input-key to current cell value (or None/"" if blank).
    """
    return {k: ws.cell(row=row, column=c).value for k, c in OBJ_COLS.items()}


def _same(a, b):
    """True when two cell values are equivalent (blank == blank)."""
    if _blank(a) and _blank(b):
        return True
    return a == b


def plan_row(ws, row, writes, dma_value, today_iso):
    """Compute what write_row WOULD do, without writing anything.

    Split out so `--dry-run` can report the same column set as a live run and
    show old → new per changed cell. Before this existed, dry runs reported an
    empty touched list, so the review step that is supposed to catch corrupt
    values (rule 27) could not see them — the 2026-08-13 VST refresh moved a
    name a full tier after a dry run printed no changes at all.

    Args:
        ws: openpyxl worksheet.
        row: Row number that would be written.
        writes: dict mapping input-key → value (same contract as write_row).
        dma_value: Value for DMA_COL, or None to skip it.
        today_iso: ISO date string for LAST_UPDATED_COL.

    Returns:
        (touched, changes) where touched is the sorted list of column indices
        that would be written, and changes is a list of
        {"col", "field", "old", "new"} dicts for the subset whose value would
        actually differ from what is in the cell now.
    """
    planned = [(OBJ_COLS[k], k, v) for k, v in writes.items()]
    if dma_value is not None:
        planned.append((DMA_COL, "50dma", dma_value))
    planned.append((LAST_UPDATED_COL, "last_updated", today_iso))

    changes = []
    for col, field, new in planned:
        old = ws.cell(row=row, column=col).value
        if not _same(old, new):
            changes.append({"col": col, "field": field, "old": old, "new": new})
    changes.sort(key=lambda c: c["col"])
    return sorted({c for c, _f, _v in planned}), changes


def write_row(ws, row, writes, dma_value, today_iso):
    """Write objective inputs and metadata to a row, never touching subjective columns.

    Args:
        ws: openpyxl worksheet.
        row: Row number to write to.
        writes: dict mapping input-key → value. Absent key = leave untouched;
                present key with None value = write blank.
        dma_value: Value for DMA_COL, or None to skip writing it.
        today_iso: ISO date string for LAST_UPDATED_COL.

    Returns:
        list[int]: Sorted list of column indices actually written.
    """
    # Plan first so the dry-run preview and the live write can never drift apart.
    touched, _changes = plan_row(ws, row, writes, dma_value, today_iso)
    for key, val in writes.items():
        ws.cell(row=row, column=OBJ_COLS[key]).value = val
    if dma_value is not None:
        ws.cell(row=row, column=DMA_COL).value = dma_value
    ws.cell(row=row, column=LAST_UPDATED_COL).value = today_iso
    return touched


# ---------------------------------------------------------------------------
# Task 5: real yfinance fetcher + orchestration
# ---------------------------------------------------------------------------

def _default_fetcher(ticker, layer):
    """Default fetcher: calls yfinance + compute_inputs. Injectable for tests.

    yfinance/batch_score imported lazily so this module stays importable
    without yfinance (the site deploy's pytest env has openpyxl only).
    """
    import yfinance as yf
    from batch_score import compute_inputs
    t = yf.Ticker(ticker)
    info = t.info or {}
    fresh, _gaps = compute_inputs(ticker, t, info, layer=layer)
    return info, fresh


def _layer_of(ws, row):
    """Return the layer string from column 3 for the given row."""
    return ws.cell(row=row, column=3).value


def refresh(targets, dry_run, scoring_path=SCORING_PATH, fetcher=None,
            dma_fetcher=None, today=None):
    """Orchestrate fetch → guards → (write|stage) for each target ticker.

    Args:
        targets: List of ticker strings (already resolved).
        dry_run: If True, do not write or save the workbook.
        scoring_path: Path to the scoring .xlsx file.
        fetcher: Callable(ticker, layer) -> (info, fresh_dict). Defaults to
                 _default_fetcher (real yfinance). Injectable for tests.
        dma_fetcher: Callable(ticker) -> float|None. Defaults to
                     pct_days_above_50dma. Injectable for tests.
        today: date object for Last Updated and staleness checks. Defaults
               to dt.date.today().

    Returns:
        dict with keys: mode_count, rows, flags, wrote.
    """
    fetcher = fetcher or _default_fetcher
    if dma_fetcher is None:
        from momentum_50dma import pct_days_above_50dma
        dma_fetcher = pct_days_above_50dma
    today = today or dt.date.today()

    wb = load_workbook(scoring_path, data_only=False)
    ws = wb["Watchlist"]
    rows = {
        str(ws.cell(row=r, column=1).value).strip().upper(): r
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    }

    report = {"mode_count": len(targets), "rows": [], "flags": [], "wrote": False}

    for ticker in targets:
        row = rows.get(ticker)
        if row is None:
            report["flags"].append(f"{ticker}: not on Watchlist — skipped.")
            continue

        layer = _layer_of(ws, row)
        try:
            info, fresh = fetcher(ticker, layer)
        except Exception as e:
            # Isolate per-ticker fetch failures: leave the row unchanged (never
            # clobber with bad data), flag it, and keep going. Matters most on
            # the ~167-name `all` run where one transient yfinance error would
            # otherwise abort the whole pass and save nothing.
            report["flags"].append(f"{ticker}: fetch error — {e} — skipped (row unchanged).")
            report["rows"].append({"ticker": ticker, "touched": [], "changes": [],
                                   "flags": [f"fetch error — {e}"]})
            if fetcher is _default_fetcher:
                time.sleep(0.3)
            continue
        existing = read_existing(ws, row)
        writes, flags = apply_guards(info, fresh, existing)

        mwf = mw_staleness_flag(ticker, layer, today)
        if mwf:
            flags.append(mwf)

        try:
            dma_value = dma_fetcher(ticker)
        except Exception as e:
            dma_value = None
            flags.append(f"50DMA: fetch error — {e}")

        # Plan unconditionally so a dry run reports exactly what a live run
        # would do; only the write itself is gated on dry_run.
        touched, changes = plan_row(ws, row, writes, dma_value, today.isoformat())
        if not dry_run:
            write_row(ws, row, writes, dma_value, today.isoformat())

        report["rows"].append({"ticker": ticker, "touched": touched,
                               "changes": changes, "flags": flags})
        report["flags"].extend(f"{ticker}: {f}" for f in flags)

        # Throttle only for the real yfinance fetcher; injected test fakes are fast
        if fetcher is _default_fetcher:
            time.sleep(0.3)

    if not dry_run:
        wb.save(scoring_path)
        report["wrote"] = True

    return report


def _fmt(v):
    return "(blank)" if _blank(v) else str(v)


def format_row_lines(row_rep, dry_run):
    """Render one row's plan as operator-facing lines: cols, then old → new.

    The report dict never reaches the human running the script; these lines are
    what the rule-27 "review before you write" step actually reads.
    """
    verb = "would write" if dry_run else "wrote"
    changes = row_rep.get("changes") or []
    head = f"  {row_rep['ticker']:<8} {verb} cols: {row_rep['touched']}"
    if not changes:
        return [head + "  — no value changes"]
    lines = [head + f"  — {len(changes)} value change(s)"]
    width = max(len(c["field"]) for c in changes)
    for c in changes:
        lines.append(f"      {c['field']:<{width}}  {_fmt(c['old'])} -> {_fmt(c['new'])}")
    return lines


def main():
    """CLI entry point: refresh objective inputs for portfolio / all / subset."""
    ap = argparse.ArgumentParser(description="Refresh objective Watchlist inputs.")
    ap.add_argument("target", nargs="+", help="'portfolio', 'all', or ticker(s)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    # Single 'portfolio' or 'all' token → pass as string; otherwise ticker list
    if len(args.target) == 1 and args.target[0] in ("portfolio", "all"):
        arg = args.target[0]
    else:
        arg = args.target

    targets = resolve_targets(arg)
    mode_label = "DRY RUN" if args.dry_run else "LIVE"
    print(f"REFRESH OBJECTIVE INPUTS — {len(targets)} names — {mode_label}")

    rep = refresh(targets, dry_run=args.dry_run)

    for r in rep["rows"]:
        for line in format_row_lines(r, args.dry_run):
            print(line)

    if rep["flags"]:
        print("\nJUDGMENT FLAGS (human ruling needed):")
        for f in rep["flags"]:
            print(f"  • {f}")

    if rep["wrote"]:
        print("\nTOTAL changes (after recalc):")
        out = subprocess.run(
            ["python3", str(ROOT / "scripts" / "recalc_watchlist.py")],
            capture_output=True, text=True,
        )
        print(out.stdout)


if __name__ == "__main__":
    main()
