"""Reweight scoring system and add PEG ratio as 5th Value metric.

Changes:
  1. Weights: Value 15→20%, AI Thesis 30→20%, Risk 10→15%
  2. Insert PEG column (col I) between P/S and Value Score
  3. Update all formulas across all rows for the column shift
  4. Add PEG to Methodology sheet
  5. Compute and display new scores (openpyxl can't evaluate formulas)
"""

from pathlib import Path
from openpyxl import load_workbook
import math

ROOT = Path(__file__).resolve().parent.parent
SCORING_PATH = ROOT / "00-master" / "ai_supply_chain_scoring.xlsx"

# ── New weights ──────────────────────────────────────────────────────
NEW_WEIGHTS = {
    "B4": 0.20,  # Value: 15 → 20%
    "B5": 0.20,  # Quality: unchanged
    "B6": 0.15,  # Growth: unchanged
    "B7": 0.20,  # AI Thesis: 30 → 20%
    "B8": 0.10,  # Momentum: unchanged
    "B9": 0.15,  # Risk: 10 → 15%
}

# ── Formula templates (row number injected at write time) ────────────
# After inserting PEG at col I(9), column layout becomes:
#   E:FwdPE  F:EV/EBITDA  G:FCFYld  H:P/S  I:PEG
#   J:ValueScore
#   K:ROIC  L:GrossMgn  M:FCFMgn  N:ND/EBITDA
#   O:QualityScore
#   P:Rev3yCagr  Q:RevYoY  R:EPSYoY
#   S:GrowthScore
#   T:AIRev  U:Position  V:Moat  W:Capacity  X:Hypersc
#   Y:AIScore
#   Z:EPSRev  AA:RelStr  AB:Insider
#   AC:MomentumScore
#   AD:CustConc  AE:GeoRisk  AF:BSRisk  AG:RegRisk
#   AH:RiskScore
#   AI:TOTAL
#   AJ:Tier

def peg_formula(r):
    return f'=IF(OR(E{r}="",R{r}="",R{r}<=0,E{r}<0),"",E{r}/R{r})'

def value_score_formula(r):
    fwd_pe  = f'IFERROR(IF(E{r}="","",IF(E{r}<0,5,IF(E{r}<=15,90,IF(E{r}<=20,75,IF(E{r}<=25,60,IF(E{r}<=30,45,IF(E{r}<=40,30,IF(E{r}<=60,15,5)))))))),"")'
    ev_ebitda = f'IFERROR(IF(F{r}="","",IF(F{r}<0,5,IF(F{r}<=10,90,IF(F{r}<=15,75,IF(F{r}<=20,60,IF(F{r}<=25,45,IF(F{r}<=35,30,IF(F{r}<=50,15,5)))))))),"")'
    fcf_yld = f'IFERROR(IF(G{r}="","",IF(G{r}>=8,100,IF(G{r}>=6,90,IF(G{r}>=4,75,IF(G{r}>=2,60,IF(G{r}>=1,45,IF(G{r}>=0,30,15))))))),"")'
    ps      = f'IFERROR(IF(H{r}="","",IF(H{r}<=2,90,IF(H{r}<=4,75,IF(H{r}<=7,60,IF(H{r}<=10,45,IF(H{r}<=15,30,IF(H{r}<=25,15,5))))))),"")'
    peg     = f'IFERROR(IF(I{r}="","",IF(I{r}<=0.5,100,IF(I{r}<=1,90,IF(I{r}<=1.5,75,IF(I{r}<=2,60,IF(I{r}<=2.5,45,IF(I{r}<=3,30,15))))))),"")'
    return f'=IFERROR(AVERAGE({fwd_pe},{ev_ebitda},{fcf_yld},{ps},{peg}),"")'

def quality_score_formula(r):
    roic    = f'IFERROR(IF(K{r}="","",IF(K{r}>=25,100,IF(K{r}>=20,90,IF(K{r}>=15,75,IF(K{r}>=10,60,IF(K{r}>=5,45,IF(K{r}>=0,30,15))))))),"")'
    gm      = f'IFERROR(IF(L{r}="","",IF(L{r}>=60,100,IF(L{r}>=50,90,IF(L{r}>=40,75,IF(L{r}>=30,60,IF(L{r}>=20,45,IF(L{r}>=10,30,15))))))),"")'
    fcf_mgn = f'IFERROR(IF(M{r}="","",IF(M{r}>=25,100,IF(M{r}>=20,90,IF(M{r}>=15,75,IF(M{r}>=10,60,IF(M{r}>=5,45,IF(M{r}>=0,30,15))))))),"")'
    nd      = f'IFERROR(IF(N{r}="","",IF(N{r}<=0,100,IF(N{r}<=1,90,IF(N{r}<=2,75,IF(N{r}<=3,60,IF(N{r}<=4,45,IF(N{r}<=5,30,15))))))),"")'
    return f'=IFERROR(AVERAGE({roic},{gm},{fcf_mgn},{nd}),"")'

def growth_score_formula(r):
    cagr  = f'IFERROR(IF(P{r}="","",IF(P{r}>=40,100,IF(P{r}>=30,90,IF(P{r}>=20,75,IF(P{r}>=15,60,IF(P{r}>=10,45,IF(P{r}>=5,30,15))))))),"")'
    yoy   = f'IFERROR(IF(Q{r}="","",IF(Q{r}>=40,100,IF(Q{r}>=30,90,IF(Q{r}>=20,75,IF(Q{r}>=10,60,IF(Q{r}>=5,45,IF(Q{r}>=0,30,15))))))),"")'
    eps   = f'IFERROR(IF(R{r}="","",IF(R{r}>=40,100,IF(R{r}>=30,90,IF(R{r}>=20,75,IF(R{r}>=10,60,IF(R{r}>=0,45,IF(R{r}>=-10,30,15))))))),"")'
    return f'=IFERROR(AVERAGE({cagr},{yoy},{eps}),"")'

def ai_score_formula(r):
    return f'=IFERROR(AVERAGE(T{r},U{r},V{r},W{r},X{r})*20,"")'

def momentum_score_formula(r):
    return f'=IFERROR(AVERAGE(Z{r},AA{r},AB{r})*20,"")'

def risk_score_formula(r):
    return f'=IFERROR(AVERAGE(AD{r},AE{r},AF{r},AG{r})*20,"")'

def total_formula(r):
    return f'=IFERROR(J{r}*Weights!$B$4 + O{r}*Weights!$B$5 + S{r}*Weights!$B$6 + Y{r}*Weights!$B$7 + AC{r}*Weights!$B$8 + AH{r}*Weights!$B$9,"")'

def tier_formula(r):
    return f'=IF(AI{r}="","",IF(AI{r}>=85,"✓✓✓",IF(AI{r}>=70,"✓✓",IF(AI{r}>=55,"✓",IF(AI{r}>=40,"?","✗")))))'


# ── Scoring functions (Python-side, mirrors the formulas) ────────────

def _score_lower_better(val, bands):
    """Score a metric where lower is better. bands = [(threshold, score), ...]"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if val < 0:
        return 5
    for thresh, score in bands:
        if val <= thresh:
            return score
    return bands[-1][1]  # worst band

def _score_higher_better(val, bands):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    for thresh, score in bands:
        if val >= thresh:
            return score
    return bands[-1][1]

def score_fwd_pe(v):
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    if v < 0: return 5
    if v <= 15: return 90
    if v <= 20: return 75
    if v <= 25: return 60
    if v <= 30: return 45
    if v <= 40: return 30
    if v <= 60: return 15
    return 5

def score_ev_ebitda(v):
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    if v < 0: return 5
    if v <= 10: return 90
    if v <= 15: return 75
    if v <= 20: return 60
    if v <= 25: return 45
    if v <= 35: return 30
    if v <= 50: return 15
    return 5

def score_fcf_yield(v):
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    if v >= 8: return 100
    if v >= 6: return 90
    if v >= 4: return 75
    if v >= 2: return 60
    if v >= 1: return 45
    if v >= 0: return 30
    return 15

def score_ps(v):
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    if v <= 2: return 90
    if v <= 4: return 75
    if v <= 7: return 60
    if v <= 10: return 45
    if v <= 15: return 30
    if v <= 25: return 15
    return 5

def score_peg(v):
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    if v <= 0.5: return 100
    if v <= 1.0: return 90
    if v <= 1.5: return 75
    if v <= 2.0: return 60
    if v <= 2.5: return 45
    if v <= 3.0: return 30
    return 15

def score_roic(v):      return _score_higher_better(v, [(25,100),(20,90),(15,75),(10,60),(5,45),(0,30),(-5,15)])
def score_gm(v):        return _score_higher_better(v, [(60,100),(50,90),(40,75),(30,60),(20,45),(10,30),(0,15)])
def score_fcf_mgn(v):   return _score_higher_better(v, [(25,100),(20,90),(15,75),(10,60),(5,45),(0,30),(-10,15)])
def score_nd_ebitda(v):
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    if v <= 0: return 100
    if v <= 1: return 90
    if v <= 2: return 75
    if v <= 3: return 60
    if v <= 4: return 45
    if v <= 5: return 30
    return 15

def score_rev_cagr(v):  return _score_higher_better(v, [(40,100),(30,90),(20,75),(15,60),(10,45),(5,30),(0,15)])
def score_rev_yoy(v):   return _score_higher_better(v, [(40,100),(30,90),(20,75),(10,60),(5,45),(0,30),(-10,15)])
def score_eps_yoy(v):   return _score_higher_better(v, [(40,100),(30,90),(20,75),(10,60),(0,45),(-10,30),(-25,15)])

def avg_non_none(vals):
    valid = [v for v in vals if v is not None]
    return sum(valid) / len(valid) if valid else None


def compute_all_scores(rows, weights):
    """Compute scores for all rows using new weights + PEG. Returns list of dicts."""
    results = []
    for row in rows:
        ticker = row["ticker"]

        # PEG
        fwd_pe = row.get("fwd_pe")
        eps_yoy = row.get("eps_yoy")
        peg = None
        if fwd_pe is not None and eps_yoy is not None and eps_yoy > 0 and fwd_pe >= 0:
            peg = fwd_pe / eps_yoy

        # Category scores
        val_scores = [score_fwd_pe(row.get("fwd_pe")), score_ev_ebitda(row.get("ev_ebitda")),
                      score_fcf_yield(row.get("fcf_yield")), score_ps(row.get("ps")),
                      score_peg(peg)]
        value = avg_non_none(val_scores)

        qual_scores = [score_roic(row.get("roic")), score_gm(row.get("gm")),
                       score_fcf_mgn(row.get("fcf_mgn")), score_nd_ebitda(row.get("nd_ebitda"))]
        quality = avg_non_none(qual_scores)

        grow_scores = [score_rev_cagr(row.get("rev_cagr")), score_rev_yoy(row.get("rev_yoy")),
                       score_eps_yoy(row.get("eps_yoy"))]
        growth = avg_non_none(grow_scores)

        # Subjective (1-5 → *20)
        ai_vals = [row.get(k) for k in ("ai_rev","position","moat","capacity","hypersc")]
        ai_valid = [v for v in ai_vals if v is not None]
        ai_score = (sum(ai_valid) / len(ai_valid) * 20) if ai_valid else None

        mom_vals = [row.get(k) for k in ("eps_rev","rel_str","insider")]
        mom_valid = [v for v in mom_vals if v is not None]
        momentum = (sum(mom_valid) / len(mom_valid) * 20) if mom_valid else None

        risk_vals = [row.get(k) for k in ("cust_conc","geo_risk","bs_risk","reg_risk")]
        risk_valid = [v for v in risk_vals if v is not None]
        risk = (sum(risk_valid) / len(risk_valid) * 20) if risk_valid else None

        # Total
        cats = [
            (value,    weights["value"]),
            (quality,  weights["quality"]),
            (growth,   weights["growth"]),
            (ai_score, weights["ai_thesis"]),
            (momentum, weights["momentum"]),
            (risk,     weights["risk"]),
        ]
        total_parts = [(s, w) for s, w in cats if s is not None]
        if total_parts:
            total = sum(s * w for s, w in total_parts) / sum(w for _, w in total_parts) * sum(w for _, w in cats)
            # Simpler: just weighted sum (same as formula, which assumes all categories present)
            total = sum(s * w for s, w in total_parts)
        else:
            total = None

        def tier(t):
            if t is None: return ""
            if t >= 85: return "✓✓✓"
            if t >= 70: return "✓✓"
            if t >= 55: return "✓"
            if t >= 40: return "?"
            return "✗"

        results.append({
            "ticker": ticker, "company": row.get("company",""),
            "layer": row.get("layer",""),
            "peg": peg,
            "value": value, "quality": quality, "growth": growth,
            "ai_score": ai_score, "momentum": momentum, "risk": risk,
            "total": total, "tier": tier(total),
        })
    return results


def _val(cell):
    """Extract numeric value from cell, return None if empty/non-numeric."""
    v = cell.value
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.startswith("="):
        return None  # formula — can't evaluate
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def main():
    print("Loading workbook...")
    wb = load_workbook(SCORING_PATH)

    # ── Step 1: Update weights ───────────────────────────────────────
    ws_w = wb["Weights"]
    for cell_ref, new_val in NEW_WEIGHTS.items():
        ws_w[cell_ref] = new_val
    print("✓ Weights updated: Value=20%, Quality=20%, Growth=15%, AI Thesis=20%, Momentum=10%, Risk=15%")

    # ── Step 2: Read all data BEFORE inserting column ────────────────
    ws = wb["Watchlist"]
    max_row = ws.max_row

    # Read data from current layout (before column shift)
    all_rows = []
    for r in range(2, max_row + 1):
        ticker = ws.cell(row=r, column=1).value
        if not ticker:
            continue
        all_rows.append({
            "row": r,
            "ticker": str(ticker).strip(),
            "company": ws.cell(row=r, column=2).value or "",
            "layer": ws.cell(row=r, column=3).value or "",
            "fwd_pe":    _val(ws.cell(row=r, column=5)),
            "ev_ebitda": _val(ws.cell(row=r, column=6)),
            "fcf_yield": _val(ws.cell(row=r, column=7)),
            "ps":        _val(ws.cell(row=r, column=8)),
            "roic":      _val(ws.cell(row=r, column=10)),
            "gm":        _val(ws.cell(row=r, column=11)),
            "fcf_mgn":   _val(ws.cell(row=r, column=12)),
            "nd_ebitda": _val(ws.cell(row=r, column=13)),
            "rev_cagr":  _val(ws.cell(row=r, column=15)),
            "rev_yoy":   _val(ws.cell(row=r, column=16)),
            "eps_yoy":   _val(ws.cell(row=r, column=17)),
            "ai_rev":    _val(ws.cell(row=r, column=19)),
            "position":  _val(ws.cell(row=r, column=20)),
            "moat":      _val(ws.cell(row=r, column=21)),
            "capacity":  _val(ws.cell(row=r, column=22)),
            "hypersc":   _val(ws.cell(row=r, column=23)),
            "eps_rev":   _val(ws.cell(row=r, column=25)),
            "rel_str":   _val(ws.cell(row=r, column=26)),
            "insider":   _val(ws.cell(row=r, column=27)),
            "cust_conc": _val(ws.cell(row=r, column=29)),
            "geo_risk":  _val(ws.cell(row=r, column=30)),
            "bs_risk":   _val(ws.cell(row=r, column=31)),
            "reg_risk":  _val(ws.cell(row=r, column=32)),
        })

    print(f"  Read {len(all_rows)} tickers from Watchlist")

    # ── Step 3: Insert PEG column at position 9 ─────────────────────
    ws.insert_cols(9)
    ws.cell(row=1, column=9, value="PEG")
    print("✓ PEG column inserted at position I (between P/S and Value Score)")

    # ── Step 4: Write formulas for all data rows ─────────────────────
    formula_map = {
         9: peg_formula,             # I: PEG
        10: value_score_formula,     # J: Value Score
        15: quality_score_formula,   # O: Quality Score
        19: growth_score_formula,    # S: Growth Score
        25: ai_score_formula,        # Y: AI Score
        29: momentum_score_formula,  # AC: Momentum Score
        34: risk_score_formula,      # AH: Risk Score
        35: total_formula,           # AI: TOTAL
        36: tier_formula,            # AJ: Tier
    }

    for r in range(2, max_row + 1):
        ticker = ws.cell(row=r, column=1).value
        if not ticker:
            continue
        for col, fn in formula_map.items():
            ws.cell(row=r, column=col, value=fn(r))

    print(f"✓ Formulas updated for {len(all_rows)} rows × 9 formula columns")

    # ── Step 5: Update Methodology sheet ─────────────────────────────
    ws_m = wb["Methodology"]
    # Find the right place to insert PEG (after P/S row)
    insert_row = None
    for r in range(1, ws_m.max_row + 1):
        v = ws_m.cell(row=r, column=1).value
        if v and "Price / Sales" in str(v):
            insert_row = r + 1
            break

    if insert_row:
        ws_m.insert_rows(insert_row)
        ws_m.cell(row=insert_row, column=1, value="PEG Ratio (lower = better)")
        ws_m.cell(row=insert_row, column=2, value="Threshold:")
        ws_m.cell(row=insert_row, column=3, value="≤0.5 → 100")
        ws_m.cell(row=insert_row, column=4, value="≤1.0 → 90")
        ws_m.cell(row=insert_row, column=5, value="≤1.5 → 75")
        ws_m.cell(row=insert_row, column=6, value="≤2.0 → 60")
        ws_m.cell(row=insert_row, column=7, value="≤2.5 → 45")
        ws_m.cell(row=insert_row, column=8, value="≤3.0 → 30")
        ws_m.cell(row=insert_row, column=9, value=">3.0 → 15")
        ws_m.cell(row=insert_row, column=10, value="PEG = Fwd P/E ÷ EPS Growth %. Undefined when EPS growth ≤ 0 (excluded from avg).")
        print("✓ PEG added to Methodology sheet")

    # ── Step 6: Save ─────────────────────────────────────────────────
    wb.save(SCORING_PATH)
    print(f"✓ Saved to {SCORING_PATH}")

    # ── Step 7: Compute new scores in Python and display ─────────────
    old_weights = {"value": 0.15, "quality": 0.20, "growth": 0.15,
                   "ai_thesis": 0.30, "momentum": 0.10, "risk": 0.10}
    new_weights = {"value": 0.20, "quality": 0.20, "growth": 0.15,
                   "ai_thesis": 0.20, "momentum": 0.10, "risk": 0.15}

    old_results = compute_all_scores(all_rows, old_weights)  # without PEG in value
    new_results = compute_all_scores(all_rows, new_weights)  # with PEG in value

    # For old results, recompute value WITHOUT PEG (4 metrics only)
    for i, row in enumerate(all_rows):
        val_scores_old = [score_fwd_pe(row.get("fwd_pe")), score_ev_ebitda(row.get("ev_ebitda")),
                          score_fcf_yield(row.get("fcf_yield")), score_ps(row.get("ps"))]
        old_value = avg_non_none(val_scores_old)
        old_results[i]["value"] = old_value
        # Recompute old total
        cats = [
            (old_value,                    0.15),
            (old_results[i]["quality"],     0.20),
            (old_results[i]["growth"],      0.15),
            (old_results[i]["ai_score"],    0.30),
            (old_results[i]["momentum"],    0.10),
            (old_results[i]["risk"],        0.10),
        ]
        total_parts = [(s, w) for s, w in cats if s is not None]
        old_results[i]["total"] = sum(s * w for s, w in total_parts) if total_parts else None
        t = old_results[i]["total"]
        if t is None:
            old_results[i]["tier"] = ""
        elif t >= 85:
            old_results[i]["tier"] = "✓✓✓"
        elif t >= 70:
            old_results[i]["tier"] = "✓✓"
        elif t >= 55:
            old_results[i]["tier"] = "✓"
        elif t >= 40:
            old_results[i]["tier"] = "?"
        else:
            old_results[i]["tier"] = "✗"

    # Sort by new total descending
    paired = list(zip(new_results, old_results))
    paired.sort(key=lambda x: x[0]["total"] or 0, reverse=True)

    print("\n" + "="*100)
    print("RESCORED WATCHLIST — New Weights + PEG")
    print(f"  Value=20%  Quality=20%  Growth=15%  AI Thesis=20%  Momentum=10%  Risk=15%")
    print("="*100)
    print(f"{'Ticker':<7} {'Company':<30} {'PEG':>5} {'Val':>5} {'Qua':>5} {'Gro':>5} {'AI':>5} {'Mom':>5} {'Rsk':>5} │ {'NEW':>6} {'Tier':<4} │ {'OLD':>6} {'Tier':<4} │ {'Δ':>6}")
    print("─"*100)

    tier_changes = []
    for new, old in paired:
        if new["total"] is None:
            continue
        delta = (new["total"] or 0) - (old["total"] or 0)
        peg_str = f"{new['peg']:.1f}" if new["peg"] is not None else "  —"

        tier_changed = new["tier"] != old["tier"]
        marker = " ←" if tier_changed else ""

        print(f"{new['ticker']:<7} {new['company'][:30]:<30} {peg_str:>5} "
              f"{new['value'] or 0:5.1f} {new['quality'] or 0:5.1f} {new['growth'] or 0:5.1f} "
              f"{new['ai_score'] or 0:5.1f} {new['momentum'] or 0:5.1f} {new['risk'] or 0:5.1f} "
              f"│ {new['total']:6.1f} {new['tier']:<4} │ {old['total'] or 0:6.1f} {old['tier']:<4} │ {delta:+6.1f}{marker}")

        if tier_changed:
            tier_changes.append((new["ticker"], old["tier"], new["tier"], delta))

    print("─"*100)

    if tier_changes:
        print(f"\n📊 TIER CHANGES ({len(tier_changes)} stocks):")
        for ticker, old_t, new_t, delta in sorted(tier_changes, key=lambda x: x[3]):
            direction = "↑" if delta > 0 else "↓"
            print(f"  {ticker:<7} {old_t} → {new_t}  ({delta:+.1f}) {direction}")

    print(f"\nTotal stocks scored: {sum(1 for n, _ in paired if n['total'] is not None)}")


if __name__ == "__main__":
    main()
