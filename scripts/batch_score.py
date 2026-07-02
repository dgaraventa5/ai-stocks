"""Batch-score every ticker in the research map.

For each ticker:
  1. yfinance pull → 11 objective inputs for the Watchlist tab
  2. Append row to 00-master/ai_supply_chain_scoring.xlsx (formulas retargeted)
  3. Scaffold per-stock/{T}/ (thesis.md, news-log.md, catalysts-watchlist.md, dirs)
  4. Build per-stock/{T}/financials.xlsx with Data flags sheet

Subjective rating cells (12 of them) are left BLANK — those happen in a
collaborative rating session with the human per
templates/rating-rubric-and-workflow.md.

Saves incrementally every 5 tickers so partial failures don't lose work.

Usage:
  python3 scripts/batch_score.py                  # full universe
  python3 scripts/batch_score.py --only NVDA AMD  # explicit subset
  python3 scripts/batch_score.py --layer 06       # one layer prefix
  python3 scripts/batch_score.py --dry-run        # parse universe, don't fetch
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import re
import time
import warnings
import json
from pathlib import Path

warnings.filterwarnings("ignore")

import yfinance as yf

import adr_currency
from openpyxl import load_workbook

from common import ROOT, flag, per_stock_dir
from new_ticker import scaffold
from yfinance_fundamentals import build_financials

SCORING_PATH = ROOT / "00-master" / "ai_supply_chain_scoring.xlsx"
MAP_PATH = ROOT / "00-master" / "ai_supply_chain_research_map.md"


def parse_universe() -> list[tuple[str, str]]:
    """Extract (ticker, layer-label) tuples from the research map.

    Stops when leaving the Layer sections (avoids false positives like 'EDGAR'
    from the Tracking Framework section)."""
    text = MAP_PATH.read_text()
    universe: list[tuple[str, str]] = []
    seen: set[str] = set()
    current_layer: str | None = None
    current_sub: str | None = None

    for line in text.splitlines():
        m = re.match(r"^## Layer (\d+)\s*[—-]+\s*(.+)$", line)
        if m:
            current_layer = f"{int(m.group(1)):02d} {m.group(2).strip()}"
            current_sub = None
            continue
        # Any other ## heading exits layer parsing
        if line.startswith("## "):
            current_layer = None
            current_sub = None
            continue
        m = re.match(r"^### (.+)$", line)
        if m:
            current_sub = m.group(1).strip()
            continue
        m = re.match(r"^- \*\*[^*]+?\s*\(([A-Z]+)\)\*\*", line)
        if m and current_layer:
            ticker = m.group(1)
            if ticker in seen:
                continue
            seen.add(ticker)
            label = current_layer + (f" / {current_sub}" if current_sub else "")
            universe.append((ticker, label))

    return universe


def existing_watchlist_tickers(ws) -> tuple[set[str], int]:
    """Return (set of tickers already in Watchlist, next empty row number)."""
    existing = set()
    next_row = 2
    for r in range(2, ws.max_row + 2):
        v = ws.cell(row=r, column=1).value
        if v:
            existing.add(str(v).strip().upper())
            next_row = r + 1
        else:
            next_row = r
            break
    return existing, next_row


def _is_missing(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False


def statement_fcf(t: yf.Ticker) -> float | None:
    """TTM FCF from the cash-flow statement: sum of last 4 quarters OCF - |capex|.

    Added 2026-06-12: yfinance info.freeCashflow is Yahoo's *levered* FCF
    estimate and does not reconcile to the statements (it printed NVDA's 47%
    TTM FCF margin as 18% and MOD's positive FCF as negative). Statement-based
    is the source of truth; callers fall back to info.freeCashflow only when
    fewer than 4 quarters are available, and must flag the fallback.
    """
    try:
        cf = t.quarterly_cashflow
        ocf = cf.loc["Operating Cash Flow"].dropna()
        capex = cf.loc["Capital Expenditure"].dropna()
        if len(ocf) >= 4 and len(capex) >= 4:
            return float(ocf.iloc[:4].sum()) - abs(float(capex.iloc[:4].sum()))
    except Exception:
        pass
    return None


_CAPACITY_JSON = Path(__file__).resolve().parent.parent / "00-master" / "capacity-mw.json"


def _secured_mw(ticker: str) -> float | None:
    """Secured gross MW from 00-master/capacity-mw.json (rule 13 cohort data)."""
    try:
        data = json.loads(_CAPACITY_JSON.read_text())
        rec = data.get(ticker)
        return rec.get("secured_gross_mw") if rec else None
    except Exception:
        return None


def _is_layer10(layer: str | None) -> bool:
    return bool(layer) and str(layer).strip().startswith("10")


def _is_layer9_capacity(layer: str | None) -> bool:
    if not layer or not str(layer).strip().startswith("09"):
        return False
    l = str(layer).lower()
    return "bitcoin" in l or "neocloud" in l


def compute_inputs(ticker: str, t: yf.Ticker, info: dict,
                   layer: str | None = None) -> tuple[dict, list[str]]:
    """Return the 11 objective inputs + a list of gap messages for the log."""
    gaps: list[str] = []
    inp: dict[str, float | None] = {
        "fwd_pe": None, "ev_ebitda": None, "fcf_yield": None, "ps": None,
        "roic": None, "gross_mgn": None, "fcf_mgn": None, "nd_ebitda": None,
        "rev_3y_cagr": None, "rev_yoy": None, "eps_yoy": None,
    }

    # ----- Value -----
    inp["fwd_pe"]    = info.get("forwardPE")
    # Col F is layer-conditional (the column header still says EV/EBITDA):
    #   Layer 10 (rule 10): EV / statement FCF, SaaS bands
    #   Layer 9 capacity cohort (rule 13): EV / secured gross MW in $M/MW,
    #     MW from 00-master/capacity-mw.json
    #   everything else: EV/EBITDA
    inp["ev_ebitda"] = info.get("enterpriseToEbitda")
    mcap = info.get("marketCap")
    fcf = statement_fcf(t)
    ev = info.get("enterpriseValue")
    if ev is None and mcap:
        ev = mcap + (info.get("totalDebt") or 0) - (info.get("totalCash") or 0)
    if _is_layer10(layer):
        if ev and fcf:
            inp["ev_ebitda"] = round(ev / fcf, 1)
        else:
            inp["ev_ebitda"] = None
            gaps.append("EV/FCF (L10): missing EV or statement FCF — left blank")
    elif _is_layer9_capacity(layer):
        mw = _secured_mw(ticker)
        if ev and mw:
            inp["ev_ebitda"] = round(ev / 1e6 / mw, 1)
        else:
            inp["ev_ebitda"] = None
            gaps.append(f"EV/MW (L9 capacity): no secured_gross_mw for {ticker} in "
                        "capacity-mw.json — add it (rule 13) or score stays blank")
    if fcf is None:
        fcf = info.get("freeCashflow")
        if fcf is not None:
            gaps.append("FCF: statement TTM unavailable (<4 quarters) — fell back to "
                        "yfinance info.freeCashflow (levered estimate, known to diverge)")
    if fcf is not None and mcap and not _is_missing(fcf) and not _is_missing(mcap):
        inp["fcf_yield"] = round(fcf / mcap * 100, 2)
    else:
        gaps.append("FCF Yield: missing freeCashflow or marketCap")
    inp["ps"] = info.get("priceToSalesTrailing12Months")

    # Foreign filer whose price currency differs from its financial-reporting
    # currency: yfinance's market-cap-based ratios (EV/EBITDA, P/S, FCF-Yield)
    # mix currencies and are garbage. Fix by putting numerator and denominator in
    # the SAME currency. Prefer a mapped local listing (cleanest — for US ADRs
    # like TSM whose ADR market cap is itself corrupt); otherwise CONVERT the
    # market cap into the financial currency by exchange rate and recompute
    # (general — any foreign primary listing, e.g. SMIC in HKD). Skip Layer-10/9
    # (col F is EV/FCF or EV/MW there, not EV/EBITDA).
    trading, financial = info.get("currency"), info.get("financialCurrency")
    if adr_currency.has_currency_mismatch(trading, financial) \
            and not _is_layer10(layer) and not _is_layer9_capacity(layer):
        fixed = adr_currency.fetch_local_ratios(ticker)
        if fixed:
            src = f"local listing {adr_currency.ADR_LOCAL.get(ticker)}"
        else:
            rate = adr_currency.fx_rate(trading, financial)
            fixed = adr_currency.ratios_via_fx(
                mcap, rate, info.get("totalRevenue"), info.get("ebitda"),
                info.get("totalDebt"), info.get("totalCash"), fcf)
            src = f"FX {trading}->{financial}"
        inp["ev_ebitda"], inp["ps"], inp["fcf_yield"] = (
            fixed["ev_ebitda"], fixed["ps"], fixed["fcf_yield"])
        got = [k for k in ("ev_ebitda", "ps", "fcf_yield") if fixed[k] is not None]
        gaps.append(f"Foreign filer {trading}/{financial}: EV/EBITDA, P/S, "
                    f"FCF-Yield via {src} ({', '.join(got) or 'none available'})")

    # ----- Quality -----
    # ROIC not exposed by yfinance.info; leave blank (don't substitute ROA)
    gaps.append("ROIC: not exposed by yfinance.info — leave blank")
    gm = info.get("grossMargins")
    if gm is not None and not _is_missing(gm):
        inp["gross_mgn"] = round(gm * 100, 2)
    rev = info.get("totalRevenue")
    if fcf is not None and rev and not _is_missing(fcf) and not _is_missing(rev):
        inp["fcf_mgn"] = round(fcf / rev * 100, 2)
    debt = info.get("totalDebt"); cash = info.get("totalCash"); ebitda = info.get("ebitda")
    if (debt is not None and cash is not None and ebitda
            and not any(_is_missing(x) for x in (debt, cash, ebitda)) and ebitda > 0):
        inp["nd_ebitda"] = round((debt - cash) / ebitda, 2)
    elif ebitda is not None and not _is_missing(ebitda) and ebitda <= 0:
        gaps.append("ND/EBITDA: EBITDA non-positive — ratio would be misleading; left blank")
    else:
        gaps.append("ND/EBITDA: missing totalDebt/totalCash/ebitda")

    # ----- Growth -----
    try:
        is_a = t.income_stmt
        if is_a is not None and not is_a.empty and "Total Revenue" in is_a.index:
            rev_row = is_a.loc["Total Revenue"]
            cols = list(rev_row.index)
            if len(cols) >= 4:
                r_now = float(rev_row.iloc[0])
                r_3y = float(rev_row.iloc[3])
                if r_3y > 0 and r_now > 0:
                    inp["rev_3y_cagr"] = round((((r_now / r_3y) ** (1 / 3)) - 1) * 100, 2)
                else:
                    gaps.append("Rev 3y CAGR: non-positive base or current value")
            else:
                gaps.append(f"Rev 3y CAGR: only {len(cols)} years of data")
        else:
            gaps.append("Rev 3y CAGR: income statement missing or no Total Revenue row")
    except Exception as e:
        gaps.append(f"Rev 3y CAGR: error {e}")

    rg = info.get("revenueGrowth")
    if rg is not None and not _is_missing(rg):
        inp["rev_yoy"] = round(rg * 100, 2)
    eg = info.get("earningsQuarterlyGrowth")
    if eg is not None and not _is_missing(eg):
        inp["eps_yoy"] = round(eg * 100, 2)

    return inp, gaps


_ROW_REF_RE = re.compile(r'(?<![\$\d])([A-Z]+)2(?!\d)')


def _retarget_formula(formula: str, new_row: int) -> str:
    """Replace row-2 cell references (J2, K2, etc.) with new_row.

    Preserves: numeric literals (25, 20, *20), absolute refs (Weights!$B$2),
    multi-digit numbers (must not match the trailing "2" inside "25").
    """
    return _ROW_REF_RE.sub(lambda m: m.group(1) + str(new_row), formula)


def append_row(ws, row_num: int, ticker: str, layer: str, inputs: dict) -> None:
    """Append a Watchlist row with values + retargeted formulas. Subjective cells blank."""
    SRC_ROW = 2
    # PEG(9), ValueScore(10), QualityScore(15), GrowthScore(19),
    # AIScore(25), MomentumScore(29), RiskScore(34), TOTAL(35), Tier(36)
    SCORE_COLS = [9, 10, 15, 19, 25, 29, 34, 35, 36]
    for col in SCORE_COLS:
        src = ws.cell(row=SRC_ROW, column=col).value
        if isinstance(src, str) and src.startswith("="):
            ws.cell(row=row_num, column=col, value=_retarget_formula(src, row_num))

    cells = {
        1: ticker,
        2: None,  # Company name filled below
        3: layer,
        4: dt.date.today().isoformat(),
        5:  inputs["fwd_pe"],
        6:  inputs["ev_ebitda"],
        7:  inputs["fcf_yield"],
        8:  inputs["ps"],
        # col 9 (PEG) is a formula — populated via SCORE_COLS
        11: inputs["roic"],
        12: inputs["gross_mgn"],
        13: inputs["fcf_mgn"],
        14: inputs["nd_ebitda"],
        16: inputs["rev_3y_cagr"],
        17: inputs["rev_yoy"],
        18: inputs["eps_yoy"],
    }
    for col, val in cells.items():
        ws.cell(row=row_num, column=col, value=val)


def process_ticker(ticker: str, layer: str, ws, row_num: int) -> tuple[str, list[str], str]:
    """Returns (status, gaps, company_name)."""
    try:
        t = yf.Ticker(ticker)
        info = t.info or {}
    except Exception as e:
        return ("error_yfinance", [f"yfinance error: {e}"], "")

    company = info.get("longName") or info.get("shortName") or ""
    if not company:
        return ("error_no_data", ["yfinance returned empty info"], "")

    inputs, gaps = compute_inputs(ticker, t, info, layer=layer)
    append_row(ws, row_num, ticker, layer, inputs)
    ws.cell(row=row_num, column=2, value=company)

    # Scaffold per-stock dir (idempotent — skips if files already exist)
    try:
        scaffold(ticker)
    except Exception as e:
        gaps.append(f"scaffold failed: {e}")

    # Build financials.xlsx — pass the same Ticker to avoid a second .info call
    try:
        # build_financials() instantiates its own yf.Ticker, but Ticker is cheap.
        # The .info pull will hit cache on yfinance side within same process.
        build_financials(ticker)
    except Exception as e:
        gaps.append(f"financials.xlsx failed: {e}")

    return ("ok", gaps, company)


def run(tickers: list[tuple[str, str]], dry_run: bool = False,
        save_every: int = 5) -> dict:
    """Returns a dict summary of the run."""
    if dry_run:
        print(f"DRY RUN — would process {len(tickers)} tickers:")
        for t, l in tickers:
            print(f"  {t:<6} {l}")
        return {}

    wb = load_workbook(SCORING_PATH)
    ws = wb["Watchlist"]
    existing, next_row = existing_watchlist_tickers(ws)

    todo = [(t, l) for t, l in tickers if t not in existing]
    print(f"Universe: {len(tickers)} | Existing: {len(existing)} | To process: {len(todo)}")
    print(f"Starting at Watchlist row {next_row}")
    print()

    log: list[dict] = []
    started = time.monotonic()

    for i, (ticker, layer) in enumerate(todo, start=1):
        t0 = time.monotonic()
        status, gaps, company = process_ticker(ticker, layer, ws, next_row)
        elapsed = time.monotonic() - t0
        log.append(dict(
            ticker=ticker, layer=layer, status=status, gaps=gaps,
            company=company, row=next_row if status == "ok" else None,
            seconds=round(elapsed, 1),
        ))

        marker = "✓" if status == "ok" else "✗"
        gap_n = len([g for g in gaps if not g.startswith("ROIC:")])  # ROIC gap is universal
        print(f"  [{i:>3}/{len(todo)}] {marker} {ticker:<6} {company[:35]:<35} "
              f"{elapsed:5.1f}s  gaps={gap_n}  {status}")

        if status == "ok":
            next_row += 1

        if i % save_every == 0:
            wb.save(SCORING_PATH)
            elapsed_total = time.monotonic() - started
            rate = i / elapsed_total
            eta = (len(todo) - i) / rate if rate > 0 else 0
            print(f"        [checkpoint saved, ~{eta:.0f}s remaining]")

        time.sleep(0.3)  # polite throttle

    wb.save(SCORING_PATH)
    elapsed_total = time.monotonic() - started
    print(f"\nDone in {elapsed_total:.0f}s.")

    return {
        "total": len(todo),
        "ok": sum(1 for r in log if r["status"] == "ok"),
        "errors": sum(1 for r in log if r["status"].startswith("error")),
        "log": log,
        "elapsed_seconds": round(elapsed_total, 1),
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", nargs="+", help="Specific tickers")
    ap.add_argument("--layer", help="Layer prefix filter, e.g., '06'")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    universe = parse_universe()
    if args.only:
        wanted = {t.upper() for t in args.only}
        universe = [(t, l) for t, l in universe if t in wanted]
    if args.layer:
        universe = [(t, l) for t, l in universe if l.startswith(args.layer)]

    summary = run(universe, dry_run=args.dry_run)

    if summary:
        # Append a summary file for the run
        out = ROOT / "tracking" / f"batch-score-{dt.date.today().isoformat()}.md"
        with out.open("w") as f:
            f.write(f"# Batch score run — {dt.date.today().isoformat()}\n\n")
            f.write(f"- Total processed: **{summary['total']}**\n")
            f.write(f"- Succeeded: **{summary['ok']}**\n")
            f.write(f"- Errors: **{summary['errors']}**\n")
            f.write(f"- Elapsed: **{summary['elapsed_seconds']}s**\n\n")
            f.write("## Per-ticker log\n\n")
            f.write("| Ticker | Company | Row | Status | Gaps | Seconds |\n")
            f.write("|---|---|---:|---|---|---:|\n")
            for r in summary["log"]:
                gap_short = "; ".join(g for g in r["gaps"] if not g.startswith("ROIC:"))[:90] or "—"
                f.write(f"| {r['ticker']} | {r['company'][:40]} | "
                        f"{r['row'] or '—'} | {r['status']} | {gap_short} | {r['seconds']} |\n")
        print(f"Summary written to {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
