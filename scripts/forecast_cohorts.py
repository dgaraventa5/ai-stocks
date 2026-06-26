"""Build a frozen equal-weight peer basket for a ticker from the Watchlist.

The peer list is frozen into resolution_rule at forecast creation so quarterly
roster churn can't retroactively change the benchmark (no membership look-ahead).
Layers with fewer than MIN_PEERS peers fall back to SMH.

openpyxl only (CI-safe). See design spec §6/§7.
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCORING_PATH = ROOT / "00-master" / "ai_supply_chain_scoring.xlsx"

TICKER_COL = 1
LAYER_COL = 3
HORIZON_TD = 63
MIN_PEERS = 6   # fewer same-layer peers than this -> SMH fallback


def read_watchlist_rows(scoring_path: Path = SCORING_PATH) -> list[tuple[str, str]]:
    """(ticker, layer_num) for every Watchlist row. layer_num = first token of col C."""
    ws = load_workbook(scoring_path, read_only=True)["Watchlist"]
    assert ws.cell(row=1, column=TICKER_COL).value == "Ticker", "Watchlist col A is not Ticker"
    assert ws.cell(row=1, column=LAYER_COL).value == "Layer", "Watchlist col C is not Layer"
    rows = []
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=TICKER_COL).value
        layer = ws.cell(row=r, column=LAYER_COL).value
        if t and layer:
            rows.append((str(t).strip().upper(), str(layer).strip().split()[0]))
    return rows


def build_frozen_cohort(ticker: str, scoring_path: Path = SCORING_PATH,
                        rows: list[tuple[str, str]] | None = None) -> tuple[str, dict]:
    """Return (layer_num, frozen benchmark rule) for `ticker`.

    rule = {"benchmark","constituents","horizon_td"} — a layer EW basket, or the
    SMH fallback when the layer is too thin. `rows` is injectable for tests.
    """
    ticker = ticker.upper()
    rows = rows if rows is not None else read_watchlist_rows(scoring_path)
    layer = next((lay for t, lay in rows if t == ticker), None)
    if layer is None:
        raise ValueError(f"{ticker} not found on Watchlist")
    peers = sorted({t for t, lay in rows if lay == layer and t != ticker})
    if len(peers) < MIN_PEERS:
        return layer, {"benchmark": "SMH", "constituents": ["SMH"], "horizon_td": HORIZON_TD}
    return layer, {"benchmark": "layer_cohort_ew", "constituents": peers, "horizon_td": HORIZON_TD}
