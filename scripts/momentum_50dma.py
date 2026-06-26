"""Compute '% of last 120 trading days with close > 50-day SMA' per ticker.

The objective Momentum metric added 2026-06-09 (Watchlist col AC, '50DMA %').
Run after price-moving events or as part of the quarterly rescore; values go
stale the same way any momentum input does.

Per CLAUDE.md §Subagent patterns: yfinance calls are serialized with a small
sleep to avoid throttling. Per §3 (Flag, don't assume): tickers with <60 days
of valid SMA comparisons (recent IPOs) are flagged and left blank; partial
windows of 60-119 days are used but flagged.

Usage:
  python3 scripts/momentum_50dma.py                # all Watchlist tickers, write
  python3 scripts/momentum_50dma.py NVDA AMD       # subset
  python3 scripts/momentum_50dma.py --dry-run      # compute + print, no write
"""
from __future__ import annotations

import argparse
import datetime as dt
import time
from pathlib import Path

import yfinance as yf
from openpyxl import load_workbook

from common import flag

_REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = str(_REPO_ROOT / '00-master/ai_supply_chain_scoring.xlsx')
DMA_COL = 29          # AC — '50DMA %'
WINDOW = 120          # trading days scored
SMA = 50              # moving-average length
MIN_DAYS = 60         # below this, leave blank + flag
MAX_STALE_DAYS = 7    # last trade older than this → halted/delisted, leave blank


def pct_days_above_50dma(ticker: str) -> float | None:
    """Share (0-100) of the last WINDOW trading days with close > 50-day SMA."""
    hist = yf.Ticker(ticker).history(period='13mo', auto_adjust=True)
    if hist is None or hist.empty:
        flag(f'{ticker}: no price history on yfinance')
        return None
    # Freshness gate (added after CTRA, delisted 2026-05, scored 70.0 off its
    # pre-delisting tail): a series that quietly stopped updating still has
    # >MIN_DAYS of history, so the length check alone doesn't catch it.
    last = hist.index[-1].date()
    age = (dt.date.today() - last).days
    if age > MAX_STALE_DAYS:
        flag(f'{ticker}: last trade {last} ({age}d ago) — halted/delisted? left blank')
        return None
    close = hist['Close'].dropna()
    sma = close.rolling(SMA).mean()
    valid = (close > sma)[sma.notna()].tail(WINDOW)
    if len(valid) < MIN_DAYS:
        flag(f'{ticker}: only {len(valid)} valid days (<{MIN_DAYS}) — left blank')
        return None
    if len(valid) < WINDOW:
        flag(f'{ticker}: partial window, {len(valid)}/{WINDOW} days (recent listing?)')
    return round(float(valid.mean()) * 100, 1)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('tickers', nargs='*', help='default: all Watchlist tickers')
    ap.add_argument('--dry-run', action='store_true', help='compute but do not write')
    args = ap.parse_args()

    wb = load_workbook(XLSX, data_only=False)
    ws = wb['Watchlist']
    assert ws.cell(row=1, column=DMA_COL).value == '50DMA %', \
        'Watchlist col AC is not 50DMA % — schema drift, abort'

    rows = {ws.cell(row=r, column=1).value: r
            for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}
    targets = [t.upper() for t in args.tickers] if args.tickers else list(rows)
    unknown = [t for t in targets if t not in rows]
    if unknown:
        flag(f'not on Watchlist, skipped: {", ".join(unknown)}')
        targets = [t for t in targets if t in rows]

    wrote = 0
    print(f'{"Ticker":<8}{"50DMA %":>9}')
    for t in targets:
        try:
            pct = pct_days_above_50dma(t)
        except Exception as e:
            flag(f'{t}: yfinance error — {e}')
            pct = None
        print(f'{t:<8}{pct if pct is not None else "—":>9}')
        if pct is not None and not args.dry_run:
            ws.cell(row=rows[t], column=DMA_COL).value = pct
            wrote += 1
        time.sleep(0.3)  # serialize, per CLAUDE.md

    if not args.dry_run:
        wb.save(XLSX)
        print(f'wrote {wrote}/{len(targets)} values to {XLSX}')
    else:
        print('(dry run — nothing written)')


if __name__ == '__main__':
    main()
