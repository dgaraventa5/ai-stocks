"""Append validated forecasts to tracking/forecasts.jsonl (CLAUDE.md rule 17).

Phase 1 seeds the portfolio in one pass:

  python3 scripts/log_forecast.py --seed-portfolio --dry-run        # propose + print
  python3 scripts/log_forecast.py --seed-portfolio --apply \
        --overrides /path/overrides.json                            # write (Dom's probs)

overrides.json: {"AVGO": 0.60, "NVDA": 0.72}  (ticker -> probability; the override
is the training signal, not the rating-derived default).

openpyxl only (no yfinance): reads ratings + builds the frozen cohort; never
fetches prices. See design spec §7.
"""
from __future__ import annotations

import sys
import json
import argparse
import datetime as dt
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

import forecast_store as store
import forecast_cohorts as cohorts
from forecast_resolvers import resolution_date_for
from refresh_objective_inputs import resolve_targets, SCORING_PATH, PORTFOLIO_PATH

DIMENSION = "momentum.rel_strength"
TEMPLATE = "REL_STRENGTH_1Q"
HORIZON_TD = 63
RELSTR_COL = 27                              # AA — 'Rel Str' (momentum subjective rating)
RELSTR_HEADER = "Rel Str"
DEFAULT_PROB = {5: 0.80, 4: 0.65, 3: 0.55, 2: 0.42, 1: 0.28}
BASE_RATE_PROB = 0.55


def _relstr_ratings(scoring_path=SCORING_PATH) -> dict:
    ws = load_workbook(scoring_path, read_only=True)["Watchlist"]
    assert ws.cell(1, RELSTR_COL).value == RELSTR_HEADER, \
        f"Watchlist col {RELSTR_COL} is not {RELSTR_HEADER!r} — schema drift, abort"
    out = {}
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, 1).value
        if not t:
            continue
        v = ws.cell(r, RELSTR_COL).value
        out[str(t).strip().upper()] = int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None
    return out


def _claim(ticker: str, rule: dict, layer: str) -> str:
    if rule["benchmark"] == "SMH":
        return f"{ticker} total return outperforms SMH over the next {rule['horizon_td']} trading days"
    return (f"{ticker} total return outperforms its frozen Layer-{layer} equal-weight "
            f"peer basket over the next {rule['horizon_td']} trading days")


def build_proposals(today: dt.date, scoring_path=SCORING_PATH, portfolio_path=PORTFOLIO_PATH) -> list[dict]:
    names = resolve_targets("portfolio", scoring_path=scoring_path, portfolio_path=portfolio_path)
    ratings = _relstr_ratings(scoring_path)
    rows = cohorts.read_watchlist_rows(scoring_path)
    res_date = resolution_date_for(today, HORIZON_TD).isoformat()
    proposals = []
    for t in names:
        layer, rule = cohorts.build_frozen_cohort(t, rows=rows)
        rating = ratings.get(t)
        prob = DEFAULT_PROB.get(rating, BASE_RATE_PROB)
        if rating is None:
            store.flag(f"{t}: no Relative Strength rating — default {BASE_RATE_PROB}")
        proposals.append(dict(
            ticker=t, layer=layer, dimension=DIMENSION, rating_value=rating,
            template=TEMPLATE, claim=_claim(t, rule, layer), probability=prob,
            resolution_date=res_date, resolution_rule=rule, status="open"))
    return proposals


def print_table(proposals: list[dict]) -> None:
    print(f'{"Ticker":<7}{"Layer":<6}{"Rating":>7}{"Prob":>7}  Benchmark')
    for p in proposals:
        rating = p["rating_value"] if p["rating_value"] is not None else "—"
        rule = p["resolution_rule"]
        print(f'{p["ticker"]:<7}{p["layer"]:<6}{str(rating):>7}{p["probability"]:>7.2f}  '
              f'{rule["benchmark"]} ({len(rule["constituents"])})')
    if proposals:
        print(f'\nresolution_date: {proposals[0]["resolution_date"]}')


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--seed-portfolio", action="store_true",
                    help="propose one REL_STRENGTH_1Q forecast per portfolio name")
    ap.add_argument("--dry-run", action="store_true", help="print proposals, write nothing")
    ap.add_argument("--apply", action="store_true", help="write proposals to the log")
    ap.add_argument("--overrides", help="JSON file mapping ticker -> probability")
    args = ap.parse_args()

    if not args.seed_portfolio:
        ap.error("Phase 1 supports only --seed-portfolio")
    if args.dry_run == args.apply:
        ap.error("choose exactly one of --dry-run / --apply")

    today = dt.date.today()
    proposals = build_proposals(today)

    if args.overrides:
        ov = {k.upper(): float(v) for k, v in json.loads(Path(args.overrides).read_text()).items()}
        for p in proposals:
            if p["ticker"] in ov:
                p["probability"] = ov[p["ticker"]]

    print_table(proposals)
    if args.dry_run:
        print("\n(dry run — nothing written)")
        return

    for p in proposals:
        snap = store.append_creation(p, today=today)
        print(f'logged {snap["id"]}  p={snap["probability"]:.2f}')
    print(f'\nwrote {len(proposals)} forecasts to {store.FORECASTS_PATH}')


if __name__ == "__main__":
    main()
