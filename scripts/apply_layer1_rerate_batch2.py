"""Apply Layer 1 directness re-ratings — batch 2, the remaining 19 names (2026-06-10).

Completes the Layer 1 AI-Thesis re-rating begun with the 9 portfolio-relevant
names. Research = per-stock/{T}/context-2026-06-{09,10}.md briefings, extracted
by 5 bucket agents; final ratings reconciled by the orchestrator to ONE
consistent standard across all 19 (the agents each saw only their bucket).

Consistent standard applied:
  D1 (current AI revenue %): pre-revenue / commodity / regulated-load-not-yet-
     flowing = 1; measurable-but-minority current DC revenue = 2; majority = 5.
  D2 (pricing power): regulated capped-ROE = 3; commodity price-taker = 2;
     oligopoly term-pricing = 4; dispatchable-scarcity bottleneck = 5.
  D5 (hyperscaler): signed+named+disclosed-terms = 4-5; signed+named but
     regulated-economics/undisclosed = 3; LOI/MOU/narrative = 2; none = 1.

Orchestrator overrides of bucket-agent proposals, for cross-name consistency
(flagged to Dom): NEE D1 3->2, NEE D2 5->4, XEL D1 hold->1, XEL D5 hold->3.
Only D1/D2/D5 touched; D3/D4 untouched (BWXT D3=5 noted off-thesis for a later
pass). NXT unchanged (already lens-correct 2026-06-05).
"""
from __future__ import annotations

from openpyxl import load_workbook

XLSX = '/Users/dom/Desktop/ai-stocks/00-master/ai_supply_chain_scoring.xlsx'
DATE = '2026-06-10'
COL = {'D1': 20, 'D2': 21, 'D5': 24}

# ticker -> dim -> (new, rationale, source, conf)
C = {
 'D': {
  'D1': (2, "DC = 28% of DEV electricity SALES VOLUME FY2025 (not revenue, mostly NoVA cloud/enterprise); AI-specific share of consolidated D ~5-10%. Highest current-DC of the regulated cohort.", 'D 10-K FY2025 2026-02-23', 'Med'),
  'D2': (3, "Regulated, allowed ROE 9.8%; GS-5 minimum-demand terms capture volume/rate-base, not scarcity pricing.", 'Virginia SCC order 2025-11-25', 'High'),
  'D5': (3, "51 GW pipeline (10.4 GW ESA-stage) but ZERO named hyperscaler counterparties, tariff-priced; only an exploratory Amazon SMR MOU. Large but not concentrated/premium.", 'D Q1 2026 slides 2026-05-01', 'High')},
 'AEP': {
  'D1': (1, "No disclosed current DC revenue/%; the 63 GW is FUTURE load by 2030. Current AI revenue unclear/<5%.", 'AEP Q1 2026 8-K 2026-05-05', 'High'),
  'D2': (3, "Earned ROE 9.1% 2025 (below allowed); Ohio take-or-pay de-risks recovery but $16B earmarked as cost offsets for existing customers. Volume x capped ROE.", 'AEP Q1 2026 call 2026-05-05', 'High'),
  'D5': (3, "63 GW contracted incremental (~90% DC) is the largest pipeline and names some campuses (SB Energy/Google/Amazon), but no signed hyperscaler ESAs with disclosed terms; 41 GW are softer TX letters of agreement. FLAG: most debatable, 4 defensible on named counterparties.", 'AEP 8-K+call 2026-05-05', 'Med')},
 'SO': {
  'D1': (1, "No hard DC %; est ~3-5% of retail kWh / low-single-digit % revenue today; $556M guaranteed large-load rev at 2028 ~= 1.9% of FY25. Unclear/<5%.", 'SO 10-K FY2025; Q1 call 2026-04-30', 'High'),
  'D2': (3, "GA Power ROE set 10.5%, base rates frozen through 2028; Dec 2025 stipulation routes large-load revenue to CUT residential bills -- DC economics shared with ratepayers by design.", 'SO 10-K FY2025; GA PSC 2026-03', 'High'),
  'D5': (3, "11+ GW contracted across 28 projects but ZERO named hyperscaler counterparties (confidential GA PSC filings), tariff-priced; only ~1.9 GW executed; pipeline shrank 6 GW Q2->Q3 2025.", 'SO Q1 2026 slides 2026-04-30', 'High')},
 'NEE': {
  'D1': (2, "Hybrid; Energy Resources holds signed hyperscaler PPAs (Google ~3.5 GW, Meta ~2.5 GW) but <2 GW DELIVERING today = low-single-digit % of consolidated revenue. Highest of regulated cohort but current share is minority. [Override: agent proposed 3; set 2 for consistency with VST/TLN at similar current %.]", '8-K Ex99 2026-04-23; Latitude Media 2025-12-08', 'Med'),
  'D2': (4, "Competitive renewables with firming pricing (+$20/MWh recontracting, bilateral hubs, 4 GW turbine slots) -- strong but not a dispatchable-scarcity bottleneck like the nuclear IPPs (D2=5). [Override: agent held 5; set 4 to rank below VST/TLN/CEG.]", 'NEE Q1 2026 call 2026-04-23', 'Med')},
 'ETR': {
  'D1': (1, "The flagged mis-rating. Fully regulated; DC load grows rate base at allowed ROE, not revenue share; upper-bound ~3-4% of Q1 rev. Was 5 (>50%) -- unsupportable.", 'ETR 8-K Ex99.1 2026-04-29', 'High'),
  'D2': (3, "'Fair Share Plus' pledges surplus to EXISTING customers ($2.65B Meta LA, >$2B AWS MS); ELL earned ROE in authorized range; zero per-MWh upside.", 'ETR 8-K 2026-04-29; The Lens 2026-04-14', 'Med'),
  'D5': (3, "Meta Hyperion IS signed/named (15-yr ESA, LPSC-approved Aug 2025, ~2.3 GW) + Google/AWS named -- but terms confidential and ETR keeps only REGULATED return, no scarcity upside. Named deal != shareholder captures it.", 'ETR 8-K 2026-04-29', 'Med')},
 'PPL': {
  'D1': (1, "Regulated holdco; only ~0.6 GW DC online in PA 2026, large-load tariff effective 2026-07-01; the 28.3 GW PA pipeline is forward. DC-tied rev low-single-digit.", 'PPL Q1 8-K 2026-05-08', 'Med'),
  'D2': (3, "Regulated capped ROE both legs (KY 9.775%); PA T&D earns wires return only -- PA energy/capacity scarcity upside accrues to merchant gens (TLN/VST/CEG), not PPL.", 'KPSC Feb 2026; FERC Op.594 2026-03-19', 'Med'),
  'D5': (3, "The 28.3 GW 'signed' are INTERCONNECTION/wires developer agreements, not power-supply contracts; AWS is one named hyperscaler; the Blackstone JV (only merchant leg) has ZERO signed ESAs 10 months on.", 'PPL Q1 8-K 2026-05-08; Utility Dive 2026-05-11', 'Med')},
 'XEL': {
  'D1': (1, "Fully regulated; only ONE DC operating in territory; Q1 C&I growth driven by Permian oil&gas electrification, not DC; Google Pine Island pre-approval; sales ramp 2029-30+. [Override: agent held 3; set 1 for cohort consistency.]", 'XEL Q1 2026 8-K; POWER 2026-02', 'Med'),
  'D5': (3, "Google ESA named+signed (NSP-MN, 15-yr) but regulated economics (Google funds all costs, benefits flow to other ratepayers). [Override: agent held 4; set 3 to match ETR's identical signed-named-but-regulated profile.]", 'XEL Q1 2026 8-K; newsroom 2026-02-24', 'Med')},
 'BE': {
  'D1': (5, "Related-party (Brookfield JV incl. major hyperscaler project) = $373.3M of $751.1M Q1 2026 rev = 49.7% floor; US mix 56%->91% YoY, product rev +208% 'driven by the AI data center industry.' DC is the MAJORITY of the business. FLAG: inference (no disclosed DC line); 4 if strict.", 'BE 10-Q Q1 2026; Q4 PR 2026-02-05', 'Med')},
 'CMI': {
  'D1': (2, "Datacenter ~$5B 2026E / ~$36.9B FY26E ~= 13-14% IF the Analyst-Day figure (2nd-hand) is accepted; 2 on verified disclosure. ~86% non-DC truck/components co; old D1=4 too high. FLAG: verify $5B vs primary IR deck (2 vs 3).", 'Cummins Analyst Day 2026-05-21; Q1 8-K 2026-05-05', 'Med'),
  'D5': (2, "No named hyperscaler customers and no disclosed $ backlog in 8-K/call/Analyst Day; materiality disclosed, identity/backlog not.", 'CMI Q1 8-K; Analyst Day 2026-05-21', 'High')},
 'GNRC': {
  'D1': (2, "Datacenter ~$75-150M of $1.06B Q1 ~= 7-14%; FY26 ~8-12%; still 52% residential. Old D1=4 (25-50%) too high.", 'GNRC Q1 8-K Ex99.1 2026-04-29', 'Med'),
  'D2': (3, "Wins on AVAILABILITY not price (~30-wk lead times vs CAT/CMI 100+ wks); ASP in-line, no premium; new entrant vs CAT/CMI; China-owned engine = tariff risk peers lack.", 'Hunterbrook 2026-02-23', 'Med'),
  'D5': (3, ">$700M disclosed DC backlog + a signed global hyperscale supply agreement (2026-06-02) but customer UNNAMED, no $/volume; best quantified backlog short of named customers.", 'GNRC Q1 call 2026-04-29; PR 2026-06-02', 'Med')},
 'PLUG': {
  'D1': (1, "Q1 2026 10-Q revenue disaggregation has NO datacenter category; DC-attributable rev ~0%. The 'DC deal' (Stream) is a one-time land/substation ASSET sale, not power solutions -- the relationship is inverted.", 'PLUG 10-Q Q1 2026; 8-K 2026-02-26', 'High')},
 'CCJ': {
  'D1': (1, "FY2025 rev 82.5% uranium + 16% fuel services to nuclear utilities; ZERO 'AI/data center/hyperscaler' in filings; AI link is 3rd-order commodity. Westinghouse $80B is equity-accounted, government counterparty, $0 current rev.", 'CCJ FY2025 6-K 2026-02-13; Q1 6-K 2026-05-05', 'High'),
  'D2': (4, "Tier-1 term-book uranium oligopolist with supply discipline (~230M lb committed) = real firming pricing; but realized $66/lb lags term $94 by ~30%, and it is not an enrichment-style monopoly.", 'CCJ FY2025 6-K; UxC 2026-05-31', 'Med'),
  'D5': (2, "No hyperscaler buys uranium/reactors under any disclosed deal; the Google-Westinghouse partnership is REVERSED (Westinghouse is Google Cloud's AI customer). Above a pure miner only via Westinghouse reactor proximity.", 'Data Center Knowledge 2025-11-20', 'High')},
 'UEC': {
  'D1': (1, "9M FY26 rev $20.2M from ONE spot inventory sale; Q3 FY26 rev $0. Zero AI/DC contract; AI appears only as macro-tailwind framing. Most indirect of the nuclear-fuel names.", 'UEC 10-Q 2026-04-30; Q3 PR 2026-06-09', 'High'),
  'D2': (2, "100% unhedged, no term contracts/floors/escalators -- maximum commodity beta by design; opportunistic timing != pricing power. Weakest pricing profile of the three.", 'UEC Q3 PR 2026-06-09', 'High'),
  'D5': (1, "Zero 'hyperscale' in 10-Q; no supply agreements/MOUs/partnerships with any hyperscaler/DC operator; conversion facility pre-license-application.", 'UEC 10-Q 2026-04-30', 'High')},
 'LEU': {
  'D1': (1, "Q1 2026 rev $76.7M = LEU segment (commodity SWU to existing fleet) + Technical Solutions (overwhelmingly the DOE HALEU contract). Government != AI; commercial HALEU is post-2029 and intermediated by SMR developers.", 'LEU Q1 2026 release 2026-05-05', 'High'),
  'D2': (4, "HALEU bottleneck is real (only NRC-licensed/operating US producer) but pricing NOT yet monetized at market (DOE fixed-price task orders), and DOE funded a 2nd enricher (General Matter). Bottleneck without demonstrated pricing power = 4.", 'ANS 2026-01; Q1 call 2026-05-07', 'Med'),
  'D5': (3, "Oklo JV deconversion discussions (pre-definitive) at Piketon adjacent to Oklo's DC-oriented campus = most concrete AI vector of the three; X-energy/TerraPower prospective, nothing contracted.", 'Oklo/Centrus PR 2026-03-09', 'High')},
 'BWXT': {
  'D1': (2, "Naval/DOE/special-materials = 67% of Q1 rev (zero AI); the 33% commercial segment is mostly Kinectrics/medical/CANDU; only the BWRX-300 SMR-component slice is indirectly AI-adjacent, single-digit % and undisclosed.", 'BWXT Q1 8-K 2026-05-04', 'High'),
  'D2': (3, "Naval sole-source is genuine but OFF-THESIS; the AI-relevant position is narrower -- sole RPV supplier for the first BWRX-300, but the SMRs hyperscalers actually picked (Oklo/Kairos/X-energy) don't anchor on BWXT. Lens reframe, not a fundamentals call.", 'BWXT GE Vernova supplier group; Darlington RPV', 'Med'),
  'D5': (2, "BWXT appears in NONE of the ~13 announced hyperscaler nuclear deals; CEO on Q1 call: 'I don't know... I suspect that we do.' No named/signed/disclosed-term hyperscaler deal.", 'Benzinga Q1 call 2026-05-04', 'High')},
 'SMR': {
  'D1': (1, "Q1 2026 revenue $565K total (effectively pre-revenue, down from $13.4M as RoPower FEED rolled off); all historical revenue was the RoPower Romania government project, zero DC/AI.", 'NuScale 10-Q Q1 2026 2026-05-07', 'High')},
 'OKLO': {
  'D1': (1, "ZERO revenue (Q1 2026 10-Q has no revenue line; net loss $33.1M). Cannot have AI revenue % of ~zero. Old D1=5 credited pipeline as revenue -- the clearest single error in the layer. First power-sale target 2028+.", 'OKLO 10-Q Q1 2026 2026-05-12', 'High'),
  'D5': (3, "Switch '12 GW' and Equinix 500 MW are NON-BINDING LOIs; Meta 1.2 GW is a signed/funded Prepayment Agreement with a specific site (keeps D5 above the pure-LOI floor) but NOT a disclosed-term PPA, timeline 2030-34. Was 5 (rated pipeline as signed). FLAG: 2 if bar is a binding PPA.", 'OKLO 10-Q 2026-05-12; Meta PR 2026-01-09', 'Med')},
 'NNE': {
  'D1': (1, "10-Q: 'have not generated any material revenue... through March 31, 2026.' STS acquisition adds logistics revenue (not AI). Old D1=4 unsupportable for a zero-revenue company.", 'NNE 10-Q Q2 FY26 2026-05-14', 'High'),
  'D5': (2, "AI-datacenter exposure is 100% narrative/MOU (Supermicro/EHC/DS Dansuk, all non-binding); a binding deal would file 8-K Item 1.01 -- none has. Named MOUs keep it above pure vapor.", 'NNE Q2 PR 2026-05-14; Supermicro MOU 2026-05-06', 'High')},
}


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)
    ws = wb['Watchlist']
    audit = wb['Rating Audit']
    rows = {ws.cell(row=r, column=1).value: r
            for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}

    applied = 0
    for tkr, dims in C.items():
        if tkr not in rows:
            print(f'!! {tkr} not on Watchlist — skipped'); continue
        r = rows[tkr]
        for dim, (new, rationale, source, conf) in dims.items():
            c = COL[dim]
            old = ws.cell(row=r, column=c).value
            ws.cell(row=r, column=c).value = new
            audit.append([DATE, tkr, dim, new, f'(was {old}) {rationale}',
                          source, conf, 'Update'])
            print(f'{tkr:<5} {dim}: {old} -> {new}')
            applied += 1
        ws.cell(row=r, column=4).value = DATE

    wb.save(XLSX)
    print(f'\napplied {applied} changes across {len(C)} names (NXT unchanged)')


if __name__ == '__main__':
    main()
