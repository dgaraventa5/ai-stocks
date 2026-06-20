"""Export friend-safe site data — THE privacy boundary for the portfolio site.

Reads the workbooks and tracking files, scales every dollar figure to the
$10,000 notional base, and writes site/data/*.json. Nothing under site/
reads repo data except through this script, so what friends can see is
auditable here. EXCLUDED BY DESIGN: real share counts, real dollar values,
cost basis, real capital.

Usage: python3 scripts/export_site_data.py [--root DIR]
Fails loudly (exit 1) on schema surprises; warns (stderr) on soft gaps.
"""
from __future__ import annotations

import datetime as dt
import json
import re
import sys
from pathlib import Path

import openpyxl

NOTIONAL = 10_000.0

LAYERS = {
    '01': 'Power Generation', '02': 'Grid Equipment', '03': 'Data Centers',
    '04': 'Semi Equipment', '05': 'Fabs & Foundries', '06': 'Silicon',
    '07': 'Optical Networking', '08': 'Servers', '09': 'Cloud',
    '10': 'Applications',
}

TARGETS_HEADERS = ['Ticker', 'Layer', 'TOTAL', 'Tier', 'Rank', 'Status',
                   'Include?', 'Override', 'Target %', 'Notes']


def fail(msg: str) -> None:
    print(f'EXPORT ERROR: {msg}', file=sys.stderr)
    raise SystemExit(1)


def warn(msg: str) -> None:
    print(f'WARN: {msg}', file=sys.stderr)


def _company_names(root: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(
        root / '00-master' / 'ai_supply_chain_scoring.xlsx', data_only=True)
    ws = wb['Watchlist']
    return {r[0]: r[1] for r in ws.iter_rows(min_row=2, values_only=True)
            if r[0]}


def export_positions(root: Path) -> list[dict]:
    wb = openpyxl.load_workbook(root / '00-master' / 'portfolio.xlsx',
                                data_only=True)
    ws = wb['Targets']
    headers = [c.value for c in ws[2]][:10]
    if headers != TARGETS_HEADERS:
        fail(f'Targets headers changed: {headers}')
    names = _company_names(root)
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        include = str(row[6]).strip().upper() if row[6] is not None else ''
        if not row[0] or include != 'Y':
            continue
        layer_num = str(row[1])[:2]
        if layer_num not in LAYERS:
            warn(f'{row[0]}: unrecognized layer {row[1]!r}')
        if row[8] is None:
            fail(f'{row[0]}: included row has no Target %')
        out.append({
            'ticker': row[0],
            'company': names.get(row[0], row[0]),
            'layer_num': layer_num,
            'layer': LAYERS.get(layer_num, str(row[1])),
            'weight': round(float(row[8]), 2),
            'notional': round(float(row[8]) / 100 * NOTIONAL),
            'score': round(float(row[2]), 1),
            'tier': row[3],
        })
    if not out:
        fail('no included positions found in Targets')
    return out


def export_performance(root: Path) -> dict:
    sp = root / 'tracking' / 'performance-series.json'
    if not sp.exists():
        fail('tracking/performance-series.json missing — run '
             'scripts/track_performance.py first')
    raw = json.loads(sp.read_text())
    cfg = json.loads(
        (root / 'tracking' / 'performance-config.json').read_text())
    for key in ('dates', 'model', 'bench', 'as_of'):
        if key not in raw:
            fail(f'performance-series.json missing key {key!r}')
    for b in ('SMH', 'QQQ', 'SPY', 'EW'):
        if b not in raw['bench']:
            fail(f'performance-series.json bench missing {b!r}')
    if not raw['model']:
        fail('performance-series.json has an empty series')
    if 'capital' not in cfg:
        fail('performance-config.json missing capital')
    capital = float(cfg['capital'])
    k = NOTIONAL / capital
    model = [round(v * k, 2) for v in raw['model']]
    bench = {name: [round(g * NOTIONAL, 2) for g in series]
             for name, series in raw['bench'].items()}

    def total(series):
        return series[-1] / series[0] - 1

    summary = {
        'total_return': total(model),
        'vs_smh': total(model) - total(bench['SMH']),
        'vs_qqq': total(model) - total(bench['QQQ']),
        'vs_spy': total(model) - total(bench['SPY']),
        'vs_ew': total(model) - total(bench['EW']),
    }

    monthly: list[dict] = []
    by_month: dict[str, list[int]] = {}
    for i, d in enumerate(raw['dates']):
        by_month.setdefault(d[:7], []).append(i)
    prev_close = {'model': model[0], 'SMH': bench['SMH'][0],
                  'QQQ': bench['QQQ'][0], 'SPY': bench['SPY'][0]}
    for month in sorted(by_month):
        last = by_month[month][-1]
        row = {'month': month}
        for key, series in (('model', model), ('SMH', bench['SMH']),
                            ('QQQ', bench['QQQ']), ('SPY', bench['SPY'])):
            row[key] = round(series[last] / prev_close[key] - 1, 6)
            prev_close[key] = series[last]
        monthly.append(row)

    return {'dates': raw['dates'], 'model': model, 'bench': bench,
            'summary': summary, 'monthly': monthly, 'as_of': raw['as_of']}


def export_changes(root: Path) -> list[dict]:
    cfg = json.loads(
        (root / 'tracking' / 'performance-config.json').read_text())
    out: list[dict] = []
    events = cfg['events']
    for i, ev in enumerate(events):
        if i == 0:
            out.append({'date': ev['date'], 'type': 'inception',
                        'ticker': None, 'note': ev['reason']})
            continue
        prev, curr = set(events[i - 1]['allocations']), set(ev['allocations'])
        for t in sorted(curr - prev):
            out.append({'date': ev['date'], 'type': 'add', 'ticker': t,
                        'note': ev['reason']})
        for t in sorted(prev - curr):
            out.append({'date': ev['date'], 'type': 'drop', 'ticker': t,
                        'note': ev['reason']})
        if curr == prev:
            out.append({'date': ev['date'], 'type': 'resize', 'ticker': None,
                        'note': ev['reason']})

    wb = openpyxl.load_workbook(
        root / '00-master' / 'ai_supply_chain_scoring.xlsx', data_only=True)
    ws = wb['Rating Audit']
    grouped: dict[tuple[str, str], dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, ticker, dim, _rating, rationale, _src, _conf, typ = row[:8]
        if not date or not typ or str(typ).strip().lower() == 'initial':
            continue
        if rationale is not None and not isinstance(rationale, str):
            # Column-shifted legacy row (an extra rating column pushed
            # Rationale→Source and the real Type into col 9). The Type slot
            # here holds a misaligned Confidence value, so the row can't be
            # trusted as a re-rate record. Skip rather than crash/mislabel.
            warn(f'Rating Audit row for {ticker} on {date} looks '
                 f'column-shifted (non-text rationale); skipping')
            continue
        key = (str(date)[:10], ticker)
        g = grouped.setdefault(key, {'dims': [], 'rationale': rationale})
        if not g['rationale'] and rationale:
            g['rationale'] = rationale
        g['dims'].append(str(dim))
    for (date, ticker), g in grouped.items():
        note = (g['rationale'] or '').strip()
        out.append({'date': date, 'type': 'rerate', 'ticker': ticker,
                    'detail': ', '.join(sorted(g['dims'])),
                    'note': note[:200] + ('…' if len(note) > 200 else '')})

    return sorted(out, key=lambda c: (c['date'], c['type']), reverse=True)


# Tolerate an editorial suffix on the heading (e.g. "## 1. One-line thesis
# [STALE - revisit]"); the section delimiter is the heading prefix, not an
# exact line match. The {placeholder} check below still rejects real templates.
THESIS_SECTION = re.compile(
    r'^## 1\. One-line thesis[^\n]*$(.*?)(?=^## |\Z)', re.M | re.S)


def export_theses(root: Path, tickers: list[str]) -> dict[str, str | None]:
    out: dict[str, str | None] = {}
    for t in tickers:
        path = root / 'per-stock' / t / 'thesis.md'
        if not path.exists():
            warn(f'{t}: no thesis.md — snippet omitted')
            out[t] = None
            continue
        m = THESIS_SECTION.search(path.read_text())
        body = '' if not m else m.group(1)
        lines = [ln.lstrip('> ').strip() for ln in body.splitlines()]
        text = ' '.join(ln for ln in lines if ln and ln != '---').strip()
        if not text or re.search(r'\{.+?\}', text):
            warn(f'{t}: thesis.md is still the template — snippet omitted')
            out[t] = None
        else:
            out[t] = text
    return out


def export_watchlist(root: Path, held: set[str] | None = None) -> list[dict]:
    """Every scored name with its six category sub-scores, TOTAL and tier.

    Scores are 0-100 (not capital), so this side-steps the dollar-scaling
    privacy boundary entirely — there are no real-dollar fields to leak.
    TOTAL/Tier are live Excel formulas with no cached value, so we recompute
    them in Python via recalc_watchlist (the project's canonical method).
    """
    from recalc_watchlist import recalc

    held = held or set()
    scoring = root / '00-master' / 'ai_supply_chain_scoring.xlsx'
    rows = recalc(str(scoring))
    names = _company_names(root)

    def r1(v):
        return round(float(v), 1) if isinstance(v, (int, float)) else None

    out = []
    for x in rows:
        layer_num = str(x['layer'] or '')[:2]
        out.append({
            'ticker': x['ticker'],
            'company': names.get(x['ticker'], x['ticker']),
            'layer_num': layer_num,
            'layer': LAYERS.get(layer_num, str(x['layer'] or '')),
            'value': r1(x['Value']), 'quality': r1(x['Quality']),
            'growth': r1(x['Growth']), 'ai': r1(x['AI']),
            'momentum': r1(x['Momentum']), 'risk': r1(x['Risk']),
            'score': r1(x['TOTAL']), 'tier': x['Tier'],
            'held': x['ticker'] in held,
        })
    if not out:
        warn('watchlist recalc returned no rows')
    # Scored names first (highest TOTAL), unscored (None) last.
    out.sort(key=lambda r: (r['score'] is None, -(r['score'] or 0)))
    return out


# Tier bands — must match recalc_watchlist.tier() and CLAUDE.md.
TIERS = [
    {'symbol': '✓✓✓', 'label': 'Top tier', 'min': 85, 'max': 100},
    {'symbol': '✓✓', 'label': 'Strong', 'min': 70, 'max': 84},
    {'symbol': '✓', 'label': 'Average', 'min': 55, 'max': 69},
    {'symbol': '?', 'label': 'Below average', 'min': 40, 'max': 54},
    {'symbol': '✗', 'label': 'Avoid', 'min': 0, 'max': 39},
]

# Category → component metrics (plain-English, friend-facing). Weights and the
# one-line descriptions are pulled LIVE from the Weights sheet in
# export_methodology; only these metric lists (which change rarely) are authored
# here. `kind` flags how the inputs are derived, for an honest explainer.
CATEGORY_METRICS = {
    'Value': {'kind': 'objective', 'metrics': [
        'Forward P/E', 'EV/EBITDA', 'Free-cash-flow yield', 'Price / sales',
        'PEG ratio (P/E ÷ growth)']},
    'Quality': {'kind': 'objective', 'metrics': [
        'Return on invested capital', 'Gross margin',
        'Free-cash-flow margin', 'Net debt / EBITDA']},
    'Growth': {'kind': 'objective', 'metrics': [
        'Revenue 3-yr CAGR', 'Revenue growth (latest quarter)',
        'EPS growth (year over year)']},
    'AI Thesis': {'kind': 'judgment', 'metrics': [
        'AI revenue as % of total', 'Position in the supply chain',
        'Moat strength', 'Capacity expansion', 'Hyperscaler exposure']},
    'Momentum': {'kind': 'mixed', 'metrics': [
        'EPS estimate revisions', 'Relative strength vs sector',
        'Insider activity', '% of days above the 50-day average price']},
    'Risk': {'kind': 'judgment', 'metrics': [
        'Customer concentration', 'Geographic / export risk',
        'Balance-sheet strength', 'Regulatory / litigation overhang',
        'Disruption / business-model durability']},
}

KIND_NOTE = {
    'objective': 'Computed from financial data, each metric scored on fixed '
                 'threshold bands.',
    'judgment': 'Hand-rated 1–5 per company against a published rubric, then '
                'scaled to 0–100.',
    'mixed': 'Three hand-rated 1–5 signals blended with one objective '
             'price-trend metric.',
}


def export_methodology(root: Path) -> dict:
    """How the scores are built — weights live from the Weights sheet, tier
    bands + metric lists from the constants above. No dollars, pure rubric."""
    wb = openpyxl.load_workbook(
        root / '00-master' / 'ai_supply_chain_scoring.xlsx', data_only=True)
    if 'Weights' not in wb.sheetnames:
        fail('scoring workbook has no Weights sheet')
    categories = []
    total = 0.0
    for row in wb['Weights'].iter_rows(min_row=2, values_only=True):
        name, weight, note = (row + (None, None, None))[:3]
        if not name or weight is None or name == 'Category':
            continue
        cm = CATEGORY_METRICS.get(name)
        if cm is None:
            warn(f'Weights category {name!r} has no metric definition')
        total += float(weight)
        categories.append({
            'name': name,
            'weight': round(float(weight), 4),
            'note': note or '',
            'kind': (cm or {}).get('kind', 'objective'),
            'kind_note': KIND_NOTE.get((cm or {}).get('kind', ''), ''),
            'metrics': (cm or {}).get('metrics', []),
        })
    if not categories:
        fail('no category weights found in Weights sheet')
    if abs(total - 1.0) > 0.001:
        warn(f'category weights sum to {total:.3f}, not 1.0')
    return {'categories': categories, 'tiers': TIERS,
            'total_note': 'Total Score is the six category scores (each 0–100) '
                          'blended by the weights above. Tier is the band the '
                          'Total falls into.'}


def export_scans(root: Path) -> list[dict]:
    lp = root / 'tracking' / 'notion-scan-links.json'
    if not lp.exists():
        warn('tracking/notion-scan-links.json missing — scans page empty')
        return []
    scans = json.loads(lp.read_text())
    for s in scans:
        if not all(s.get(k) for k in ('date', 'title', 'url')):
            fail(f'malformed scan link entry: {s}')
    linked = {s['date'] for s in scans}
    for md in sorted(root.glob('tracking/weekly-news-scan-*.md')):
        date = md.stem.replace('weekly-news-scan-', '')
        if date not in linked:
            warn(f'scan {date} has no Notion link entry — omitted from site')
    return sorted(scans, key=lambda s: s['date'], reverse=True)


def main(root: Path | None = None) -> None:
    root = root or Path(__file__).resolve().parent.parent
    out_dir = root / 'site' / 'data'
    out_dir.mkdir(parents=True, exist_ok=True)

    positions = export_positions(root)
    performance = export_performance(root)
    changes = export_changes(root)
    theses = export_theses(root, [p['ticker'] for p in positions])
    watchlist = export_watchlist(root, held={p['ticker'] for p in positions})
    methodology = export_methodology(root)
    scans = export_scans(root)
    cfg = json.loads(
        (root / 'tracking' / 'performance-config.json').read_text())
    meta = {
        'generated_at': dt.datetime.now(dt.timezone.utc).isoformat(
            timespec='seconds'),
        'as_of': performance['as_of'],
        'last_rebalance': cfg['events'][-1]['date'],
        'holdings': len(positions),
        'scored': sum(1 for r in watchlist if r['score'] is not None),
        'notional': int(NOTIONAL),
    }
    for name, data in (('positions', positions), ('performance', performance),
                       ('changes', changes), ('theses', theses),
                       ('watchlist', watchlist), ('methodology', methodology),
                       ('scans', scans), ('meta', meta)):
        (out_dir / f'{name}.json').write_text(json.dumps(data) + '\n')
        print(f'wrote site/data/{name}.json')


if __name__ == '__main__':
    if len(sys.argv) == 1:
        main()
    elif len(sys.argv) == 3 and sys.argv[1] == '--root':
        main(Path(sys.argv[2]))
    else:
        fail(f'usage: export_site_data.py [--root DIR] (got {sys.argv[1:]})')
