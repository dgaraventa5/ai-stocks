"""Generate an executable trade ticket from a model event — spec §B (2026-08-09).

Holdings come from the latest reconciliation snapshot (ACTUAL account state,
never the previous model event — partial fills and drift must not compound,
B2). Reference prices come from the existing yfinance price path. The output
is an append-only JSON ticket under tracking/live/tickets/ (gitignored, §E)
that ONLY scripts/execute_ticket.py — run by Dom — can turn into orders.

Called automatically by refresh_targets on every real model event (B1), or
manually:
  python3 scripts/generate_trade_ticket.py            # from current Targets
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

import trade_ticket as tt

_REPO_ROOT = Path(__file__).resolve().parent.parent
LIVE_DIR = _REPO_ROOT / 'tracking' / 'live'


def _flag(msg: str) -> None:
    print(f'FLAG: {msg}', file=sys.stderr)


def _latest_snapshot(live_dir: Path) -> dict | None:
    snaps = sorted((Path(live_dir) / 'recon').glob('snapshot-*.json'))
    return json.loads(snaps[-1].read_text()) if snaps else None


def _yf_last_close(ticker: str) -> float | None:
    import yfinance as yf   # lazy (minimal-CI convention)
    hist = yf.Ticker(ticker).history(period='5d', auto_adjust=False)
    if hist is None or hist.empty:
        return None
    return round(float(hist['Close'].iloc[-1]), 2)


def generate(target_weights: dict[str, float], event: dict, *,
             live_dir: Path = LIVE_DIR, prices_fn=_yf_last_close,
             now: str | None = None) -> Path | None:
    """Build + write the ticket. Returns the path, or None (flagged) when no
    reconciliation snapshot exists — deltas from assumed holdings are exactly
    the compounding error B2 forbids, so we refuse rather than guess."""
    live_dir = Path(live_dir)
    snap = _latest_snapshot(live_dir)
    if snap is None:
        _flag('no recon snapshot under tracking/live/recon/ — ticket NOT '
              'generated. Run reconcile_account.py first (need actual '
              'holdings, not assumptions — spec B2).')
        return None

    cfg = dict(tt.DEFAULTS)
    cfg_path = live_dir / 'executor-config.json'
    if cfg_path.exists():
        cfg.update(json.loads(cfg_path.read_text()))

    tickers = sorted(set(target_weights) | set(snap['positions']))
    prices = {}
    for t in tickers:
        px = prices_fn(t)
        if px:
            prices[t] = px
    result = tt.compute_orders(target_weights, snap['positions'],
                               snap['cash'], prices, cfg)
    for u in result['untradeable']:
        _flag(f"{u['ticker']}: {u['reason']}")
    for s in result['skipped']:
        _flag(f"{s['ticker']}: {s['reason']}")

    now = now or dt.datetime.now(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    ticket = tt.build_ticket(result, basis_event=event, created_at=now,
                             cfg=cfg, account_state_as_of=snap['as_of'])

    tdir = live_dir / 'tickets'
    tdir.mkdir(parents=True, exist_ok=True)
    path = tdir / f"ticket-{ticket['ticket_id']}.json"
    n = 1
    while path.exists():   # append-only: a regenerated ticket is a NEW file (B3)
        n += 1
        path = tdir / f"ticket-{ticket['ticket_id']}-{n}.json"
    path.write_text(json.dumps(ticket, indent=2) + '\n')
    print(f'ticket written: {path} ({len(ticket["orders"])} orders, '
          f'{len(ticket["suppressed"])} dust-suppressed, '
          f'{len(ticket["untradeable"])} untradeable)')
    return path


def on_model_event(event: dict, weights: dict[str, float]) -> Path | None:
    """B1 hook target: refresh_targets calls this on every real fired event."""
    return generate(weights, event)


def _targets_weights() -> tuple[dict[str, float], dict]:
    from openpyxl import load_workbook   # lazy
    wb = load_workbook(_REPO_ROOT / '00-master' / 'portfolio.xlsx')
    ws = wb['Targets']
    hdr = [c.value for c in ws[2]]
    col = {name: i for i, name in enumerate(hdr)}
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[0] != 'TOTAL' and row[col['Include?']] == 'Y':
            out[row[0]] = (row[col['Target %']] or 0) / 100.0
    title = str(ws.cell(row=1, column=1).value or '')
    return out, title


if __name__ == '__main__':
    weights, title = _targets_weights()
    event = {'date': dt.date.today().isoformat(), 'kind': 'manual',
             'reason': f'manual generation from committed Targets ({title})'}
    p = generate(weights, event)
    sys.exit(0 if p else 1)
