"""Apply directness re-ratings to the competitive zone — Layers 2-10 top-40 (2026-06-10).

Completes the watchlist's move onto one consistent directness basis: Layer 1
done earlier today; this is the 35 non-Layer-1 names in the current top 40,
researched by 7 bucket agents and reconciled by the orchestrator.

Theme: D5 systematically cut for semicap (sells to FABS, not hyperscalers);
D1 trimmed for diversified names where AI is a slice not a current majority;
D2 cut for cyclical commodities (memory/HDD); one D1 UPGRADE (TSM, HPC now 61%).
20 of 35 names held unchanged (genuine AI exposure confirmed). Only D1/D2/D5
touched. Orchestrator consistency calls vs bucket agents: MRVL D1 5->4 (strict
AI-majority test, matching AMD/AVGO); LSCC D5 held 3 (not upgraded to 4 —
distributor-masked, indirect, consistent with the semicap D5 cuts).
"""
from __future__ import annotations

from openpyxl import load_workbook

XLSX = '/Users/dom/Desktop/ai-stocks/00-master/ai_supply_chain_scoring.xlsx'
DATE = '2026-06-10'
COL = {'D1': 20, 'D2': 21, 'D5': 24}

C = {
 'GEV': {'D1': (2, "AI exposure is ORDERS/backlog not current revenue: $2.4B DC equipment orders ~=13% of Q1 orders, but disclosed current DC revenue is <10% of a broad grid+gas+wind co. Distinguish backlog from revenue.", 'GEV Q1 2026 8-K 2026-04-22', 'Med')},
 'TER': {'D1': (4, "Genuine AI majority but trimmed for precision: ~70% of rev AI-related, SemiTest 87%, compute ~75% of SoC test -- but 'AI-related' is mgmt framing incl. non-AI HPC, and Robotics is non-AI. 25-50% band.", 'TER Q1 2026 8-K/call 2026-04-29', 'High'),
         'D5': (3, "Sells test to OSATs/IDMs/fabless, NOT hyperscalers; indirect pull from hyperscaler custom-silicon ASICs being tested keeps it at 3 (was 4, overstated a direct relationship).", 'TER Q1 2026 call', 'High')},
 'LRCX': {'D1': (3, "AI is a demand tailwind, not a disclosed revenue majority -- sells etch/dep into all WFE; AI-specific (adv packaging/HBM) is a modest slice of a broad cyclical base incl. NAND recovery.", 'LRCX Q3 FY2026 2026-04-22', 'High'),
          'D5': (2, "Sells to FABS (TSMC/Samsung/SK Hynix/Micron/Intel), zero direct hyperscaler relationship; 4 was narrative inflation.", 'LRCX Q3 FY2026', 'High')},
 'ARM': {'D1': (4, "IP not equipment, but AI-infrastructure is still a minority: data-center royalty doubled YoY yet smartphone/IoT/auto remain the largest base; CSS to ~50% of royalties is FUTURE.", 'ARM FY2026 6-K 2026-05', 'Med'),
         'D5': (4, "Most direct hyperscaler link in semicap (AWS Graviton/Google Axion/MSFT Cobalt pay per-chip royalties) but it's licensing not co-design, and hyperscalers are part of a minority data-center segment; 4 not 5.", 'ARM FY2026 6-K', 'High')},
 'ASML': {'D1': (3, "AI drives the demand/guidance but ASML sells litho into all advanced logic+DRAM; no disclosed AI line; EUV ~65% of sales is leading-edge broadly, only partly AI-specific.", 'ASML Q1 2026 2026-04-15', 'High'),
          'D5': (2, "Sells exclusively to FABS (TSMC/Samsung/Intel/SK Hynix); no hyperscaler relationship; 4 was pure narrative inflation. D2=5 EUV monopoly intact.", 'ASML Q1 2026', 'High')},
 'AMAT': {'D1': (3, "Same as LRCX: leading-edge/HBM/adv-packaging are the fastest-growing WFE but not a disclosed majority of a base that still includes large ICAPS/trailing-edge + China (non-AI).", 'AMAT Q2 FY2026 8-K 2026-05-14', 'High'),
          'D5': (2, "Sells to FABS not hyperscalers; no direct hyperscaler relationship; 4 inflated.", 'AMAT Q2 FY2026 8-K', 'High')},
 'TSM': {'D1': (5, "UPGRADE. HPC platform = 61% of Q1 2026 revenue (up from 55% prior qtr), overwhelmingly AI accelerator/datacenter wafers (NVDA/AMD/AVGO/custom-ASIC). Cleared >50%; old 4 was a stale sub-majority model.", 'TSMC Q1 2026 slides 2026-04-16', 'High')},
 'AMD': {'D1': (4, "CUT: Data Center is 56% of rev but blends EPYC server CPUs w/ Instinct GPUs; Instinct alone ~17-19% of total. 43% of co (Client/Gaming/Embedded) is non-AI. AI-infra is 25-50% band, not a majority. (May re-cross to 5 in 2-3 qtrs.)", 'AMD Q1 2026 8-K', 'Med'),
         },
 'MRVL': {'D1': (4, "Data center 76% of rev but AI-specific run-rate ~47% of total -- below the >50% majority bar applied to AMD/AVGO. [Override: agent held 5 on 76% DC mix; cut to 4 for consistency.] Borderline; could be 5 if all AI-cluster DC silicon counts.", 'MRVL Q1 FY26', 'Med')},
 'ALAB': {'D2': (4, "Best product today (Scorpio/PCIe Gen6 leadership) but smallest/newest in a contestable connectivity tier (AVGO/MRVL/NVLink can encroach) -- 'early leader, competition coming' = 4 not 5. D1=5 pure-play holds.", 'Astera Q1 FY26 2026-05-05', 'Med')},
 'MU': {'D2': (4, "CUT: memory is a cyclical commodity, MU a 3-player price-taker on the broad DRAM/NAND cycle (~44% of rev is commodity Mobile/Client/Auto). HBM4 sold-out fixed-price LTAs are real but ring-fenced pricing power, not a whole-business bottleneck. D1=5 (DC 56%) holds.", 'MU Q2 FY26 8-K', 'Med')},
 'SNDK': {'D1': (3, "CUT: disclosed datacenter ~14.5% of total revenue (10-25% band); ~85% is still commodity/consumer NAND; HBF in qualification = zero revenue. Prior 4 over-credited 'majority of GROWTH' as majority of revenue.", 'Sandisk Q1 FY26 / Q3 FY26 10-Q', 'Med'),
          'D2': (3, "CUT: pure NAND is the most commoditized memory (5-6 suppliers, price-taker on the cycle); the +33-38% QoQ ESSD pricing is cycle-upswing, not structural power.", 'Sandisk Q3 FY26', 'Med'),
          'D5': (3, "CUT: HBF qualification = pipeline not revenue; datacenter ~14.5% routed via OEMs/channel, no disclosed hyperscaler >10% concentration.", 'Sandisk Q1 FY26', 'Med')},
 'FN': {'D1': (4, "CUT: contract optical mfr where AI is a large slice not ~100%: datacom transceivers ~$260M (21%, down 6% QoQ) + AI-driven DCI ~$197M ~= $457M ~38% of rev; rest is non-AI telecom/industrial/auto. NVIDIA >30% is the AI lever (and risk).", 'FN Q3 FY26', 'Med')},
 'DELL': {'D1': (3, "CUT: AI-optimized servers ~$16.1B of $43.8B Q1 FY27 ~= 37% (FY27 outlook ~36%); fast-growing but a MINORITY of total, rest is commodity CSG/PCs + traditional storage, and at LOW margin.", 'Dell Q1 FY27 8-K 2026-05-28', 'High')},
 'STX': {'D2': (4, "CUT: STX/WDC duopoly has real current pricing power ($/TB rising, sold out to late-2027) but HDD is a CYCLICAL near-commodity, not a structural non-cyclical chokepoint (reserve 5 for EUV/leading-edge). Pricing mean-reverts. D1=4/D5=5 hold.", 'STX context 2026-06-05; Q3 FY26 8-K', 'Med')},
}


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)
    ws = wb['Watchlist']
    audit = wb['Rating Audit']
    rows = {ws.cell(row=r, column=1).value: r
            for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}

    applied = 0
    for tkr, dims in C.items():
        if tkr not in rows:
            print(f'!! {tkr} not on Watchlist — skipped'); continue
        r = rows[tkr]
        for dim, (new, rationale, source, conf) in dims.items():
            c = COL[dim]
            old = ws.cell(row=r, column=c).value
            ws.cell(row=r, column=c).value = new
            audit.append([DATE, tkr, dim, new, f'(was {old}) {rationale}',
                          source, conf, 'Update'])
            print(f'{tkr:<5} {dim}: {old} -> {new}')
            applied += 1
        ws.cell(row=r, column=4).value = DATE

    wb.save(XLSX)
    print(f'\napplied {applied} changes across {len(C)} names (20 of 35 held unchanged)')


if __name__ == '__main__':
    main()
