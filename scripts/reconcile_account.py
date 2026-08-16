"""Account reconciliation & monitoring — §D of the agentic-execution spec.

Read-only, deterministic, stdlib only. Account state arrives as an explicit
JSON payload (Claude pulls it with MCP READ tools in attended sessions —
O1 answered NO 2026-08-09, so there is no headless fetch path; this script
never talks to the network and never places orders).

Outputs:
  tracking/live/recon/snapshot-{date}.json   full fidelity (GITIGNORED, §E)
  tracking/live/trading-halt.flag            created on anomalies; never sells
  tracking/live-status.json                  COMMITTED — sanitized: booleans,
                                             dates, counts, tickers only
  tracking/live-vs-model.json                COMMITTED — relative % only

Usage (attended session):
  python3 scripts/reconcile_account.py --account-json state.json
  python3 scripts/reconcile_account.py --account-json state.json \
      --external-flow 500.00   # declare a deposit (withdrawal: negative)

External cash flows (deposits/withdrawals) are DECLARED, never inferred: an
undeclared equity jump is indistinguishable from an executor bug, so it halts
(D3). A declared flow goes to the append-only ledger
tracking/live/recon/flows.jsonl (gitignored, §E) where every later run — the
attended session and the launchd cron alike — can explain the same move, and
the live-vs-model baseline is divisor-adjusted so a deposit doesn't print as
performance.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
LIVE_DIR = _REPO_ROOT / 'tracking' / 'live'
STATUS_PATH = _REPO_ROOT / 'tracking' / 'live-status.json'
LVM_PATH = _REPO_ROOT / 'tracking' / 'live-vs-model.json'
SERIES_PATH = _REPO_ROOT / 'tracking' / 'performance-series.json'

DRIFT_ALERT = 0.05        # relative drift threshold (D2)
EQUITY_TOL = 0.03         # unexplained-equity tolerance (D3)
OPEN_STATES = ('queued', 'confirmed', 'unconfirmed', 'partially_filled')
DEAD_STATES = ('cancelled', 'expired', 'rejected', 'failed', 'voided')


# §E sanitizer gates — called before EVERY committed write; fail closed on any
# unknown field. Same standing as the site privacy gate: never weaken.
_STATUS_SCHEMA = {'as_of': str, 'halted': bool, 'positions': int,
                  'open_orders': int, 'drift_flags': list,
                  'regen_needed': list, 'anomaly_count': int}
_LVM_ENTRY_SCHEMA = {'date': str, 'live_pct': (float, int, type(None)),
                     'model_pct': (float, int, type(None)),
                     'shortfall_pct': (float, int, type(None))}


def assert_sanitized_status(status: dict) -> None:
    for key, val in status.items():
        if key not in _STATUS_SCHEMA:
            raise ValueError(f'live-status field {key!r} not in the sanitized '
                             f'allowlist (§E) — refusing to commit it')
        if not isinstance(val, _STATUS_SCHEMA[key]):
            raise ValueError(f'live-status field {key!r} has type '
                             f'{type(val).__name__}, expected '
                             f'{_STATUS_SCHEMA[key]}')
    for lst in (status.get('drift_flags', []), status.get('regen_needed', [])):
        for item in lst:
            if not isinstance(item, str):
                raise ValueError(f'{item!r}: ticker lists may contain only '
                                 f'ticker strings (§E)')


def assert_sanitized_lvm(doc: dict) -> None:
    if set(doc) != {'baseline_date', 'series'}:
        raise ValueError(f'live-vs-model keys {sorted(doc)} != '
                         f"['baseline_date', 'series'] (§E)")
    for entry in doc['series']:
        for key, val in entry.items():
            if key not in _LVM_ENTRY_SCHEMA:
                raise ValueError(f'live-vs-model field {key!r} not in the '
                                 f'relative-percentages allowlist (§E)')
            if not isinstance(val, _LVM_ENTRY_SCHEMA[key]):
                raise ValueError(f'live-vs-model field {key!r}: bad type')


# §D external-flow ledger — Dom-declared deposits/withdrawals. Append-only,
# gitignored (real dollars). Declarations are idempotent per (date, amount) so
# the attended run and the cron's same-day recon can't double-count one flow.

def _flows_path(live_dir: Path) -> Path:
    return Path(live_dir) / 'recon' / 'flows.jsonl'


def read_flows(live_dir: Path) -> list[dict]:
    path = _flows_path(live_dir)
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text().splitlines()
            if line.strip()]


def declare_flow(live_dir: Path, date: str, amount: float) -> None:
    """Record an external cash flow (deposit > 0, withdrawal < 0), dated to
    the recon that first sees it. No-op on zero or on an exact duplicate."""
    if not amount:
        return
    entry = {'date': date, 'amount': round(float(amount), 2)}
    if entry in read_flows(live_dir):
        print(f'external flow already declared for {date} — ledger unchanged')
        return
    path = _flows_path(live_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('a') as fh:
        fh.write(json.dumps(entry) + '\n')
    print(f'external flow declared: {amount:+.2f} on {date}')


def guard_flow_is_delta(flow: float, cash: float, prior: dict | None) -> None:
    """Refuse a declared flow that looks like a BALANCE rather than a DELTA.

    --external-flow wants the amount that moved. A resulting cash balance
    passed by mistake is also a float of dollars, so nothing catches it: on
    2026-08-14 a deposit was declared as the post-deposit balance, over-stating
    the flow by the pre-existing cash. That over-declaration then inflated the
    baseline divisor and printed a large phantom negative live return. The
    fingerprint is the declared deposit equalling current cash while the
    account already held cash. Prior cash ~0 is the honest case (balance ==
    delta) — no guard there.
    """
    if flow <= 0 or not prior:
        return
    prior_cash = float(prior.get('cash', 0.0))
    if prior_cash < 0.01 or abs(flow - float(cash)) >= 0.01:
        return
    raise ValueError(
        f'declared flow {flow:,.2f} equals the resulting cash balance — '
        f'--external-flow takes the amount that MOVED, not the resulting '
        f'balance. The account already held {prior_cash:,.2f} in cash before '
        f'this recon; did you mean {float(cash) - prior_cash:,.2f}? '
        f'(rule 3: flagged, not guessed)')


def flows_between(live_dir: Path, after: str, through: str) -> float:
    """Sum of declared flows dated in (after, through] — the window between
    the prior snapshot and the current recon."""
    return sum(f['amount'] for f in read_flows(live_dir)
               if after < f['date'] <= through)


def _receipts(live_dir: Path) -> list[dict]:
    rdir = live_dir / 'receipts'
    if not rdir.is_dir():
        return []
    return [json.loads(p.read_text()) for p in sorted(rdir.glob('receipt-*.json'))]


def verify_fills(receipts: list[dict], orders: list[dict]):
    """Match receipt orders to live order states by order_id (D1)."""
    by_id = {o.get('order_id'): o for o in orders}
    fills: dict[str, dict[str, str]] = {}
    regen: list[str] = []
    for r in receipts:
        states: dict[str, str] = {}
        for o in r.get('orders', []):
            live = by_id.get(o.get('order_id'))
            state = (live or {}).get('state') or o.get('state', 'unknown')
            states[o['ticker']] = state
            if state in DEAD_STATES:
                regen.append(o['ticker'])   # unfilled + dead → regenerate
        fills[r['ticket_id']] = states
    return fills, sorted(set(regen))


def drift_check(state: dict, target_weights: dict[str, float],
                threshold: float = DRIFT_ALERT) -> list[dict]:
    """Relative per-name drift of actual vs target weight (D2). Flags only —
    drift never halts; it feeds B2's delta-from-actuals rule."""
    equity = state['equity']
    out = []
    for t, tw in sorted(target_weights.items()):
        if not tw or not equity:
            continue
        pos = state['positions'].get(t, {})
        actual = pos.get('shares', 0.0) * pos.get('price', 0.0) / equity
        rel = abs(actual - tw) / tw
        if rel > threshold:
            out.append({'ticker': t, 'rel_drift_pct': round(rel * 100, 1)})
    return out


def _prior_snapshot(live_dir: Path, as_of: str) -> dict | None:
    snaps = sorted((live_dir / 'recon').glob('snapshot-*.json'))
    prior = [p for p in snaps if p.stem.split('snapshot-')[1] < as_of]
    return json.loads(prior[-1].read_text()) if prior else None


def detect_anomalies(state: dict, target_weights: dict, receipts: list[dict],
                     prior: dict | None, declared_flow: float = 0.0) -> list[str]:
    """D3 halt conditions. Any entry here raises the kill switch (which stops
    future EXECUTIONS only — it never sells anything). declared_flow: net
    Dom-declared external cash flow since the prior snapshot (ledger above) —
    an EXPLAINED equity move; undeclared moves still halt."""
    anomalies = []
    known = set(target_weights) | {o['ticker'] for r in receipts
                                   for o in r.get('orders', [])}
    for t in sorted(state['positions']):
        if t not in known:
            anomalies.append(
                f'position {t} has unknown provenance (not in roster, not in '
                f'any receipt) — possible executor bug or account misuse')
    if state['cash'] < 0:
        anomalies.append('cash negative')
    if prior is not None:
        # Expected equity: prior shares marked at today's prices + prior cash.
        # Composition changes (fills since prior) make this ill-posed — skip
        # honestly rather than false-alarm (rule 3).
        filled_since = any(o.get('state') == 'filled'
                           for o in state.get('orders', []))
        if not filled_since:
            expected = prior['cash'] + declared_flow
            for t, pos in prior['positions'].items():
                px = state['positions'].get(t, {}).get('price', pos.get('price'))
                expected += pos['shares'] * px
            if expected > 0:
                gap = abs(state['equity'] - expected) / expected
                if gap > EQUITY_TOL:
                    anomalies.append(
                        f'equity move unexplained by market moves of held '
                        f'names ({gap:.1%} vs expected) — if this is a '
                        f'deposit/withdrawal, declare it with --external-flow')
    return anomalies


def live_vs_model(state: dict, live_dir: Path, model_series_path: Path,
                  lvm_path: Path) -> None:
    """Implementation-shortfall line (D5): cumulative live vs model return
    since the live baseline. COMMITTED — relative percentages only."""
    base_path = live_dir / 'recon' / 'baseline.json'
    flows = read_flows(live_dir)
    if not base_path.exists():
        # Flows dated at/before creation are embedded in the creation equity.
        base_path.write_text(json.dumps(
            {'date': state['as_of'], 'equity': state['equity'],
             'applied_flows': [f for f in flows
                               if f['date'] <= state['as_of']]}) + '\n')
    base = json.loads(base_path.read_text())
    base.setdefault('applied_flows', [])

    # Divisor adjustment (index-style): a declared external flow F at current
    # equity E scales the baseline by E/(E−F), so live_pct is unchanged by the
    # flow itself and later market moves compound correctly. Applied exactly
    # once per ledger entry (applied_flows), whichever run sees it first.
    pending = [f for f in flows
               if f['date'] > base['date'] and f not in base['applied_flows']]
    if pending:
        total = sum(f['amount'] for f in pending)
        if state['equity'] - total > 0:
            base['equity'] = round(
                base['equity'] * state['equity'] / (state['equity'] - total), 4)
        else:
            print(f'FLAG: declared flow {total:+.2f} >= current equity — '
                  f'baseline NOT adjusted; live-vs-model needs a manual '
                  f'baseline reset (rule 3: flagged, not guessed)')
        base['applied_flows'].extend(pending)
        base_path.write_text(json.dumps(base) + '\n')
    live_pct = round((state['equity'] / base['equity'] - 1) * 100, 2)

    model_pct = None
    try:
        series = json.loads(Path(model_series_path).read_text())
        idx = {d: v for d, v in zip(series['dates'], series['model'])}

        def at_or_before(day):
            past = [d for d in series['dates'] if d <= day]
            return idx[past[-1]] if past else None
        m0, m1 = at_or_before(base['date']), at_or_before(state['as_of'])
        if m0 and m1:
            model_pct = round((m1 / m0 - 1) * 100, 2)
    except (OSError, ValueError, KeyError):
        pass   # model series unavailable → model_pct stays None (flagged below)

    entry = {'date': state['as_of'], 'live_pct': live_pct,
             'model_pct': model_pct,
             'shortfall_pct': (round(live_pct - model_pct, 2)
                               if model_pct is not None else None)}
    try:
        doc = json.loads(Path(lvm_path).read_text())
    except (OSError, ValueError):
        doc = {'baseline_date': base['date'], 'series': []}
    doc['series'] = [e for e in doc['series'] if e['date'] != entry['date']]
    doc['series'].append(entry)
    doc['series'].sort(key=lambda e: e['date'])
    assert_sanitized_lvm(doc)                    # §E gate before committed write
    Path(lvm_path).write_text(json.dumps(doc, indent=2) + '\n')


def run(state: dict, *, live_dir: Path, target_weights: dict[str, float],
        model_series_path: Path = SERIES_PATH,
        status_path: Path = STATUS_PATH, lvm_path: Path = LVM_PATH,
        external_flow: float = 0.0) -> dict:
    live_dir = Path(live_dir)
    (live_dir / 'recon').mkdir(parents=True, exist_ok=True)
    prior = _prior_snapshot(live_dir, state['as_of'])
    if external_flow:
        guard_flow_is_delta(external_flow, state['cash'], prior)
        declare_flow(live_dir, state['as_of'], external_flow)
    receipts = _receipts(live_dir)
    declared = flows_between(live_dir, prior['as_of'] if prior else '',
                             state['as_of'])

    fills, regen = verify_fills(receipts, state.get('orders', []))
    drift = drift_check(state, target_weights)
    anomalies = detect_anomalies(state, target_weights, receipts, prior,
                                 declared_flow=declared)

    halt_path = live_dir / 'trading-halt.flag'
    if anomalies:
        halt_path.write_text(
            f'{state["as_of"]} auto-halt (reconcile_account.py):\n'
            + '\n'.join(f'- {a}' for a in anomalies) + '\n')
        print(f'TRADING HALT raised: {"; ".join(anomalies)}')
    halted = halt_path.exists()

    # Full-fidelity snapshot — gitignored (§E). This is the actuals source for
    # the next ticket's share deltas (B2).
    snap_path = live_dir / 'recon' / f'snapshot-{state["as_of"]}.json'
    snap_path.write_text(json.dumps(
        {**state, 'fills': fills, 'anomalies': anomalies}, indent=2) + '\n')

    live_vs_model(state, live_dir, model_series_path, lvm_path)

    # Committed status — SANITIZED: booleans, dates, counts, tickers, relative
    # percentages. No dollars, no share counts, no order ids (§E).
    open_orders = sum(1 for o in state.get('orders', [])
                      if o.get('state') in OPEN_STATES)
    status = {
        'as_of': state['as_of'],
        'halted': halted,
        'positions': len(state['positions']),
        'open_orders': open_orders,
        'drift_flags': [d['ticker'] for d in drift],
        'regen_needed': regen,
        'anomaly_count': len(anomalies),
    }
    assert_sanitized_status(status)              # §E gate before committed write
    Path(status_path).write_text(json.dumps(status, indent=2) + '\n')

    for d in drift:
        print(f'drift flag: {d["ticker"]} {d["rel_drift_pct"]}% from target '
              f'(> {DRIFT_ALERT:.0%} relative)')
    for t in regen:
        print(f'unfilled order dead: {t} — regenerate on next model event '
              f'or --regen-unfilled')
    print(f'recon {state["as_of"]}: {len(state["positions"])} positions, '
          f'{open_orders} open orders, halted={halted}')
    return {'halted': halted, 'drift': drift, 'fills': fills,
            'regen_needed': regen, 'anomalies': anomalies,
            'snapshot': snap_path}


def _targets_from_workbook() -> dict[str, float]:
    from openpyxl import load_workbook   # lazy (minimal-CI convention)
    wb = load_workbook(_REPO_ROOT / '00-master' / 'portfolio.xlsx')
    ws = wb['Targets']
    hdr = [c.value for c in ws[2]]
    col = {name: i for i, name in enumerate(hdr)}
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[0] != 'TOTAL' and row[col['Include?']] == 'Y':
            out[row[0]] = (row[col['Target %']] or 0) / 100.0
    return out


if __name__ == '__main__':
    ap = argparse.ArgumentParser(
        description='Reconcile live account state against the model')
    ap.add_argument('--account-json', required=True,
                    help='account state JSON pulled via MCP READ tools '
                         '(attended session — O1: no headless OAuth)')
    ap.add_argument('--external-flow', type=float, default=0.0,
                    metavar='AMOUNT',
                    help='declare an external cash flow since the prior '
                         'snapshot (deposit > 0, withdrawal < 0): explains '
                         'the equity move to the D3 anomaly check and '
                         'divisor-adjusts the live-vs-model baseline')
    args = ap.parse_args()
    payload = json.loads(Path(args.account_json).read_text())
    run(payload, live_dir=LIVE_DIR, target_weights=_targets_from_workbook(),
        external_flow=args.external_flow)
