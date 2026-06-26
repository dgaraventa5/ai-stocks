"""Rebuild all computed-column formulas in the Watchlist tab.

Why this exists (2026-06-12): openpyxl `delete_rows` shifts cell values but
does NOT rewrite formula references, so any row deletion silently breaks every
formula below it (found: 106 of 135 rows referencing the wrong row after the
CTRA + one other removal). Run this script after ANY structural row change.

It also implements the layer-conditional EV-column bands:
  - Standard layers: col F = EV/EBITDA, standard bands
  - Layer 10 (rule 10, added 2026-05-26): col F = EV/FCF, SaaS bands
    (<=20->100 ... >100->15). These bands were documented in CLAUDE.md but
    never actually written into the workbook formulas until 2026-06-12.
  - Layer 9 capacity cohort (rule 13, added 2026-06-12): col F = EV per
    secured MW ($M/MW), replacement-cost-anchored bands.

Usage:
  python3 scripts/rebuild_watchlist_formulas.py            # rebuild + save
  python3 scripts/rebuild_watchlist_formulas.py --check    # report drift only
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = str(_REPO_ROOT / '00-master/ai_supply_chain_scoring.xlsx')

# --- EV-column (col F) band segments -----------------------------------------

F_STANDARD = ('IFERROR(IF(F{r}="","",IF(F{r}<0,5,IF(F{r}<=10,90,IF(F{r}<=15,75,'
              'IF(F{r}<=20,60,IF(F{r}<=25,45,IF(F{r}<=35,30,IF(F{r}<=50,15,5)))))))),"")')

# Rule 10: Layer 10 col F holds EV/FCF, SaaS-calibrated bands. Negative FCF -> 5.
F_L10_EVFCF = ('IFERROR(IF(F{r}="","",IF(F{r}<0,5,IF(F{r}<=20,100,IF(F{r}<=30,90,'
               'IF(F{r}<=40,75,IF(F{r}<=55,60,IF(F{r}<=75,45,IF(F{r}<=100,30,15)))))))),"")')

# Rule 13: Layer 9 capacity cohort col F holds EV / secured MW in $M per MW.
# Anchor: ~$10-12M/MW replacement cost for an energized AI-ready data center
# (shell + power, excl. IT gear). Below ~half replacement = deep discount.
F_L9_EVMW = ('IFERROR(IF(F{r}="","",IF(F{r}<=2,100,IF(F{r}<=4,90,IF(F{r}<=6,75,'
             'IF(F{r}<=9,60,IF(F{r}<=12,45,IF(F{r}<=18,30,15))))))),"")')


def is_layer10(layer: str) -> bool:
    return bool(layer) and layer.strip().startswith('10')


def is_layer9_capacity(layer: str) -> bool:
    """Rule 13 cohort: Layer 9 bitcoin-miner pivots and neoclouds.

    Hyperscalers keep standard EV/EBITDA.
    """
    if not layer or not layer.strip().startswith('09'):
        return False
    l = layer.lower()
    return 'bitcoin' in l or 'neocloud' in l


def f_segment_for(layer: str) -> str:
    if is_layer10(layer):
        return F_L10_EVFCF
    if is_layer9_capacity(layer):
        return F_L9_EVMW
    return F_STANDARD


# --- full formula templates ({r} = row, {F} = EV-column segment) --------------

VALUE = ('=IFERROR(AVERAGE('
         'IFERROR(IF(E{r}="","",IF(E{r}<0,5,IF(E{r}<=15,90,IF(E{r}<=20,75,IF(E{r}<=25,60,'
         'IF(E{r}<=30,45,IF(E{r}<=40,30,IF(E{r}<=60,15,5)))))))),""),'
         '{F},'
         'IFERROR(IF(G{r}="","",IF(G{r}>=8,100,IF(G{r}>=6,90,IF(G{r}>=4,75,IF(G{r}>=2,60,'
         'IF(G{r}>=1,45,IF(G{r}>=0,30,15))))))),""),'
         'IFERROR(IF(H{r}="","",IF(H{r}<=2,90,IF(H{r}<=4,75,IF(H{r}<=7,60,IF(H{r}<=10,45,'
         'IF(H{r}<=15,30,IF(H{r}<=25,15,5))))))),""),'
         'IFERROR(IF(I{r}="","",IF(I{r}<=0.5,100,IF(I{r}<=1,90,IF(I{r}<=1.5,75,IF(I{r}<=2,60,'
         'IF(I{r}<=2.5,45,IF(I{r}<=3,30,15))))))),"")),"")')

QUALITY = ('=IFERROR(AVERAGE('
           'IFERROR(IF(K{r}="","",IF(K{r}>=25,100,IF(K{r}>=20,90,IF(K{r}>=15,75,IF(K{r}>=10,60,'
           'IF(K{r}>=5,45,IF(K{r}>=0,30,15))))))),""),'
           'IFERROR(IF(L{r}="","",IF(L{r}>=60,100,IF(L{r}>=50,90,IF(L{r}>=40,75,IF(L{r}>=30,60,'
           'IF(L{r}>=20,45,IF(L{r}>=10,30,15))))))),""),'
           'IFERROR(IF(M{r}="","",IF(M{r}>=25,100,IF(M{r}>=20,90,IF(M{r}>=15,75,IF(M{r}>=10,60,'
           'IF(M{r}>=5,45,IF(M{r}>=0,30,15))))))),""),'
           'IFERROR(IF(N{r}="","",IF(N{r}<=0,100,IF(N{r}<=1,90,IF(N{r}<=2,75,IF(N{r}<=3,60,'
           'IF(N{r}<=4,45,IF(N{r}<=5,30,15))))))),"")),"")')

GROWTH = ('=IFERROR(AVERAGE('
          'IFERROR(IF(P{r}="","",IF(P{r}>=40,100,IF(P{r}>=30,90,IF(P{r}>=20,75,IF(P{r}>=15,60,'
          'IF(P{r}>=10,45,IF(P{r}>=5,30,15))))))),""),'
          'IFERROR(IF(Q{r}="","",IF(Q{r}>=40,100,IF(Q{r}>=30,90,IF(Q{r}>=20,75,IF(Q{r}>=10,60,'
          'IF(Q{r}>=5,45,IF(Q{r}>=0,30,15))))))),""),'
          'IFERROR(IF(R{r}="","",IF(R{r}>=40,100,IF(R{r}>=30,90,IF(R{r}>=20,75,IF(R{r}>=10,60,'
          'IF(R{r}>=0,45,IF(R{r}>=-10,30,15))))))),"")),"")')

AI = '=IFERROR(AVERAGE(T{r},U{r},V{r},W{r},X{r})*20,"")'

MOMENTUM = ('=IFERROR(AVERAGE('
            'IFERROR(IF(Z{r}="","",Z{r}*20),""),'
            'IFERROR(IF(AA{r}="","",AA{r}*20),""),'
            'IFERROR(IF(AB{r}="","",AB{r}*20),""),'
            'IFERROR(IF(AC{r}="","",IF(AC{r}>=85,100,IF(AC{r}>=70,90,IF(AC{r}>=55,75,'
            'IF(AC{r}>=40,60,IF(AC{r}>=25,40,20)))))),"")),"")')

# R5 Disruption Risk (col AL=38, added 2026-06-17): appended after Tier to avoid
# shifting Risk Score/TOTAL/Tier and breaking downstream scripts. 5 = most durable.
RISK = '=IFERROR(AVERAGE(AE{r},AF{r},AG{r},AH{r},AL{r})*20,"")'

TOTAL = ('=IFERROR(J{r}*Weights!$B$4 + O{r}*Weights!$B$5 + S{r}*Weights!$B$6 + '
         'Y{r}*Weights!$B$7 + AD{r}*Weights!$B$8 + AI{r}*Weights!$B$9,"")')

TIER = ('=IF(AJ{r}="","",IF(AJ{r}>=85,"✓✓✓",IF(AJ{r}>=70,"✓✓",'
        'IF(AJ{r}>=55,"✓",IF(AJ{r}>=40,"?","✗")))))')

COLS = {10: VALUE, 15: QUALITY, 19: GROWTH, 25: AI, 30: MOMENTUM,
        35: RISK, 36: TOTAL, 37: TIER}


def check_drift(ws) -> list[tuple[int, str, int]]:
    drifted = []
    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=1).value:
            continue
        f = ws.cell(row=r, column=10).value
        if isinstance(f, str) and f.startswith('='):
            m = re.search(r'E(\d+)', f)
            if m and int(m.group(1)) != r:
                drifted.append((r, ws.cell(row=r, column=1).value, int(m.group(1))))
    return drifted


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('--check', action='store_true', help='report drift, change nothing')
    args = ap.parse_args()

    wb = load_workbook(XLSX)
    ws = wb['Watchlist']

    drifted = check_drift(ws)
    print(f'{len(drifted)} rows with drifted formula refs')

    if args.check:
        for r, t, ref in drifted[:10]:
            print(f'  row {r} ({t}) Value formula references row {ref}')
        return

    rebuilt = 0
    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=1).value:
            continue
        layer = str(ws.cell(row=r, column=3).value or '')
        fseg = f_segment_for(layer).replace('{r}', str(r))
        for col, template in COLS.items():
            formula = template.replace('{F}', fseg).replace('{r}', str(r))
            ws.cell(row=r, column=col).value = formula
        rebuilt += 1

    wb.save(XLSX)
    print(f'Rebuilt formulas for {rebuilt} rows '
          f'(standard / Layer-10 EV-FCF / Layer-9 EV-MW conditional bands applied).')


if __name__ == '__main__':
    main()
