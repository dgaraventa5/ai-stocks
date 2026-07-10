"""Web-search fallback for the weekly scan — date-verified, source-filtered, full-watchlist.

WHY THIS EXISTS
---------------
The primary weekly-scan path (`weekly_scan_runner.py`) pulls
`data.sec.gov/submissions/CIK{cik}.json` per ticker — a deterministic, dated list
of every filing in the window. When SEC EDGAR egress is blocked (as it has been on
every cloud session since 2026-06-12), the fallback is web search. Done ad-hoc, that
fallback leaked material events for weeks (e.g. the AVGO CFO transition missed across
four scans; the MU×Anthropic deal missed across two) because:
  1. no real recency filter — "TICKER news July 2026" biases to pages that *say* the
     month, not pages *published* in the window, and result dates went unverified;
  2. headline bias — quiet 8-Ks that generate no press slipped through;
  3. coverage narrowed to holdings, not the full watchlist.

WHAT THIS FIXES (the #2–#4 fixes from the 2026-07-10 scan retro)
  #2  allowed_domains → forces quasi-primary sources (SEC mirrors, the wire services)
  #3  query by *filing*, not by vibe ("TICKER 8-K filed", "TICKER press release")
  #4  full-watchlist coverage, tiered (holdings get more queries than the long tail)

BOUNDARY (important)
  Python cannot call the agent's WebSearch tool. So this script does everything Python
  *can*: (a) generate the exact query plan the agent should execute, with the scan
  window and domain allowlist baked in; and (b) verify the harvested results — parse
  publication dates, drop anything outside the window, flag undated results as
  UNVERIFIED (never silently keep them — CLAUDE.md rule 3), and classify material vs.
  routine by a materiality-keyword net that catches the headline-light misses.

USAGE
  # 1. Emit the query plan for the agent to execute (default window = last 8 days):
  python3 scripts/web_scan.py --plan
  python3 scripts/web_scan.py --plan --from 2026-07-03 --to 2026-07-10 --json plan.json

  # 2. After running the WebSearch calls, drop the harvested hits into a JSON list of
  #    {ticker, title, url, date?, snippet?} and verify them:
  python3 scripts/web_scan.py --verify harvested.json --from 2026-07-03 --to 2026-07-10
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))  # recalc_watchlist uses flat sibling imports

from common import flag  # noqa: E402  (after sys.path tweak)
SCORING_PATH = ROOT / "00-master" / "ai_supply_chain_scoring.xlsx"
PORTFOLIO_PATH = ROOT / "00-master" / "portfolio.xlsx"

# --- #2: quasi-primary sources. Restricting a query to these raises the odds a hit is
# a filing/press release with a real date, not an SEO aggregator. SEC first, then the
# wire services companies file through, then the filing mirrors that stamp dates.
PRIMARY_DOMAINS = ["sec.gov", "businesswire.com", "globenewswire.com", "prnewswire.com"]
FILING_MIRROR_DOMAINS = ["stocktitan.net", "sec.gov"]

# Foreign ADRs that don't file 8-Ks with EDGAR (home-market disclosure only). Kept in
# sync with weekly_scan_runner.KNOWN_FOREIGN_NO_EDGAR; these still get a news query.
KNOWN_FOREIGN_NO_EDGAR = {"SBGSY", "TOELY", "BESIY", "ABBNY", "HTHIY", "SBGSY",
                          "5347.TWO", "HHUSF", "0981.HK"}

# --- #3 + materiality net. Each category's keywords are matched (case-insensitive,
# word-ish) against a result's title+snippet. A hit in ANY category => material. This
# is the net that would have flagged "Broadcom names Amie Thuener CFO" — a headline
# with no earnings, no big number, that a vibe-search buried.
MATERIAL_KEYWORDS: dict[str, list[str]] = {
    "leadership": [
        "ceo", "cfo", "cto", "coo", "chief financial officer", "chief executive",
        "chief operating officer", "chief technology officer", "resign", "retire",
        "steps down", "step down", "stepping down", "appoint", "succeed", "successor",
        "departure", "named president", "interim", "transition",
    ],
    "m_and_a": [
        "acqui", "merger", "to merge", "divest", "takeover", "buyout", "spin off",
        "spin-off", "strategic agreement", "strategic partnership", "joint venture",
        "to acquire", "stake in",
    ],
    "capacity_contract": [
        "gigawatt", "megawatt", " gw ", " mw ", "capacity", "supply agreement",
        "take-or-pay", "take or pay", "multi-year", "multiyear", " ppa", "power purchase",
        "backlog", "new customer", "design win", "wins contract", "awarded",
    ],
    "guidance_earnings": [
        "guidance", "guides", "raises outlook", "cuts outlook", "lowers", "warns",
        "preliminary results", "quarterly results", "reports q", "earnings",
        "beats estimates", "misses estimates", "profit warning",
    ],
    "financing": [
        "notes offering", "senior notes", "bond offering", "debt offering",
        "equity offering", "convertible", "private placement", "at-the-market",
        "buyback", "share repurchase", "dividend increase", "capital raise",
    ],
    "legal_reg": [
        "lawsuit", "sued", "antitrust", "fine", "investigation", "probe", "raid",
        "sec charges", "settlement", "recall", "subpoena", "injunction", "delist",
    ],
    "going_concern": [
        "going concern", "auditor change", "restatement", "bankruptcy", "chapter 11",
        "material impairment", "default",
    ],
}

MONTHS = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july", "august",
     "september", "october", "november", "december"], start=1)}


# ---------------------------------------------------------------------------
# Watchlist / portfolio loading
# ---------------------------------------------------------------------------

def _live_tiers() -> dict[str, str]:
    """Live per-ticker tier from recalc_watchlist. The sheet's Tier/TOTAL columns are
    recalc-maintained VALUES that read as blank via data_only (rule 20), so tiering off
    the sheet silently collapses every non-holding to one query. Fall back to {} if
    recalc can't run (keeps the plan working offline — holdings still get priority)."""
    try:
        import recalc_watchlist as rc
        return {r["ticker"]: (r.get("Tier") or "").strip() for r in rc.recalc()}
    except Exception as e:  # pragma: no cover - offline/degraded path
        flag(f"web_scan: recalc unavailable ({e}); tiering by holdings only")
        return {}


def load_watchlist() -> list[dict]:
    tiers = _live_tiers()
    wb = load_workbook(SCORING_PATH, data_only=True)
    ws = wb["Watchlist"]
    rows = []
    for r in range(2, ws.max_row + 1):
        tkr = ws.cell(row=r, column=1).value
        if not tkr:
            break
        tkr = str(tkr).strip().upper()
        rows.append(dict(
            ticker=tkr,
            company=str(ws.cell(row=r, column=2).value or "").strip(),
            layer=str(ws.cell(row=r, column=3).value or "").strip(),
            score=ws.cell(row=r, column=35).value,
            tier=tiers.get(tkr, str(ws.cell(row=r, column=36).value or "").strip()),
        ))
    return rows


def load_holdings() -> set[str]:
    if not PORTFOLIO_PATH.exists():
        return set()
    wb = load_workbook(PORTFOLIO_PATH, data_only=True)
    ws = wb["Targets"]
    held = set()
    for row in ws.iter_rows(values_only=True):
        if row and len(row) > 6 and row[6] == "Y" and row[0] not in (None, "Ticker"):
            held.add(str(row[0]).strip().upper())
    return held


# ---------------------------------------------------------------------------
# #4: query-plan generation (full watchlist, tiered)
# ---------------------------------------------------------------------------

def _window_phrase(scan_from: date, scan_to: date) -> str:
    """A recency anchor for the query string. Same-month → 'Month YYYY';
    cross-month → 'Mon–Mon YYYY'. It's a hint, not a filter (WebSearch has no date
    param) — the real filter is verify_results()."""
    if scan_from.year == scan_to.year and scan_from.month == scan_to.month:
        return scan_from.strftime("%B %Y")
    return f"{scan_from.strftime('%B')}-{scan_to.strftime('%B %Y')}"


def build_query_plan(watchlist: list[dict], holdings: set[str],
                     scan_from: date, scan_to: date) -> list[dict]:
    """One entry per WebSearch call the agent should run.

    Tiering (fixes #4 without exploding the call count):
      - Holdings + high-tier (✓✓✓/✓✓): 3 queries each — a filing query (SEC-scoped),
        a press-release query (wire-scoped), and an open news query. This is where a
        missed event costs the most.
      - Everyone else: 1 filing-oriented query, SEC/StockTitan-scoped, so the long tail
        is covered at low cost and each hit is likely a dated filing mirror.
    """
    win = _window_phrase(scan_from, scan_to)
    plan: list[dict] = []
    for row in watchlist:
        tkr, co = row["ticker"], row["company"]
        tier = row["tier"]
        is_priority = tkr in holdings or tier in ("✓✓✓", "✓✓")
        label = f"{tkr} ({co})" if co else tkr

        if is_priority:
            plan.append(dict(ticker=tkr, kind="filing",
                             query=f"{co or tkr} {tkr} 8-K SEC filing {win}",
                             allowed_domains=FILING_MIRROR_DOMAINS))
            plan.append(dict(ticker=tkr, kind="press",
                             query=f"{co or tkr} press release {win}",
                             allowed_domains=PRIMARY_DOMAINS))
            plan.append(dict(ticker=tkr, kind="news",
                             query=f"{co or tkr} {tkr} news {win}",
                             allowed_domains=None))
        elif tkr in KNOWN_FOREIGN_NO_EDGAR:
            # No EDGAR 8-K; home-market disclosure only → a single news query.
            plan.append(dict(ticker=tkr, kind="news",
                             query=f"{co or tkr} {tkr} news {win}",
                             allowed_domains=None))
        else:
            plan.append(dict(ticker=tkr, kind="filing",
                             query=f"{co or tkr} {tkr} 8-K filing {win}",
                             allowed_domains=FILING_MIRROR_DOMAINS))
    return plan


# ---------------------------------------------------------------------------
# Date parsing + verification (the recency filter that was missing)
# ---------------------------------------------------------------------------

def parse_date(text: str, default_year: int) -> date | None:
    """Best-effort extraction of a publication date from a title/snippet/URL.

    Handles ISO (2026-07-09), 'July 9, 2026', 'July 9' (year defaulted to the scan
    window's), '7/9/2026', '7/9/26', and URL date paths like /2026/07/09/.
    Returns None if nothing parseable is found — the caller flags that as UNVERIFIED
    rather than assuming it's in-window (CLAUDE.md rule 3).
    """
    if not text:
        return None
    t = text.strip()

    # ISO 2026-07-09
    m = re.search(r"\b(20\d{2})-(\d{1,2})-(\d{1,2})\b", t)
    if m:
        y, mo, d = map(int, m.groups())
        return _safe_date(y, mo, d)

    # URL date path /2026/07/09/
    m = re.search(r"/(20\d{2})/(\d{1,2})/(\d{1,2})[/\-]", t)
    if m:
        y, mo, d = map(int, m.groups())
        return _safe_date(y, mo, d)

    # 'July 9, 2026' or 'July 9 2026'
    m = re.search(r"\b([A-Za-z]{3,9})\.?\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(20\d{2})\b", t)
    if m:
        mon = _month(m.group(1))
        if mon:
            return _safe_date(int(m.group(3)), mon, int(m.group(2)))

    # 'July 9' (no year) → default to the scan window's year
    m = re.search(r"\b([A-Za-z]{3,9})\.?\s+(\d{1,2})(?:st|nd|rd|th)?\b", t)
    if m:
        mon = _month(m.group(1))
        if mon:
            return _safe_date(default_year, mon, int(m.group(2)))

    # '7/9/2026' or '7/9/26'
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", t)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        return _safe_date(y, mo, d)

    return None


def _month(name: str) -> int | None:
    name = name.lower()
    if name in MONTHS:
        return MONTHS[name]
    for full, idx in MONTHS.items():
        if full.startswith(name):  # 'jul' → july, 'sept' → september
            return idx
    return None


def _safe_date(y: int, mo: int, d: int) -> date | None:
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def classify(title: str, snippet: str = "") -> list[str]:
    """Return the list of materiality categories a result trips. Empty => routine."""
    blob = f" {title} {snippet} ".lower()
    hits = []
    for cat, kws in MATERIAL_KEYWORDS.items():
        if any(kw in blob for kw in kws):
            hits.append(cat)
    return hits


def verify_results(results: list[dict], scan_from: date, scan_to: date) -> dict:
    """Partition harvested hits into material / routine / unverified / out-of-window.

    Each result: {ticker, title, url, date?, snippet?}. If `date` is absent, we try to
    parse one from date+title+url+snippet. No parseable date => UNVERIFIED (surfaced,
    not dropped, not assumed in-window).
    """
    material, routine, unverified, out_of_window = [], [], [], []
    for r in results:
        raw_date = r.get("date")
        d = None
        if raw_date:
            d = parse_date(str(raw_date), scan_to.year)
        if d is None:
            # fall back to scraping a date out of title/url/snippet
            for field in ("title", "url", "snippet"):
                d = parse_date(str(r.get(field, "")), scan_to.year)
                if d:
                    break

        cats = classify(r.get("title", ""), r.get("snippet", ""))
        enriched = {**r, "parsed_date": d.isoformat() if d else None, "categories": cats}

        if d is None:
            unverified.append(enriched)
        elif d < scan_from or d > scan_to:
            out_of_window.append(enriched)
        elif cats:
            material.append(enriched)
        else:
            routine.append(enriched)

    material.sort(key=lambda x: (x.get("ticker", ""), x.get("parsed_date") or ""))
    return dict(material=material, routine=routine, unverified=unverified,
                out_of_window=out_of_window)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _default_window() -> tuple[date, date]:
    # weekly cadence; the scan looks back 8 days to overlap the prior run by a day
    # (so an event on the boundary isn't lost between two scans).
    to = date.today()
    return to - timedelta(days=8), to


def cmd_plan(args) -> None:
    scan_from = date.fromisoformat(args.from_) if args.from_ else _default_window()[0]
    scan_to = date.fromisoformat(args.to) if args.to else _default_window()[1]
    watchlist = load_watchlist()
    holdings = load_holdings()
    plan = build_query_plan(watchlist, holdings, scan_from, scan_to)

    priority = sum(1 for p in plan if p["kind"] in ("filing", "press", "news")
                   and p["ticker"] in holdings)
    print(f"# Web-scan query plan  window={scan_from}..{scan_to}  "
          f"tickers={len(watchlist)}  queries={len(plan)}  holdings={len(holdings)}")
    print("# Run each as WebSearch(query=..., allowed_domains=...). "
          "Then harvest hits to JSON and `--verify`.\n")
    for p in plan:
        dom = f"  allowed_domains={p['allowed_domains']}" if p["allowed_domains"] else ""
        print(f"[{p['ticker']:<8} {p['kind']:<6}] {p['query']}{dom}")

    if args.json:
        Path(args.json).write_text(json.dumps(
            dict(scan_from=scan_from.isoformat(), scan_to=scan_to.isoformat(), plan=plan),
            indent=2))
        print(f"\nplan written to {args.json}")


def cmd_verify(args) -> None:
    scan_from = date.fromisoformat(args.from_) if args.from_ else _default_window()[0]
    scan_to = date.fromisoformat(args.to) if args.to else _default_window()[1]
    payload = json.loads(Path(args.verify).read_text())
    results = payload if isinstance(payload, list) else payload.get("results", [])
    out = verify_results(results, scan_from, scan_to)

    print(f"# Verified against window {scan_from}..{scan_to}\n")
    print(f"MATERIAL: {len(out['material'])} | ROUTINE: {len(out['routine'])} | "
          f"UNVERIFIED(no date): {len(out['unverified'])} | "
          f"OUT-OF-WINDOW: {len(out['out_of_window'])}\n")

    if out["material"]:
        print("## ⚠️ Material (in-window, keyword-flagged)")
        for m in out["material"]:
            print(f"  {m.get('ticker',''):<8} {m['parsed_date']}  "
                  f"[{','.join(m['categories'])}]  {m.get('title','')}")
    if out["unverified"]:
        print("\n## ❓ Unverified — no parseable date; DO NOT assume in-window, re-check manually")
        for u in out["unverified"]:
            print(f"  {u.get('ticker',''):<8} {u.get('title','')}  ({u.get('url','')})")
    if args.json:
        Path(args.json).write_text(json.dumps(out, indent=2))
        print(f"\nverified results written to {args.json}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--plan", action="store_true", help="emit the query plan")
    ap.add_argument("--verify", metavar="HARVESTED.json",
                    help="verify harvested WebSearch hits against the window")
    ap.add_argument("--from", dest="from_", metavar="YYYY-MM-DD")
    ap.add_argument("--to", metavar="YYYY-MM-DD")
    ap.add_argument("--json", metavar="OUT.json", help="also write output as JSON")
    args = ap.parse_args()

    if args.plan:
        cmd_plan(args)
    elif args.verify:
        cmd_verify(args)
    else:
        ap.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
