"""Apply Layer 1 AI-Thesis re-ratings under the 'directness' lens (2026-06-10).

Context: Dom + Claude agreed the portfolio was holding second-derivative AI-power
plays (nat-gas E&Ps whose AI link runs through Henry Hub) while the direct plays
(IPPs with signed hyperscaler PPAs) sat on the bench. Per-ticker /refresh-context
research (briefings in per-stock/{T}/context-2026-06-09.md) re-scored D1/D2/D5 for
the 9 researched Layer 1 names. The other 20 Layer 1 names remain on the prior
(generous) basis until their own research pass — watchlist is mixed-basis until then.

Only D1 (col T=20), D2 (col U=21), D5 (col X=24) are touched. D3/D4 untouched.
Each change appends an Update row to Rating Audit with rationale + source + conf.
"""
from __future__ import annotations

from openpyxl import load_workbook

XLSX = '/Users/dom/Desktop/ai-stocks/00-master/ai_supply_chain_scoring.xlsx'
DATE = '2026-06-10'
COL = {'D1': 20, 'D2': 21, 'D5': 24}

# ticker -> dim -> (new_rating, rationale, source, confidence)
CHANGES = {
    'EQT': {
        'D1': (1, "Q1 2026 revenue 100% commodity (no 'data center' in 10-Q); two datacenter-linked supply deals are index-plus priced to power developers and don't flow until Shippingport fall-2026/Homer City 2028. ~0% current AI revenue.",
               'EQT 10-Q Q1 2026 (2026-04-22); S&P Global 2025-07-21', 'High'),
        'D2': (2, "Price-taker: index-plus/TETCO M-2 pricing, Q2 2026 strategic curtailments into weak prices; AI mechanism is in-basin basis (2nd-derivative). Scale+midstream are cost advantages, not pricing power.",
               'EQT 10-Q Q1 2026; S&P Global 2025-07-21', 'High'),
        'D5': (2, "No hyperscaler counterparty; deals are with power developers (Homer City Redevelopment, Frontier Group). Chain hyperscaler->unnamed tenant->developer->EQT gas = indirect.",
               'Businesswire 2025-07-15; NGI', 'High'),
    },
    'RRC': {
        'D1': (1, "~0% current AI revenue; only signed power-linked deal is 75 MMcf/d to an unnamed Midwest power plant starting late 2027 (~5% of volumes when live), datacenter linkage unconfirmed. Q1 release: zero 'data center' mentions.",
               'RRC 8-K 2026-04-21 & 2026-02-24', 'High'),
        'D2': (2, "Structural price-taker, sells below Henry Hub (FY26 guide NYMEX -$0.35/-0.45); Q1 +$0.18 premium was seasonal. One unstarted premium contract on ~5% of volumes.",
               'RRC 8-K 2026-04-21', 'High'),
        'D5': (1, "Zero signed hyperscaler/developer deals; Fort Cherry alliance partners are frac-services + real-estate (no tenant named); NextEra South Mon 'potential participation' only.",
               'Liberty Energy Apr 2025; RRC Q1 2026 call', 'High'),
    },
    'AR': {
        'D1': (1, "~0% current AI; Q1 2026 revenue 67% gas (LNG-corridor)/26% NGL (export)/2% oil, no datacenter mention in 10-Q. Only AI engagement is >5 Bcf/d of RFP participation, no wins.",
               'AR 10-Q Q1 2026 (2026-04-29); Q1 call', 'High'),
        'D2': (2, "Price-taker with superior logistics; premiums (90% Tier-1 points) derive from firm-transport + LNG/LPG export benchmarks, not negotiated with AI buyers.",
               'AR 10-Q Q1 2026; investor materials', 'High'),
        'D5': (1, "Zero signed hyperscaler/developer deals; closest is unnamed-counterparty RFPs + HG dry-gas 'optionality' + geographic-proximity narrative.",
               'AR Q1 2026 call; 10-Q', 'High'),
    },
    'EXE': {
        'D1': (1, "0% direct AI; generous indirect ~1-3% via gas-demand share runs through Henry Hub (excluded by directness lens). 10-Q mentions 'data center' once, as a price driver not a customer.",
               'EXE 10-Q Q1 2026 (2026-04-28)', 'High'),
        'D2': (2, "Price-taker, demonstrated: Q1 avg sales $4.92/Mcf vs NYMEX $5.04 (discount), hedges >65% of volumes, Delfin LNG SPA priced at Henry Hub. Scale = cost advantage not pricing power.",
               'EXE 10-Q Q1 2026', 'High'),
        'D5': (1, "Narrative-stage; no signed deal with a named hyperscaler/developer; negotiating with PJM power-provider intermediaries, unsigned. Peers (DTE, Energy Transfer) have signed deals EXE lacks.",
               'EXE Q1 2026 call (2026-04-29)', 'High'),
    },
    'CEG': {
        'D5': (5, "Best-in-layer: two named hyperscalers signed 20-yr -- MSFT 835 MW + Meta 1,121 MW (~2 GW) + CyrusOne >1,100 MW (deliveries from Q4 2026). Multiple anchor hyperscalers.",
               'CEG 10-Q 2026-05-11; company PR 2025-06-03', 'High'),
    },
    'VST': {
        'D1': (2, "Signed hyperscaler PPAs pre-revenue (Meta 2,609 MW starts late 2026; AWS 1,200 MW Q4 2027); current AI-tied revenue ~= PJM capacity ~$1.24B (~6% of rev) driven by datacenter demand. Was 5 (>50%) -- corrected to current reality.",
               'Vistra IR 2026-01-09; Q1 2026 call', 'High'),
    },
    'TLN': {
        'D1': (2, "~5% of current revenue is AWS-contracted (front-of-meter transition completed Apr 2026, ~300 MW flowing); scales to 20-30% at full ramp. Band 2 reflects revenue actually flowing now.",
               'TLN 10-Q Q1 2026 (2026-05-05); PR 2025-06-11', 'High'),
        'D5': (4, "Single named hyperscaler anchor (AWS 1,920 MW/17yr/$18B, ~77% of Susquehanna at ramp). Rubric band 5 requires MULTIPLE hyperscalers; TLN has one (large) -> band 4. Single-customer concentration also sits in R1. FLAG: magnitude argues for 5.",
               'TLN PR 2025-06-11; Power Magazine 2025-06', 'Medium'),
    },
    'NRG': {
        'D1': (1, "Current datacenter revenue ~=0.01% (5 MW contracted 2026); even full 445 MW (2032) <1% of >$30B revenue. 10-Q: zero 'hyperscaler' mentions.",
               'NRG 10-Q Q1 2026 (2026-05-06); Q1 slides', 'High'),
        'D2': (3, "Small premium-priced datacenter slice (445 MW >$80/MWh) + real 5.4 GW GE Vernova turbine-slot lockup (monetizes 2029+), but bulk is ~25 GW merchant fleet price-taking (Q1 -46% EPS miss on mild weather alone).",
               'NRG 10-Q Q1 2026; Utility Dive', 'Medium'),
        'D5': (2, "No signed deal with a named hyperscaler; 445 MW counterparties undisclosed, 6.5 GW Menlo/PowLan still LOIs, GE Vernova JV anchor unsigned. Below TLN/CEG.",
               'NRG Q1 2026 slides; Utility Dive', 'High'),
    },
    'DUK': {
        'D1': (1, "Regulated utility; 7.6 GW of signed datacenter ESAs begin delivering H2 2027 -- contracted AI revenue today ~=0%, legacy datacenter load not broken out.",
               'DUK Q1 2026 call (2026-05-05); Utility Dive', 'High'),
        'D2': (3, "Fully regulated capture: $103B capex -> ~9.6% rate-base growth -> 5-7% EPS CAGR at allowed ROE (~10%). ESA safeguards de-risk capex but don't capture scarcity upside; no pricing power.",
               'DUK Q1 2026 call; Fortune 2026-04-25', 'High'),
        'D5': (3, "Large datacenter exposure (7.6 GW signed ESAs, 15.4 GW pipeline) but zero hyperscalers named by the issuer, regulated economics, no per-customer terms disclosed.",
               'DUK Q1 2026 call; Business North Carolina', 'Medium'),
    },
}


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)
    ws = wb['Watchlist']
    audit = wb['Rating Audit']

    rows = {ws.cell(row=r, column=1).value: r
            for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}

    applied = 0
    for tkr, dims in CHANGES.items():
        if tkr not in rows:
            print(f'!! {tkr} not on Watchlist — skipped')
            continue
        r = rows[tkr]
        for dim, (new, rationale, source, conf) in dims.items():
            c = COL[dim]
            old = ws.cell(row=r, column=c).value
            ws.cell(row=r, column=c).value = new
            audit.append([DATE, tkr, dim, new,
                          f'(was {old}) {rationale}', source, conf, 'Update'])
            print(f'{tkr:<5} {dim}: {old} -> {new}')
            applied += 1
        ws.cell(row=r, column=4).value = DATE  # Last Updated

    wb.save(XLSX)
    print(f'\napplied {applied} rating changes across {len(CHANGES)} names')


if __name__ == '__main__':
    main()
