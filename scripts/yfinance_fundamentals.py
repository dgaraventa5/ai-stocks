"""Pull fundamentals from yfinance and write to per-stock/{TICKER}/financials.xlsx.

Per CLAUDE.md §3 (Flag, don't assume): missing fields are explicitly noted
in a "Data flags" sheet, not silently coerced. Per §4: history sheets are
left as plain values (already historical, no formulas to drive), but the
"Summary" tab uses formulas to derive ratios so changing the inputs
re-derives the numbers.

Usage:
  python3 scripts/yfinance_fundamentals.py NVDA
"""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from common import flag, per_stock_dir

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _is_missing(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _write_df(ws, df) -> None:
    """Write a pandas DataFrame to a sheet with the index as the first column."""
    if df is None or df.empty:
        ws.append(["(no data)"])
        return
    # Columns = period end dates (most recent first)
    ws.append(["Line item"] + [str(c.date()) if hasattr(c, "date") else str(c)
                               for c in df.columns])
    _style_header(ws)
    for idx, row in df.iterrows():
        out = [str(idx)] + [None if _is_missing(v) else float(v) for v in row.tolist()]
        ws.append(out)
    # Number format for numeric cells (USD, millions implicit since yfinance returns raw)
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            ws.cell(row=r, column=c).number_format = "#,##0"
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22 if c == 1 else 16


# yfinance annual columns are most-recent-first, so col B = latest FY.
_ANNUAL_COL = {"FY-3": "D", "FY-2": "C", "FY-1": "B"}
_CANDIDATES = {
    "rev": ["Total Revenue", "Operating Revenue"],
    "gp": ["Gross Profit"],
    "oi": ["Operating Income", "Total Operating Income As Reported"],
    "eb": ["EBITDA", "Normalized EBITDA"],
    "fcf": ["Free Cash Flow"],
    "cap": ["Capital Expenditure"],
    "nd": ["Net Debt"],
    "sh": ["Ordinary Shares Number", "Share Issued", "Common Stock Shares Outstanding"],
}


def _find_row(ws, candidates) -> int | None:
    labels = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    for c in candidates:
        if c in labels:
            return labels[c]
    return None


def add_fy_summary(wb, info: dict, flags: list) -> None:
    """Add a formula-driven 'FY Summary' tab: FY-3/FY-2/FY-1/TTM on a GAAP basis.

    Raw line-items reference the statement sheets; every ratio is an Excel
    formula (CLAUDE.md rule 4). TTM flows = Σ last 4 quarters; TTM stock items
    (net debt, shares) = most-recent quarter. Missing line-items degrade to a
    blank cell + a flag (rule 3), never a crash. Fixed 16-row layout so ratio
    formula references stay stable regardless of which inputs are present.
    """
    required = ["Income stmt (annual)", "Income stmt (quarterly)",
                "Cash flow (annual)", "Cash flow (quarterly)",
                "Balance sheet (annual)"]
    if any(s not in wb.sheetnames for s in required):
        flags.append("FY Summary: skipped — statement sheets missing")
        return
    ai, qi = wb["Income stmt (annual)"], wb["Income stmt (quarterly)"]
    ca, cq = wb["Cash flow (annual)"], wb["Cash flow (quarterly)"]
    ba = wb["Balance sheet (annual)"]
    qb = wb["Balance sheet (quarterly)"] if "Balance sheet (quarterly)" in wb.sheetnames else None

    last_q = min(5, qi.max_column)              # cap TTM at 4 quarters (cols B:E)
    last_q_letter = get_column_letter(last_q)
    if last_q < 5:
        flags.append(f"FY Summary: TTM uses only {last_q - 1} quarters (<4 available)")

    R = {
        "rev_a": _find_row(ai, _CANDIDATES["rev"]), "rev_q": _find_row(qi, _CANDIDATES["rev"]),
        "gp_a": _find_row(ai, _CANDIDATES["gp"]), "gp_q": _find_row(qi, _CANDIDATES["gp"]),
        "oi_a": _find_row(ai, _CANDIDATES["oi"]), "oi_q": _find_row(qi, _CANDIDATES["oi"]),
        "eb_a": _find_row(ai, _CANDIDATES["eb"]), "eb_q": _find_row(qi, _CANDIDATES["eb"]),
        "fcf_a": _find_row(ca, _CANDIDATES["fcf"]), "fcf_q": _find_row(cq, _CANDIDATES["fcf"]),
        "cap_a": _find_row(ca, _CANDIDATES["cap"]), "cap_q": _find_row(cq, _CANDIDATES["cap"]),
        "nd_a": _find_row(ba, _CANDIDATES["nd"]), "nd_q": _find_row(qb, _CANDIDATES["nd"]) if qb else None,
        "sh_a": _find_row(ba, _CANDIDATES["sh"]), "sh_q": _find_row(qb, _CANDIDATES["sh"]) if qb else None,
    }
    for key, label in [("rev_a", "Total Revenue"), ("gp_a", "Gross Profit"),
                       ("oi_a", "Operating Income"), ("eb_a", "EBITDA"),
                       ("fcf_a", "Free Cash Flow"), ("cap_a", "Capital Expenditure"),
                       ("nd_a", "Net Debt"), ("sh_a", "share count")]:
        if R[key] is None:
            flags.append(f"FY Summary: '{label}' not found in statements — left blank")

    def a_ref(sheet, row, fycol, scale="", neg=False):
        col = _ANNUAL_COL[fycol]
        if row is None or column_index_from_string(col) > sheet.max_column:
            return ""
        return f"={'-' if neg else ''}'{sheet.title}'!{col}{row}{scale}"

    def ttm_ref(sheet, row, neg=False):
        if row is None:
            return ""
        return f"={'-' if neg else ''}SUM('{sheet.title}'!B{row}:{last_q_letter}{row})"

    def mrq_ref(row, scale=""):
        if qb is None or row is None:
            return ""
        return f"='{qb.title}'!B{row}{scale}"

    if "FY Summary" in wb.sheetnames:
        del wb["FY Summary"]
    fy = wb.create_sheet("FY Summary", 1)

    def hdr(ws, col):
        ci = column_index_from_string(col)
        return ws.cell(row=1, column=ci).value if ci <= ws.max_column else ""

    fy.append(["Fiscal-year summary (GAAP, $ except where noted)",
               "FY-3", "FY-2", "FY-1", "TTM", "Source / formula"])
    for cell in fy[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    fy.append(["Fiscal period end", hdr(ai, "D"), hdr(ai, "C"), hdr(ai, "B"),
               f"TTM → {qi.cell(row=1, column=2).value}",
               "annual statement headers; TTM = Σ last 4 quarters"])

    def flow(label, a_sheet, a_row, q_sheet, q_row, src, neg=False):
        fy.append([label,
                   a_ref(a_sheet, a_row, "FY-3", neg=neg),
                   a_ref(a_sheet, a_row, "FY-2", neg=neg),
                   a_ref(a_sheet, a_row, "FY-1", neg=neg),
                   ttm_ref(q_sheet, q_row, neg=neg), src])

    def ratio(label, num_row, den_row, src):
        fy.append([label] + [f'=IFERROR({get_column_letter(c)}{num_row}/'
                             f'{get_column_letter(c)}{den_row},"")' for c in (2, 3, 4, 5)]
                  + [src])

    flow("Revenue", ai, R["rev_a"], qi, R["rev_q"], "Total Revenue")                  # r3
    flow("Gross profit", ai, R["gp_a"], qi, R["gp_q"], "Gross Profit (GAAP)")         # r4
    ratio("Gross margin %", 4, 3, "GAAP = gross profit / revenue")                    # r5
    flow("Operating income", ai, R["oi_a"], qi, R["oi_q"], "Operating Income")        # r6
    ratio("Operating margin %", 6, 3, "= operating income / revenue")                 # r7
    flow("EBITDA", ai, R["eb_a"], qi, R["eb_q"], "EBITDA")                            # r8
    flow("Free cash flow", ca, R["fcf_a"], cq, R["fcf_q"], "FCF = OCF − capex")       # r9
    ratio("FCF margin %", 9, 3, "= FCF / revenue")                                    # r10
    flow("Capex", ca, R["cap_a"], cq, R["cap_q"], "Capex (shown positive)", neg=True)  # r11
    ratio("Capex / revenue %", 11, 3, "= capex / revenue")                            # r12
    fy.append(["Net debt", a_ref(ba, R["nd_a"], "FY-3"), a_ref(ba, R["nd_a"], "FY-2"),  # r13
               a_ref(ba, R["nd_a"], "FY-1"),
               mrq_ref(R["nd_q"]) or a_ref(ba, R["nd_a"], "FY-1"),
               "Net debt; TTM = MRQ"])
    ratio("Net debt / EBITDA", 13, 8, "= net debt / EBITDA")                          # r14
    fy.append(["Share count (period-end, M)",                                          # r15
               a_ref(ba, R["sh_a"], "FY-3", scale="/1e6"),
               a_ref(ba, R["sh_a"], "FY-2", scale="/1e6"),
               a_ref(ba, R["sh_a"], "FY-1", scale="/1e6"),
               mrq_ref(R["sh_q"], "/1e6") or a_ref(ba, R["sh_a"], "FY-1", scale="/1e6"),
               "shares ÷ 1e6; TTM = MRQ"])
    fy.append(["Share count YoY %", "", '=IFERROR(C15/B15-1,"")',                       # r16
               '=IFERROR(D15/C15-1,"")', '=IFERROR(E15/D15-1,"")', "vs prior period"])

    pct_rows = {5, 7, 10, 12, 16}
    for r in range(3, fy.max_row + 1):
        for c in range(2, 6):
            fy.cell(row=r, column=c).number_format = "0.0%" if r in pct_rows else "#,##0"
    for c in range(2, 6):
        fy.cell(row=14, column=c).number_format = '0.00"x"'
        fy.cell(row=15, column=c).number_format = "#,##0.0"
    fy.column_dimensions["A"].width = 28
    for c in "BCDE":
        fy.column_dimensions[c].width = 15
    fy.column_dimensions["F"].width = 52

    # Non-GAAP gross-margin gap flag (generic: VMware-style amort. in COGS, etc.)
    if R["gp_q"] and R["rev_q"]:
        gp = sum((qi.cell(row=R["gp_q"], column=c).value or 0) for c in range(2, last_q + 1))
        rev = sum((qi.cell(row=R["rev_q"], column=c).value or 0) for c in range(2, last_q + 1))
        stmt_gm = gp / rev if rev else None
        info_gm = info.get("grossMargins")
        if stmt_gm and info_gm and info_gm - stmt_gm > 0.03:
            flags.append(
                f"Gross margin: GAAP (statement) ~{stmt_gm:.0%} vs yfinance "
                f"info.grossMargins {info_gm:.0%} — info is a non-GAAP basis; "
                f"FY Summary uses GAAP")


def build_financials(ticker: str, out_path: Path | None = None) -> Path:
    t = yf.Ticker(ticker)
    info = t.info or {}

    wb = Workbook()

    # ---------- Summary tab (formula-driven ratios) ----------
    s = wb.active
    s.title = "Summary"
    s.append(["Field", "Value", "Source / formula"])
    _style_header(s)

    flags: list[str] = []

    def put(label, value, source):
        if _is_missing(value):
            flags.append(f"{label}: missing from yfinance")
            s.append([label, "(missing — see Data flags)", source])
        else:
            s.append([label, value, source])

    put("Ticker", ticker.upper(), "input")
    put("Name", info.get("longName"), "yfinance info.longName")
    put("Sector", info.get("sector"), "yfinance info.sector")
    put("Industry", info.get("industry"), "yfinance info.industry")
    put("Market cap ($)", info.get("marketCap"), "yfinance info.marketCap")
    put("Enterprise value ($)", info.get("enterpriseValue"), "yfinance info.enterpriseValue")
    put("Trailing P/E", info.get("trailingPE"), "yfinance info.trailingPE")
    put("Forward P/E", info.get("forwardPE"), "yfinance info.forwardPE")
    put("P/S (TTM)", info.get("priceToSalesTrailing12Months"), "yfinance")
    put("Gross margin (TTM)", info.get("grossMargins"), "yfinance info.grossMargins")
    put("Operating margin (TTM)", info.get("operatingMargins"), "yfinance info.operatingMargins")
    put("Net margin (TTM)", info.get("profitMargins"), "yfinance info.profitMargins")
    put("ROE", info.get("returnOnEquity"), "yfinance info.returnOnEquity")
    put("Revenue growth (YoY)", info.get("revenueGrowth"), "yfinance info.revenueGrowth")
    put("Total cash ($)", info.get("totalCash"), "yfinance info.totalCash")
    put("Total debt ($)", info.get("totalDebt"), "yfinance info.totalDebt")
    # Statement-based TTM FCF (2026-06-12): info.freeCashflow is Yahoo's
    # levered estimate and does not reconcile to the cash-flow statements.
    from batch_score import statement_fcf
    _fcf = statement_fcf(t)
    if _fcf is not None:
        put("FCF (TTM, $)", _fcf, "statement: TTM OCF - |capex| (yfinance quarterly_cashflow)")
    else:
        put("FCF (TTM, $)", info.get("freeCashflow"),
            "yfinance info.freeCashflow (FALLBACK levered estimate — statement TTM unavailable)")
    put("Shares outstanding", info.get("sharesOutstanding"), "yfinance info.sharesOutstanding")
    put("Current price ($)", info.get("currentPrice") or info.get("regularMarketPrice"),
        "yfinance info.currentPrice")

    # Derived ratios as formulas, not pasted values (per CLAUDE.md §4).
    # Find row numbers by scanning column A.
    def row_of(label: str) -> int | None:
        for r in range(2, s.max_row + 1):
            if s.cell(row=r, column=1).value == label:
                return r
        return None

    fcf_row = row_of("FCF (TTM, $)")
    mcap_row = row_of("Market cap ($)")
    ev_row = row_of("Enterprise value ($)")
    if fcf_row and mcap_row:
        s.append(["FCF yield", f"=IFERROR(B{fcf_row}/B{mcap_row},\"\")",
                  f"=B{fcf_row}/B{mcap_row}"])
        s.cell(row=s.max_row, column=2).number_format = "0.00%"
    if fcf_row and ev_row:
        s.append(["FCF / EV", f"=IFERROR(B{fcf_row}/B{ev_row},\"\")",
                  f"=B{fcf_row}/B{ev_row}"])
        s.cell(row=s.max_row, column=2).number_format = "0.00%"

    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 22
    s.column_dimensions["C"].width = 36

    # ---------- Statements ----------
    for sheet_name, df in [
        ("Income stmt (annual)", t.income_stmt),
        ("Income stmt (quarterly)", t.quarterly_income_stmt),
        ("Balance sheet (annual)", t.balance_sheet),
        ("Balance sheet (quarterly)", t.quarterly_balance_sheet),
        ("Cash flow (annual)", t.cashflow),
        ("Cash flow (quarterly)", t.quarterly_cashflow),
    ]:
        ws = wb.create_sheet(sheet_name)
        try:
            _write_df(ws, df)
        except Exception as e:
            flags.append(f"{sheet_name}: yfinance returned error — {e}")
            ws.append([f"(error: {e})"])

    # ---------- FY Summary (formula-driven FY-3/FY-2/FY-1/TTM, GAAP) ----------
    add_fy_summary(wb, info, flags)

    # ---------- Data flags ----------
    fl = wb.create_sheet("Data flags")
    fl.append(["Field", "Issue"])
    _style_header(fl)
    if not flags:
        fl.append(["—", "no gaps detected"])
    else:
        for f in flags:
            label, _, issue = f.partition(":")
            fl.append([label.strip(), issue.strip()])
    fl.column_dimensions["A"].width = 30
    fl.column_dimensions["B"].width = 60

    if flags:
        for f in flags:
            flag(f)

    out = out_path or (per_stock_dir(ticker, create=True) / "financials.xlsx")
    wb.save(out)
    print(f"wrote {out}")
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="Pull fundamentals from yfinance.")
    ap.add_argument("ticker")
    ap.add_argument("--out", type=Path, default=None)
    args = ap.parse_args()
    build_financials(args.ticker, args.out)


if __name__ == "__main__":
    main()
