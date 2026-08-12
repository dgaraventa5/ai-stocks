"""refresh() gate + freeze-snapshot behaviour, over a temp workbook with
recalc / network / log_rebalance mocked."""
import sys
from pathlib import Path

import openpyxl
import pytest

pytest.importorskip('yfinance')
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'scripts'))

import refresh_targets as rt


def _build_portfolio(path, holdings, aux_sheets=True):
    """holdings: [(ticker, layer, score, tier)] written as Include?=Y rows.
    aux_sheets=False omits Reconciliation + Positions, mirroring the slimmed
    workbook the notional privacy pass (PR #9) left behind."""
    wb = openpyxl.Workbook()
    sz = wb.active
    sz.title = 'Sizing Rules'
    sz.append(['Portfolio Sizing Rules'])
    sz.append(['Parameter', 'Value', 'Notes'])
    sz.append(['Cash buffer %', 0, ''])
    sz.append(['Max single position %', 1.0, ''])  # no cap: let tier ordering show
    sz.append(['Tier-based base weight', None, None, None, None])
    sz.append(['Tier', 'Score Floor', 'Weight @ Floor', 'Weight @ Ceiling',
               'Score Ceiling'])
    sz.append(['✓✓✓', 85, 0.12, 0.12, 100])
    sz.append(['✓✓', 70, 0.03, 0.10, 85])
    sz.append(['✓', 55, 0.01, 0.03, 70])
    sz.append(['?', 40, 0, 0, 55])
    sz.append(['✗', 0, 0, 0, 40])
    sz.append(['Portfolio Value ($)', 10000, ''])
    tg = wb.create_sheet('Targets')
    tg.append(['Target Portfolio (test)'])
    tg.append(['Ticker', 'Layer', 'TOTAL', 'Tier', 'Rank', 'Status',
               'Include?', 'Override', 'Target %', 'Notes'])
    for i, (t, lay, sc, tier) in enumerate(holdings, 1):
        tg.append([t, lay, sc, tier, i, 'HOLD', 'Y', None,
                   round(100 / len(holdings), 2), None])
    if aux_sheets:
        wb.create_sheet('Reconciliation')
        wb.create_sheet('Positions').append(['Ticker', 'Company', 'Shares'])
    wb.create_sheet('README').append(['Last built', 'old'])
    wb.save(path)


def _mock_env(monkeypatch, live, cfg):
    monkeypatch.setattr(rt, 'recalc', lambda: live)
    monkeypatch.setattr(rt, 'last_trade_age_days', lambda t: 1)
    monkeypatch.setattr(rt, 'current_price', lambda t: 100.0)
    monkeypatch.setattr(rt.time, 'sleep', lambda *_a, **_k: None)
    monkeypatch.setattr(rt, 'load_cfg', lambda: cfg)
    # Pin legacy tier/score modes so these tests stay hermetic against the
    # committed portfolio-config.json (live modes flipped to inverse_vol/rank
    # at the 2026-08-07 migration). v2-mode tests override after _mock_env.
    monkeypatch.setattr(rt, 'load_pcfg', lambda: {
        'selection': {'mode': 'score'},
        'sizing': {'mode': 'tier', 'lookback': 60, 'sigma_floor': 0.005,
                   'max_weight': 0.12, 'min_weight': 0.03, 'drift_band': 0.25},
        'shadows': {'top': 15, 'next': 25, 'tail': 40}})
    calls, saves = [], []

    def fake_log(cfg_, w, reason, tiers=None, kind='membership'):
        calls.append({'reason': reason, 'tiers': tiers, 'weights': dict(w),
                      'kind': kind})
        return {'allocations': {k: v * 10000 for k, v in w.items()}, 'cash': 0.0}

    monkeypatch.setattr(rt, 'log_rebalance', fake_log)
    # Capture config saves (frozen-run pending-clock persistence) so tests can
    # assert on them AND so no test can ever write the real config file.
    monkeypatch.setattr(rt, 'save_cfg',
                        lambda cfg_: saves.append(dict(cfg_.get('exit_pending', {}))),
                        raising=False)
    return calls, saves


def test_refresh_fires_on_tier_change(monkeypatch, tmp_path):
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 78.0, '✓✓')])
    live = [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0, 'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 78.0, 'Tier': '✓✓'}]
    cfg = {'inception': '2026-05-26', 'events': [{
        'date': '2026-06-18', 'reason': 'seed',
        'allocations': {'NVDA': 500.0, 'TSM': 500.0}, 'cash': 0.0,
        'tiers': {'NVDA': '✓✓', 'TSM': '✓✓'}}]}   # NVDA was ✓✓, now ✓✓✓
    calls, _saves = _mock_env(monkeypatch, live, cfg)

    rt.refresh(portfolio=str(path))

    assert len(calls) == 1
    assert calls[0]['tiers'] == {'NVDA': '✓✓✓', 'TSM': '✓✓'}
    assert 'tier: NVDA ✓✓→✓✓✓' in calls[0]['reason']
    # Targets rewritten: NVDA (✓✓✓) now outweighs TSM (✓✓).
    tg = openpyxl.load_workbook(path)['Targets']
    w = {r[0]: r[8] for r in tg.iter_rows(min_row=3, values_only=True) if r[0]}
    assert w['NVDA'] > w['TSM']


def test_refresh_frozen_when_unchanged(monkeypatch, tmp_path):
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 78.0, '✓✓')])
    live = [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0, 'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 78.0, 'Tier': '✓✓'}]
    cfg = {'inception': '2026-05-26', 'events': [{
        'date': '2026-06-18', 'reason': 'seed',
        'allocations': {'NVDA': 500.0, 'TSM': 500.0}, 'cash': 0.0,
        'tiers': {'NVDA': '✓✓✓', 'TSM': '✓✓'}}]}   # tiers already match → no change
    calls, _saves = _mock_env(monkeypatch, live, cfg)
    before = path.read_bytes()

    rt.refresh(portfolio=str(path))

    assert calls == []                     # no rebalance event logged
    assert path.read_bytes() == before     # workbook untouched (frozen snapshot)


def test_pending_rebalance_true_on_tier_change(monkeypatch, tmp_path):
    """The rule-25 gate signal: pending_rebalance() is True when a held name's tier
    moved since the last event — and, being a dry run, writes NOTHING (a read-only
    probe, safe to call from a test/CI gate)."""
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 78.0, '✓✓')])
    live = [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0, 'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 78.0, 'Tier': '✓✓'}]
    cfg = {'inception': '2026-05-26', 'events': [{
        'date': '2026-06-18', 'reason': 'seed',
        'allocations': {'NVDA': 500.0, 'TSM': 500.0}, 'cash': 0.0,
        'tiers': {'NVDA': '✓✓', 'TSM': '✓✓'}}]}   # NVDA was ✓✓, now ✓✓✓
    calls, _saves = _mock_env(monkeypatch, live, cfg)
    before = path.read_bytes()

    assert rt.pending_rebalance(portfolio=str(path)) is True
    assert calls == []                     # dry run: nothing logged
    assert path.read_bytes() == before     # dry run: workbook untouched


def test_pending_rebalance_false_when_frozen(monkeypatch, tmp_path):
    """False when tiers/membership match the last event — so the gate passes exactly
    when the committed Targets already reflect the live scores (the steady state)."""
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 78.0, '✓✓')])
    live = [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0, 'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 78.0, 'Tier': '✓✓'}]
    cfg = {'inception': '2026-05-26', 'events': [{
        'date': '2026-06-18', 'reason': 'seed',
        'allocations': {'NVDA': 500.0, 'TSM': 500.0}, 'cash': 0.0,
        'tiers': {'NVDA': '✓✓✓', 'TSM': '✓✓'}}]}   # tiers already match
    _mock_env(monkeypatch, live, cfg)

    assert rt.pending_rebalance(portfolio=str(path)) is False


def test_refresh_tolerates_missing_recon_and_positions(monkeypatch, tmp_path):
    """The notional privacy pass (PR #9) removed Reconciliation + Positions.
    refresh() must run on the slim workbook (Sizing Rules + Targets only),
    writing just the Targets snapshot — not crash on a missing sheet (the crash
    that forced the ad-hoc Targets hand-edits this whole change replaces)."""
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 78.0, '✓✓')], aux_sheets=False)
    live = [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0, 'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 78.0, 'Tier': '✓✓'}]
    cfg = {'inception': '2026-05-26', 'events': [{
        'date': '2026-06-18', 'reason': 'seed',
        'allocations': {'NVDA': 500.0, 'TSM': 500.0}, 'cash': 0.0,
        'tiers': {'NVDA': '✓✓', 'TSM': '✓✓'}}]}   # NVDA ✓✓ → ✓✓✓ fires
    calls, _saves = _mock_env(monkeypatch, live, cfg)

    rt.refresh(portfolio=str(path))            # must not raise

    assert len(calls) == 1                     # rebalance still logged
    wb = openpyxl.load_workbook(path)
    assert 'Reconciliation' not in wb.sheetnames   # not recreated (kept slim)
    w = {r[0]: r[8] for r in wb['Targets'].iter_rows(min_row=3, values_only=True)
         if r[0]}
    assert w['NVDA'] > w['TSM']                # Targets snapshot still written


# ---- exit-pending clock persistence (2026-08-02 design) --------------------
# The 2-run exit confirm clock lives in performance-config.json (top-level
# `exit_pending` map), NOT the Targets sheet Status column: freeze-snapshot
# gating only rewrites the sheet when a rebalance fires, so a sheet-persisted
# clock never started (GOOGL/META/AMZN, observed 2026-08-02).

def _seed_cfg(exit_pending=None):
    """Two-name steady state (tiers match live) so only the exit path moves.
    shadow_events pre-seeded to the rosters _update_band_shadows would compute
    for the two-name live list, so band upkeep is a no-op in these tests and
    the save assertions isolate the pending-clock behavior."""
    cfg = {'inception': '2026-05-26', 'events': [{
        'date': '2026-06-18', 'reason': 'seed',
        'allocations': {'NVDA': 500.0, 'TSM': 500.0}, 'cash': 0.0,
        'tiers': {'NVDA': '✓✓✓', 'TSM': '✓✓'}}],
        'shadow_events': {'BAND_TOP': [
            {'date': '2026-07-30', 'roster': ['NVDA', 'TSM']}]}}
    if exit_pending is not None:
        cfg['exit_pending'] = exit_pending
    return cfg


def _below_exit_live():
    # TSM 71.0 < default exit score 73.0, but still ✓✓ (no tier crossing) and
    # still held while pending → membership unchanged → snapshot frozen.
    return [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0, 'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 71.0, 'Tier': '✓✓'}]


def test_exit_pending_clock_persists_on_frozen_run(monkeypatch, tmp_path):
    """First leg: below-exit with no other change keeps the workbook frozen but
    MUST persist the clock to the config — the bug was discarding it."""
    import datetime as dt
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 71.0, '✓✓')])
    calls, saves = _mock_env(monkeypatch, _below_exit_live(), _seed_cfg())
    before = path.read_bytes()

    rep = rt.refresh(portfolio=str(path))

    today = dt.date.today().isoformat()
    assert calls == []                          # no rebalance event
    assert path.read_bytes() == before          # workbook frozen (byte-identical)
    assert saves and saves[-1] == {'TSM': today}   # clock persisted to config
    assert rep['pending'] == {'TSM': today}


def test_exit_confirms_after_prior_day_pending(monkeypatch, tmp_path):
    """Second leg: a clock from a PRIOR date + still below exit → EXIT confirms
    (membership change fires the rebalance) and the clock entry clears."""
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 71.0, '✓✓')])
    cfg = _seed_cfg(exit_pending={'TSM': '2026-07-30'})
    calls, saves = _mock_env(monkeypatch, _below_exit_live(), cfg)

    rep = rt.refresh(portfolio=str(path))

    assert len(calls) == 1
    assert '-TSM' in calls[0]['reason']
    assert cfg['exit_pending'] == {}            # cleared before log_rebalance saves
    assert rep['pending'] == {}
    tg = openpyxl.load_workbook(path)['Targets']
    row = next(r for r in tg.iter_rows(min_row=3, values_only=True)
               if r[0] == 'TSM')
    assert row[6] == 'N'                        # excluded from the book
    assert str(row[5]).startswith('EXIT (pending since 2026-07-30')


def test_same_day_pending_does_not_confirm_or_rewrite(monkeypatch, tmp_path):
    """Two runs on the same date are one data point: no confirm, and an
    unchanged clock map is not redundantly re-saved."""
    import datetime as dt
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 71.0, '✓✓')])
    today = dt.date.today().isoformat()
    calls, saves = _mock_env(monkeypatch, _below_exit_live(),
                             _seed_cfg(exit_pending={'TSM': today}))
    before = path.read_bytes()

    rep = rt.refresh(portfolio=str(path))

    assert calls == []                          # still pending, no exit
    assert saves == []                          # map unchanged → no config write
    assert path.read_bytes() == before
    assert rep['pending'] == {'TSM': today}


def test_exit_pending_clears_on_recovery(monkeypatch, tmp_path):
    """Score back above exit resets the clock (2 CONSECUTIVE runs required) —
    and the reset itself must be persisted on an otherwise-frozen run."""
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 78.0, '✓✓')])
    live = [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0, 'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 78.0, 'Tier': '✓✓'}]
    calls, saves = _mock_env(monkeypatch, live,
                             _seed_cfg(exit_pending={'TSM': '2026-07-30'}))
    before = path.read_bytes()

    rep = rt.refresh(portfolio=str(path))

    assert calls == []
    assert path.read_bytes() == before
    assert saves and saves[-1] == {}            # cleared clock persisted
    assert rep['pending'] == {}


def test_dry_run_never_mutates_pending(monkeypatch, tmp_path):
    """--dry-run / the rule-25 gate must be able to probe repeatedly without
    starting or advancing clocks."""
    import datetime as dt
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 71.0, '✓✓')])
    cfg = _seed_cfg()
    calls, saves = _mock_env(monkeypatch, _below_exit_live(), cfg)

    rep = rt.refresh(dry_run=True, portfolio=str(path))

    assert saves == [] and calls == []
    assert 'exit_pending' not in cfg            # untouched
    today = dt.date.today().isoformat()
    assert rep['pending'] == {'TSM': today}     # ...but the report shows it


def test_pending_rebalance_true_when_confirm_due(monkeypatch, tmp_path):
    """Once a clock is a day old and the name is still below exit, the rule-25
    gate turns True — forcing the confirming run (the gate was blind to this)."""
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 71.0, '✓✓')])
    cfg = _seed_cfg(exit_pending={'TSM': '2026-07-30'})
    calls, saves = _mock_env(monkeypatch, _below_exit_live(), cfg)
    before = path.read_bytes()

    assert rt.pending_rebalance(portfolio=str(path)) is True
    assert calls == [] and saves == []          # read-only probe
    assert cfg.get('exit_pending') == {'TSM': '2026-07-30'}
    assert path.read_bytes() == before


# ---- v2: rank-mode selection + band shadow upkeep (spec 2026-08-07 B1/C1) ---

def _rank_pcfg(sel='rank'):
    return {'selection': {'mode': sel},
            'sizing': {'mode': 'tier', 'lookback': 60, 'sigma_floor': 0.005,
                       'max_weight': 0.12, 'min_weight': 0.03,
                       'drift_band': 0.25},
            'shadows': {'top': 15, 'next': 25, 'tail': 40}}


def _rank_live(n=20):
    """T01 (score 90, rank 1) .. T20 (score 71, rank 20), all same tier."""
    return [{'ticker': f'T{i:02d}', 'layer': '06 Silicon',
             'TOTAL': 90.0 - (i - 1), 'Tier': '✓✓'} for i in range(1, n + 1)]


def _rank_cfg(prior, exit_pending=None):
    cfg = {'inception': '2026-05-26', 'events': [{
        'date': '2026-06-18', 'reason': 'seed',
        'allocations': {t: 500.0 for t in prior}, 'cash': 0.0,
        'tiers': {t: '✓✓' for t in prior}}]}
    if exit_pending is not None:
        cfg['exit_pending'] = exit_pending
    return cfg


def _rank_prior():
    return [f'T{i:02d}' for i in range(1, 14)] + ['T17', 'T19']


def test_rank_mode_entry_and_deadband(monkeypatch, tmp_path):
    """Outsiders at ranks 14-15 ENTER; rank-16 outsider stays out (dead-band);
    rank-17 incumbent HOLDs; rank-19 incumbent starts the exit clock (not an
    immediate exit — the rule-26 2-run confirm wraps rank crossings too)."""
    import datetime as dt
    path = tmp_path / 'portfolio.xlsx'
    prior = _rank_prior()
    _build_portfolio(path, [(t, '06 Silicon', 90.0, '✓✓') for t in prior])
    calls, _saves = _mock_env(monkeypatch, _rank_live(), _rank_cfg(prior))
    monkeypatch.setattr(rt, 'load_pcfg', _rank_pcfg)

    rep = rt.refresh(portfolio=str(path))

    assert rep['entered'] == ['T14', 'T15']
    assert rep['exited'] == []                       # T19 pending, still held
    today = dt.date.today().isoformat()
    assert rep['pending'] == {'T19': today}          # clock started
    assert len(calls) == 1 and calls[0]['kind'] == 'membership'
    tg = openpyxl.load_workbook(path)['Targets']
    inc = {r[0]: r[6] for r in tg.iter_rows(min_row=3, values_only=True) if r[0]}
    assert inc['T16'] == 'N'                         # dead-band outsider
    assert inc['T17'] == 'Y'                         # dead-band incumbent
    assert inc['T19'] == 'Y'                         # held while pending


def test_rank_mode_exit_confirms_next_run(monkeypatch, tmp_path):
    """A prior-dated clock + still below rank M -> the exit confirms."""
    path = tmp_path / 'portfolio.xlsx'
    prior = _rank_prior()
    _build_portfolio(path, [(t, '06 Silicon', 90.0, '✓✓') for t in prior])
    calls, _saves = _mock_env(monkeypatch, _rank_live(),
                              _rank_cfg(prior, exit_pending={'T19': '2026-07-30'}))
    monkeypatch.setattr(rt, 'load_pcfg', _rank_pcfg)

    rep = rt.refresh(portfolio=str(path))

    assert 'T19' in rep['exited']
    assert rep['pending'] == {}                      # clock consumed
    assert len(calls) == 1
    assert '-T19' in calls[0]['reason']


def test_shadow_events_appended_on_real_run(monkeypatch, tmp_path):
    path = tmp_path / 'portfolio.xlsx'
    prior = _rank_prior()
    _build_portfolio(path, [(t, '06 Silicon', 90.0, '✓✓') for t in prior])
    cfg = _rank_cfg(prior)
    calls, _saves = _mock_env(monkeypatch, _rank_live(), cfg)
    monkeypatch.setattr(rt, 'load_pcfg', _rank_pcfg)
    import datetime as dt
    today = dt.date.today().isoformat()

    rt.refresh(portfolio=str(path))

    top = cfg['shadow_events']['BAND_TOP']
    assert len(top) == 1 and top[0]['date'] == today
    assert len(top[0]['roster']) == 15               # ranks 1-15
    assert top[0]['roster'][0] == 'T01'
    nxt = cfg['shadow_events']['BAND_NEXT']
    assert nxt[0]['roster'] == [f'T{i:02d}' for i in range(16, 21)]
    assert 'BAND_TAIL' not in cfg['shadow_events'] or \
        cfg['shadow_events']['BAND_TAIL'] == []      # only 20 names live
    before = {k: [dict(e) for e in v] for k, v in cfg['shadow_events'].items()}
    rt.refresh(portfolio=str(path))                  # same-day re-run
    assert cfg['shadow_events'] == before            # idempotent, no dupes


def test_shadow_events_not_touched_on_dry_run(monkeypatch, tmp_path):
    path = tmp_path / 'portfolio.xlsx'
    prior = _rank_prior()
    _build_portfolio(path, [(t, '06 Silicon', 90.0, '✓✓') for t in prior])
    cfg = _rank_cfg(prior)
    calls, saves = _mock_env(monkeypatch, _rank_live(), cfg)
    monkeypatch.setattr(rt, 'load_pcfg', _rank_pcfg)

    rt.refresh(dry_run=True, portfolio=str(path))

    assert 'shadow_events' not in cfg
    assert saves == [] and calls == []


# ---- v2: inverse-vol sizing mode + monthly drift-band re-size (spec A1/A2) --

def _invvol_pcfg():
    return {'selection': {'mode': 'score'},
            'sizing': {'mode': 'inverse_vol', 'lookback': 60,
                       'sigma_floor': 0.005, 'max_weight': 1.0,
                       'min_weight': 0.0, 'drift_band': 0.25},
            'shadows': {'top': 15, 'next': 25, 'tail': 40}}


def _vol_frame():
    """NVDA at 2x TSM's daily vol -> inverse-vol targets 1/3 vs 2/3."""
    pd = pytest.importorskip('pandas')
    np = pytest.importorskip('numpy')
    idx = pd.bdate_range('2026-05-01', periods=61)
    return pd.DataFrame(
        {'NVDA': 100 * np.cumprod([1] + [1.02, 0.98] * 30)[:61],
         'TSM': 100 * np.cumprod([1] + [1.01, 0.99] * 30)[:61]}, index=idx)


def _two_name_live():
    return [{'ticker': 'NVDA', 'layer': '06 Silicon', 'TOTAL': 86.0,
             'Tier': '✓✓✓'},
            {'ticker': 'TSM', 'layer': '05 Fabs', 'TOTAL': 78.0, 'Tier': '✓✓'}]


def _invvol_env(monkeypatch, tmp_path, cfg):
    path = tmp_path / 'portfolio.xlsx'
    _build_portfolio(path, [('NVDA', '06 Silicon', 86.0, '✓✓✓'),
                            ('TSM', '05 Fabs', 78.0, '✓✓')])
    calls, saves = _mock_env(monkeypatch, _two_name_live(), cfg)
    monkeypatch.setattr(rt, 'load_pcfg', _invvol_pcfg)
    monkeypatch.setattr(rt, '_price_frame', lambda tickers, lb: _vol_frame())
    return path, calls, saves


def test_invvol_mode_weights_from_vol_not_score(monkeypatch, tmp_path):
    """Higher-vol NVDA gets the SMALLER weight despite the higher score —
    the score chooses the names, trailing vol chooses the sizes (spec A0/A1)."""
    cfg = _seed_cfg()
    path, calls, _saves = _invvol_env(monkeypatch, tmp_path, cfg)

    rt.refresh(portfolio=str(path), resize=True)     # force a re-size event

    assert len(calls) == 1 and calls[0]['kind'] == 'manual_resize'
    w = calls[0]['weights']
    assert w['TSM'] == pytest.approx(2 * w['NVDA'], rel=1e-2)   # half the vol


def test_monthly_resize_fires_outside_band(monkeypatch, tmp_path):
    cfg = _seed_cfg()                                # no sizing_state yet
    path, calls, _saves = _invvol_env(monkeypatch, tmp_path, cfg)
    # Book drifted to 60/40; inverse-vol targets ~33/67 -> both outside ±25%.
    monkeypatch.setattr(rt, 'current_weights',
                        lambda c: {'NVDA': 0.60, 'TSM': 0.40})
    import datetime as dt

    rt.refresh(portfolio=str(path))

    assert len(calls) == 1 and calls[0]['kind'] == 'resize_monthly'
    assert 'NVDA' in calls[0]['reason'] and 'TSM' in calls[0]['reason']
    assert cfg['sizing_state']['last_resize_check'] == \
        dt.date.today().isoformat()[:7]


def test_monthly_resize_noop_inside_band(monkeypatch, tmp_path):
    cfg = _seed_cfg()
    path, calls, saves = _invvol_env(monkeypatch, tmp_path, cfg)
    probes = []

    def cur(c):
        probes.append(1)
        # match the invvol targets (1/3, 2/3) -> inside the band
        return {'NVDA': 1 / 3, 'TSM': 2 / 3}

    monkeypatch.setattr(rt, 'current_weights', cur)
    import datetime as dt

    rt.refresh(portfolio=str(path))

    assert calls == []                               # no event
    assert cfg['sizing_state']['last_resize_check'] == \
        dt.date.today().isoformat()[:7]              # month stamped
    assert saves                                     # stamp persisted (frozen run)
    rt.refresh(portfolio=str(path))                  # same month: no re-check
    assert probes == [1]                             # evaluated exactly once


def test_dry_run_never_touches_sizing_state(monkeypatch, tmp_path):
    cfg = _seed_cfg()
    path, calls, saves = _invvol_env(monkeypatch, tmp_path, cfg)
    monkeypatch.setattr(rt, 'current_weights',
                        lambda c: {'NVDA': 0.60, 'TSM': 0.40})

    rt.refresh(dry_run=True, portfolio=str(path))

    assert 'sizing_state' not in cfg
    assert calls == [] and saves == []


def test_offline_gate_stays_networkless_in_invvol_mode(monkeypatch, tmp_path):
    """pending_rebalance (rule-25 gate) must not fetch prices under invvol."""
    cfg = _seed_cfg()
    path, calls, saves = _invvol_env(monkeypatch, tmp_path, cfg)

    def boom(tickers, lb):
        raise AssertionError('offline gate fetched prices')

    monkeypatch.setattr(rt, '_price_frame', boom)
    assert rt.pending_rebalance(portfolio=str(path)) is False
    assert 'sizing_state' not in cfg and saves == []


# ---- tradability filter (spec 2026-08-11) ----------------------------------

def _tradability_pcfg():
    p = _rank_pcfg()
    p['selection']['tradable_only'] = True
    return p


def _mixed_live():
    """Ranks 1-17 with a foreign name at raw rank 3 and one at 16."""
    tickers = ['T01', 'T02', '6861.T', 'T04', 'T05', 'T06', 'T07', 'T08',
               'T09', 'T10', 'T11', 'T12', 'T13', 'T14', 'T15', '6268.T', 'T17']
    return [{'ticker': t, 'layer': '11 Robotics', 'TOTAL': 90.0 - i,
             'Tier': '✓✓'} for i, t in enumerate(tickers)]


def test_tradable_only_blocks_foreign_entry_and_exits_held(monkeypatch, tmp_path):
    """Held 6861.T exits IMMEDIATELY (no pending clock); foreign outsiders
    never enter; tradable names below them shift up into the entry band."""
    path = tmp_path / 'portfolio.xlsx'
    prior = ['T01', 'T02', '6861.T', 'T04', 'T05', 'T06', 'T07', 'T08',
             'T09', 'T10', 'T11', 'T12', 'T13', 'T14']
    _build_portfolio(path, [(t, '11 Robotics', 90.0, '✓✓') for t in prior])
    calls, _saves = _mock_env(monkeypatch, _mixed_live(), _rank_cfg(prior))
    monkeypatch.setattr(rt, 'load_pcfg', _tradability_pcfg)

    rep = rt.refresh(portfolio=str(path))

    assert '6861.T' in rep['exited']                 # immediate, this run
    assert rep['pending'] == {}                      # never on the clock
    assert 'T15' in rep['entered']                   # shifted into rank<=15
    assert '6268.T' not in rep['entered']
    assert len(calls) == 1 and calls[0]['kind'] == 'membership'
    assert '6861.T' not in calls[0]['weights']
    tg = openpyxl.load_workbook(path)['Targets']
    rows = {r[0]: r for r in tg.iter_rows(min_row=3, values_only=True) if r[0]}
    assert rows['6861.T'][6] == 'N'
    assert 'untradable' in str(rows['6861.T'][5])    # Status column


def test_tradable_only_off_keeps_foreign(monkeypatch, tmp_path):
    """Flag off (default): foreign names rank and hold exactly as before."""
    path = tmp_path / 'portfolio.xlsx'
    prior = ['T01', 'T02', '6861.T', 'T04', 'T05', 'T06', 'T07', 'T08',
             'T09', 'T10', 'T11', 'T12', 'T13', 'T14']
    _build_portfolio(path, [(t, '11 Robotics', 90.0, '✓✓') for t in prior])
    calls, _saves = _mock_env(monkeypatch, _mixed_live(), _rank_cfg(prior))
    monkeypatch.setattr(rt, 'load_pcfg', _rank_pcfg)   # no tradable_only key

    rep = rt.refresh(portfolio=str(path))

    assert rep['exited'] == []
    assert '6861.T' not in rep.get('pending', {})
