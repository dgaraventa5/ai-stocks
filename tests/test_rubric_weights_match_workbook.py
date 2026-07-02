"""P0 regression gate: the rating rubric is a MIRROR; the Weights tab is the
single source of truth.

Asserts every category weight *quoted* in templates/rating-rubric-and-workflow.md
(both the canonical weights table and the inline "NN% category weight" headers)
equals the live Weights tab in 00-master/ai_supply_chain_scoring.xlsx.

This is the cheap regression gate against the F6 class of drift surfaced in the
2026-07-01 scoring critique: the rubric said AI Thesis 30% while the model used
20%, so anyone rating a stock off the rubric calibrated to weights the engine
does not use. If these ever diverge again, this test fails the build.
"""
import re
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
RUBRIC = REPO / 'templates' / 'rating-rubric-and-workflow.md'
XLSX = REPO / '00-master' / 'ai_supply_chain_scoring.xlsx'

CATEGORIES = ['Value', 'Quality', 'Growth', 'AI Thesis', 'Momentum', 'Risk']


def workbook_weights():
    """The authoritative weights, as integer percents keyed by category name."""
    wb = load_workbook(XLSX, data_only=False)
    ws = wb['Weights']
    out = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 2:
            continue
        name, wt = row[0], row[1]
        if name in CATEGORIES and isinstance(wt, (int, float)):
            out[name] = round(wt * 100)
    return out


def test_workbook_weights_are_complete_and_sum_to_100():
    w = workbook_weights()
    assert set(w) == set(CATEGORIES), f"Weights tab categories = {sorted(w)}"
    assert sum(w.values()) == 100, f"weights sum to {sum(w.values())}, not 100"


def test_rubric_canonical_table_matches_workbook():
    """The rubric must carry a canonical 6-row weights table, one row per
    category, each matching the live Weights tab exactly."""
    text = RUBRIC.read_text()
    w = workbook_weights()
    for cat in CATEGORIES:
        m = re.search(rf'^\|\s*{re.escape(cat)}\s*\|\s*(\d+)%\s*\|', text, re.M)
        assert m, f"rubric canonical weights table is missing a row for {cat!r}"
        assert int(m.group(1)) == w[cat], (
            f"rubric canonical table quotes {cat} at {m.group(1)}% "
            f"but the Weights tab says {w[cat]}%")


def test_rubric_inline_headers_match_workbook():
    """Every inline '### <Category> (... NN% category weight)' header must match
    the live Weights tab. This is the exact F6 defect: AI Thesis quoted at 30%."""
    text = RUBRIC.read_text()
    w = workbook_weights()
    mentions = re.findall(r'###\s+([A-Za-z ]+?)\s+\([^)]*?(\d+)% category weight',
                          text)
    assert mentions, "expected at least one inline 'NN% category weight' header"
    for cat_raw, pct in mentions:
        cat = cat_raw.strip()
        assert cat in w, f"inline weight header names an unknown category {cat!r}"
        assert int(pct) == w[cat], (
            f"inline header quotes {cat} at {pct}% "
            f"but the Weights tab says {w[cat]}%")
