import json
import datetime as dt
from pathlib import Path
from openpyxl import Workbook, load_workbook
import scripts.refresh_objective_inputs as roi


def _make_scoring(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    headers = ["Ticker", "Company", "Layer", "Last Updated", "Fwd P/E"]
    ws.append(headers)
    for t in ["NVDA", "TSM", "MU", "GEV"]:
        ws.append([t, f"{t} Inc", "06 Silicon", None, None])
    p = tmp_path / "scoring.xlsx"
    wb.save(p)
    return p


def _make_portfolio(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Targets"
    ws.append(["Target Portfolio", None, None, None, None, None])
    ws.append(["Ticker", "Layer", "TOTAL", "Tier", "Rank", "Status"])
    ws.append(["NVDA", "06", 84.0, "✓✓", 1, "HOLD"])
    ws.append(["MU", "06", 79.9, "✓✓", 2, "HOLD"])
    ws.append(["XYZ", "06", 70.0, "✓✓", 3, None])  # not HOLD
    p = tmp_path / "portfolio.xlsx"
    wb.save(p)
    return p


def test_resolve_portfolio_returns_hold_tickers(tmp_path):
    sp = _make_scoring(tmp_path)
    pp = _make_portfolio(tmp_path)
    assert roi.resolve_targets("portfolio", scoring_path=sp, portfolio_path=pp) == ["NVDA", "MU"]


def test_resolve_all_returns_every_watchlist_ticker(tmp_path):
    sp = _make_scoring(tmp_path)
    pp = _make_portfolio(tmp_path)
    assert roi.resolve_targets("all", scoring_path=sp, portfolio_path=pp) == ["NVDA", "TSM", "MU", "GEV"]


def test_resolve_subset_drops_unknowns(tmp_path):
    sp = _make_scoring(tmp_path)
    pp = _make_portfolio(tmp_path)
    assert roi.resolve_targets(["nvda", "ZZZZ"], scoring_path=sp, portfolio_path=pp) == ["NVDA"]


L9 = "09 Compute-as-a-Service / Neocloud GPU clouds"
TODAY = dt.date(2026, 6, 24)


def _cap(tmp_path, payload):
    p = tmp_path / "capacity-mw.json"
    p.write_text(json.dumps(payload))
    return p


def test_mw_flag_missing_entry(tmp_path):
    cj = _cap(tmp_path, {})
    f = roi.mw_staleness_flag("CRWV", L9, TODAY, capacity_json=cj)
    assert f is not None and "EV/MW" in f and "missing" in f.lower()


def test_mw_flag_stale_asof(tmp_path):
    # Real capacity-mw.json uses the key "asof" (no underscore) — see _meta/CRWV.
    cj = _cap(tmp_path, {"CRWV": {"secured_gross_mw": 300, "asof": "2026-03-01"}})
    f = roi.mw_staleness_flag("CRWV", L9, TODAY, capacity_json=cj)
    assert f is not None and "EV/MW" in f and "days old" in f.lower()


def test_mw_flag_fresh_returns_none(tmp_path):
    cj = _cap(tmp_path, {"CRWV": {"secured_gross_mw": 300, "asof": "2026-06-01"}})
    assert roi.mw_staleness_flag("CRWV", L9, TODAY, capacity_json=cj) is None


def test_mw_flag_non_l9_returns_none(tmp_path):
    cj = _cap(tmp_path, {})
    assert roi.mw_staleness_flag("NVDA", "06 Silicon", TODAY, capacity_json=cj) is None


# ---------------------------------------------------------------------------
# Task 3: apply_guards tests
# ---------------------------------------------------------------------------

def _blank_existing():
    return {k: None for k in roi.OBJ_COLS}


def test_guard_currency_blanks_adr_cols():
    info = {"financialCurrency": "TWD"}
    fresh = {"fwd_pe": 23.4, "ev_ebitda": 15.0, "fcf_yield": 3.2, "ps": 8.0,
             "roic": 20.0, "gross_mgn": 55.0, "fcf_mgn": 30.0, "nd_ebitda": 0.5,
             "rev_3y_cagr": 25.0, "rev_yoy": 30.0, "eps_yoy": 40.0}
    writes, flags = roi.apply_guards(info, fresh, _blank_existing())
    assert writes["ps"] is None and writes["ev_ebitda"] is None and writes["fcf_yield"] is None
    assert writes["fwd_pe"] == 23.4  # kept
    assert any("ADR" in f or "currency" in f.lower() for f in flags)


def test_guard_keeps_existing_on_fetched_none():
    info = {"financialCurrency": "USD"}
    fresh = {k: None for k in roi.OBJ_COLS}
    fresh["fwd_pe"] = 40.0
    existing = _blank_existing()
    existing["roic"] = 18.5  # human-curated
    writes, flags = roi.apply_guards(info, fresh, existing)
    assert "roic" not in writes  # untouched, not clobbered with None
    assert writes["fwd_pe"] == 40.0
    assert any("roic" in f.lower() for f in flags)


def test_guard_eps_yoy_preserves_blank():
    info = {"financialCurrency": "USD"}
    fresh = {k: 10.0 for k in roi.OBJ_COLS}
    fresh["eps_yoy"] = 50.0
    existing = {k: 9.0 for k in roi.OBJ_COLS}
    existing["eps_yoy"] = None  # deliberately blank (GEV-style)
    writes, flags = roi.apply_guards(info, fresh, existing)
    assert "eps_yoy" not in writes
    assert any("eps yoy" in f.lower() and "blank" in f.lower() for f in flags)


def test_guard_eps_yoy_withholds_extreme():
    info = {"financialCurrency": "USD"}
    fresh = {k: 10.0 for k in roi.OBJ_COLS}
    fresh["eps_yoy"] = 412.0
    existing = {k: 9.0 for k in roi.OBJ_COLS}
    writes, flags = roi.apply_guards(info, fresh, existing)
    assert "eps_yoy" not in writes
    assert any("412" in f or "extreme" in f.lower() for f in flags)


def test_guard_eps_yoy_normal_writes():
    info = {"financialCurrency": "USD"}
    fresh = {k: 10.0 for k in roi.OBJ_COLS}
    fresh["eps_yoy"] = 80.0
    existing = {k: 9.0 for k in roi.OBJ_COLS}
    writes, flags = roi.apply_guards(info, fresh, existing)
    assert writes["eps_yoy"] == 80.0


def test_guard_eps_yoy_keeps_existing_on_fetched_none():
    info = {"financialCurrency": "USD"}
    fresh = {k: 10.0 for k in roi.OBJ_COLS}
    fresh["eps_yoy"] = None  # yfinance earningsQuarterlyGrowth frequently None
    existing = {k: 9.0 for k in roi.OBJ_COLS}
    existing["eps_yoy"] = 80.0  # previously-valid, must not be clobbered
    writes, flags = roi.apply_guards(info, fresh, existing)
    assert "eps_yoy" not in writes
    assert any("eps_yoy" in f.lower() and ("kept" in f.lower() or "no data" in f.lower()) for f in flags)


# ---------------------------------------------------------------------------
# Task 4: row read/write helpers
# ---------------------------------------------------------------------------

def _ws_one_row():
    wb = Workbook()
    ws = wb.active
    # 38 columns; put a sentinel in every subjective col so we can prove they're untouched
    for c in range(1, 39):
        ws.cell(row=2, column=c, value=f"S{c}")
    return ws


def test_write_row_only_touches_objective_and_meta():
    ws = _ws_one_row()
    writes = {"fwd_pe": 41.1, "ps": None}
    touched = roi.write_row(ws, 2, writes, dma_value=78.0, today_iso="2026-06-24")
    assert ws.cell(row=2, column=5).value == 41.1
    assert ws.cell(row=2, column=8).value is None
    assert ws.cell(row=2, column=29).value == 78.0
    assert ws.cell(row=2, column=4).value == "2026-06-24"
    # subjective cols untouched
    for c in (20, 21, 22, 23, 24, 26, 27, 28, 31, 32, 33, 34, 38):
        assert ws.cell(row=2, column=c).value == f"S{c}"
    # score/formula cols untouched
    for c in (9, 10, 15, 19, 25, 30, 35, 36, 37):
        assert ws.cell(row=2, column=c).value == f"S{c}"
    assert set(touched) == {4, 5, 8, 29}


def test_read_existing_pulls_objective_keys():
    ws = _ws_one_row()
    ws.cell(row=2, column=11, value=18.5)
    existing = roi.read_existing(ws, 2)
    assert existing["roic"] == 18.5
    assert set(existing) == set(roi.OBJ_COLS)


# ---------------------------------------------------------------------------
# Task 5: refresh() orchestration
# ---------------------------------------------------------------------------

def test_refresh_dry_run_writes_nothing(tmp_path):
    # build a scoring workbook with NVDA + a layer in col 3
    wb = Workbook(); ws = wb.active; ws.title = "Watchlist"
    ws.append(["Ticker", "Company", "Layer", "Last Updated"] + [None]*34)
    ws.append(["NVDA", "Nvidia", "06 Silicon", "2026-01-01"] + [None]*34)
    sp = tmp_path / "scoring.xlsx"; wb.save(sp)

    def fake_fetch(ticker, layer):
        info = {"financialCurrency": "USD"}
        fresh = {k: 10.0 for k in roi.OBJ_COLS}
        return info, fresh

    rep = roi.refresh(["NVDA"], dry_run=True, scoring_path=sp,
                      fetcher=fake_fetch, dma_fetcher=lambda t: 70.0,
                      today=dt.date(2026, 6, 24))
    # nothing written: Last Updated still original
    ws2 = load_workbook(sp)["Watchlist"]
    assert ws2.cell(row=2, column=4).value == "2026-01-01"
    assert ws2.cell(row=2, column=5).value is None
    assert rep["wrote"] is False


def test_refresh_live_writes_and_bumps_last_updated(tmp_path):
    wb = Workbook(); ws = wb.active; ws.title = "Watchlist"
    ws.append(["Ticker", "Company", "Layer", "Last Updated"] + [None]*34)
    ws.append(["NVDA", "Nvidia", "06 Silicon", "2026-01-01"] + [None]*34)
    sp = tmp_path / "scoring.xlsx"; wb.save(sp)

    def fake_fetch(ticker, layer):
        return {"financialCurrency": "USD"}, {k: 10.0 for k in roi.OBJ_COLS}

    roi.refresh(["NVDA"], dry_run=False, scoring_path=sp,
                fetcher=fake_fetch, dma_fetcher=lambda t: 70.0,
                today=dt.date(2026, 6, 24))
    ws2 = load_workbook(sp)["Watchlist"]
    assert ws2.cell(row=2, column=4).value == "2026-06-24"
    assert ws2.cell(row=2, column=5).value == 10.0
    assert ws2.cell(row=2, column=29).value == 70.0


def test_guard_ev_ebitda_nd_ebitda_blank_through_on_none():
    """ev_ebitda and nd_ebitda must blank-through on fresh None (guard-2/3/4 wins over guard-5).
    Contrast: roic fresh None over non-blank existing → kept prior (guard-5 still applies)."""
    info = {"financialCurrency": "USD"}
    # Fresh: ev_ebitda + nd_ebitda are None (e.g. non-positive EBITDA); all others have values
    fresh = {k: 10.0 for k in roi.OBJ_COLS}
    fresh["ev_ebitda"] = None
    fresh["nd_ebitda"] = None
    fresh["roic"] = None  # also None — but roic is NOT in _BLANK_THROUGH_ON_NONE
    # Existing: every key has a prior non-blank value
    existing = {k: 9.0 for k in roi.OBJ_COLS}

    writes, flags = roi.apply_guards(info, fresh, existing)

    # ev_ebitda and nd_ebitda: present in writes as None (cell blanked, not kept stale)
    assert "ev_ebitda" in writes and writes["ev_ebitda"] is None, (
        "ev_ebitda fresh None over non-blank existing must blank-through (guard-2)"
    )
    assert "nd_ebitda" in writes and writes["nd_ebitda"] is None, (
        "nd_ebitda fresh None over non-blank existing must blank-through (guard-2)"
    )
    # A flag must mention the blanking for at least one of these keys
    assert any(
        ("ev_ebitda" in f or "nd_ebitda" in f) and "blank" in f.lower()
        for f in flags
    ), f"Expected a blanked flag for ev_ebitda/nd_ebitda in: {flags}"

    # Contrast: roic fresh None over non-blank existing → key ABSENT from writes (kept prior)
    assert "roic" not in writes, (
        "roic fresh None over non-blank existing must keep prior (guard-5), not blank-through"
    )


def test_refresh_isolates_per_ticker_fetch_error(tmp_path):
    # A transient fetch error on one ticker must not abort the whole run:
    # the bad row is left unchanged + flagged; later tickers still process.
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    ws.append(["Ticker", "Company", "Layer", "Last Updated"] + [None] * 34)
    ws.append(["BAD", "Bad Inc", "06 Silicon", "2026-01-01"] + [None] * 34)
    ws.append(["NVDA", "Nvidia", "06 Silicon", "2026-01-01"] + [None] * 34)
    sp = tmp_path / "scoring.xlsx"
    wb.save(sp)

    def flaky(ticker, layer):
        if ticker == "BAD":
            raise RuntimeError("yfinance boom")
        return {"financialCurrency": "USD"}, {k: 10.0 for k in roi.OBJ_COLS}

    rep = roi.refresh(["BAD", "NVDA"], dry_run=False, scoring_path=sp,
                      fetcher=flaky, dma_fetcher=lambda t: 70.0,
                      today=dt.date(2026, 6, 24))
    ws2 = load_workbook(sp)["Watchlist"]
    # BAD row untouched (no clobber with bad data)
    assert ws2.cell(row=2, column=4).value == "2026-01-01"
    assert ws2.cell(row=2, column=5).value is None
    # NVDA still processed
    assert ws2.cell(row=3, column=4).value == "2026-06-24"
    assert ws2.cell(row=3, column=5).value == 10.0
    # error surfaced as a flag
    assert any("BAD" in f and "fetch error" in f.lower() for f in rep["flags"])


def test_module_imports_without_yfinance():
    # The site deploy runs `pytest tests` with ONLY openpyxl+pytest installed
    # (no yfinance). This module must import cleanly there — yfinance/batch_score
    # imports are lazy. Simulate by blocking yfinance in a subprocess.
    import subprocess, sys, textwrap
    code = textwrap.dedent('''
        import sys
        class _Block:
            def find_spec(self, name, path=None, target=None):
                if name == "yfinance" or name.startswith("yfinance."):
                    raise ModuleNotFoundError("No module named 'yfinance'")
                return None
        sys.meta_path.insert(0, _Block())
        import scripts.refresh_objective_inputs as roi
        assert "yfinance" not in sys.modules, "yfinance imported at module load"
        assert "batch_score" not in sys.modules, "batch_score imported at module load"
        # pure paths must work without yfinance
        roi.apply_guards({"financialCurrency": "USD"},
                         {k: 1.0 for k in roi.OBJ_COLS},
                         {k: None for k in roi.OBJ_COLS})
        roi.mw_staleness_flag("NVDA", "06 Silicon", __import__("datetime").date(2026,6,24))
    ''')
    repo_root = Path(__file__).resolve().parent.parent
    r = subprocess.run([sys.executable, "-c", code], capture_output=True, text=True,
                       cwd=str(repo_root))
    assert r.returncode == 0, f"import-without-yfinance failed:\n{r.stderr}"
