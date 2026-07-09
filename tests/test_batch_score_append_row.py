"""append_row (batch_score) column-ownership + workflow regression.

Guards the 2026-07-08 off-by-one fix. append_row's formula list (SCORE_COLS) had
drifted from the live Watchlist layout — it still named MomentumScore(29)/
RiskScore(34)/TOTAL(35)/Tier(36) while the real columns are Momentum Score=30,
Risk Score=35, TOTAL=36, Tier=37 (50DMA%=29 and Reg Risk=34 are INPUTS). The
guard `src.startswith("=")` meant the bogus input cols were harmlessly skipped,
but Momentum Score(30) and Tier(37) were simply absent — so a row added by
batch_score had a blank Momentum Score and blank Tier (hence a broken TOTAL, which
references Momentum) unless the mandatory rebuild_watchlist_formulas.py follow-up
ran and papered over it.

Post-fix contract (option a, single source of truth):
  * append_row uniquely owns PEG(9); it writes NO category/TOTAL/Tier formula.
  * rebuild_watchlist_formulas.py owns the other eight computed columns
    {10,15,19,25,30,35,36,37}.
  * The documented two-step workflow (append_row -> rebuild) yields a complete row.

batch_score imports yfinance at module load, so this module is skipped in the
minimal deploy CI (openpyxl+pytest only), where batch_score isn't exercised.
"""
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

pytest.importorskip("yfinance")  # batch_score imports yfinance at module load

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
import batch_score  # noqa: E402
import rebuild_watchlist_formulas as rebuild  # noqa: E402

PEG_COL = 9
MOMENTUM_COL, RISK_COL, TOTAL_COL, TIER_COL = 30, 35, 36, 37

# The columns rebuild_watchlist_formulas.py owns. append_row must NOT write these
# (writing them is the duplicated knowledge that let the two lists drift apart).
REBUILD_OWNED = set(rebuild.COLS)
# Layout guard: fail loudly here if the Watchlist columns ever shift again, so the
# rest of this file can't silently test the wrong columns.
assert REBUILD_OWNED == {10, 15, 19, 25, 30, 35, 36, 37}, REBUILD_OWNED

# The 11 objective inputs append_row expects (values are arbitrary — we assert on
# which cells become formulas, not on their numbers).
INPUTS = {
    "fwd_pe": 25.0, "ev_ebitda": 12.0, "fcf_yield": 3.0, "ps": 8.0,
    "roic": 15.0, "gross_mgn": 60.0, "fcf_mgn": 20.0, "nd_ebitda": 1.0,
    "rev_3y_cagr": 30.0, "rev_yoy": 25.0, "eps_yoy": 18.0,
}


def _source_ws():
    """A miniature Watchlist whose row 2 mirrors the live formula/value layout:
    PEG(9) and the category/TOTAL/Tier cols are formulas; Value(10)/Quality(15)
    are recalc-maintained VALUES (rule 20); 50DMA%(29)/Reg Risk(34) are input
    VALUES. This is the layout append_row copies row-2 formulas from."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"

    header = [None] * 38
    for col, name in {
        1: "Ticker", 3: "Layer", 5: "Fwd P/E", 9: "PEG",
        10: "Value Score", 15: "Quality Score", 18: "EPS YoY %",
        19: "Growth Score", 25: "AI Score", 29: "50DMA %",
        30: "Momentum Score", 34: "Reg Risk", 35: "Risk Score",
        36: "TOTAL", 37: "Tier", 38: "Disrupt Risk",
    }.items():
        header[col - 1] = name
    ws.append(header)

    # row 2 — a realistic scored source row
    ws.cell(row=2, column=1, value="SRC")
    ws.cell(row=2, column=3, value="06 Silicon")   # standard layer -> F_STANDARD
    ws.cell(row=2, column=5, value=25.0)           # Fwd P/E (E)
    ws.cell(row=2, column=18, value=18.0)          # EPS YoY (R)
    ws.cell(row=2, column=PEG_COL,
            value='=IF(OR(E2="",R2="",R2<=0,E2<0),"",E2/R2)')
    ws.cell(row=2, column=10, value=55.0)          # Value Score — VALUE (rule 20)
    ws.cell(row=2, column=15, value=60.0)          # Quality Score — VALUE (rule 20)
    # category/TOTAL/Tier formulas present in the source row (as on the live sheet)
    for col in (19, 25, 30, 35, 36, 37):
        ws.cell(row=2, column=col,
                value=f'=IFERROR({get_column_letter(col)}2,"")')
    ws.cell(row=2, column=29, value=84.2)          # 50DMA % input
    ws.cell(row=2, column=34, value=4)             # Reg Risk input
    ws.cell(row=2, column=38, value=5)             # Disrupt Risk input
    return ws


def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")


def test_append_row_owns_only_peg_and_delegates_the_rest():
    """append_row writes the PEG(9) formula (retargeted) and leaves every
    rebuild-owned category/TOTAL/Tier column blank — the single-source-of-truth
    partition that stops the two formula lists from drifting apart again."""
    ws = _source_ws()
    r = 5
    batch_score.append_row(ws, r, "NVDA", "06 Silicon", INPUTS)

    # append_row's unique job: the PEG formula, retargeted to the new row.
    peg = ws.cell(row=r, column=PEG_COL).value
    assert isinstance(peg, str) and peg.startswith("="), f"PEG not a formula: {peg!r}"
    assert f"E{r}" in peg and f"R{r}" in peg, f"PEG not retargeted to row {r}: {peg!r}"
    assert "E2" not in peg and "R2" not in peg, f"PEG still references row 2: {peg!r}"

    # Everything rebuild owns must be untouched by append_row. Before the fix,
    # append_row copied Growth(19)/AI(25)/Risk(35)/TOTAL(36) here — the duplication.
    wrote = sorted(c for c in REBUILD_OWNED if _is_formula(ws.cell(row=r, column=c)))
    assert wrote == [], (
        f"append_row wrote formulas into rebuild-owned columns {wrote}; those belong "
        "to rebuild_watchlist_formulas.py (run it after batch_score)."
    )


def test_append_row_then_rebuild_yields_momentum_and_tier():
    """The documented workflow: append_row, then rebuild_watchlist_formulas.py,
    produces a row with a populated Momentum Score(30), Tier(37) and TOTAL(36) —
    the exact cells the off-by-one bug left blank."""
    ws = _source_ws()
    r = 5
    batch_score.append_row(ws, r, "NVDA", "06 Silicon", INPUTS)

    # Apply rebuild_watchlist_formulas.py's logic to the new row (as main() does).
    layer = str(ws.cell(row=r, column=3).value or "")
    fseg = rebuild.f_segment_for(layer).replace("{r}", str(r))
    for col, template in rebuild.COLS.items():
        ws.cell(row=r, column=col).value = (
            template.replace("{F}", fseg).replace("{r}", str(r)))

    mom = ws.cell(row=r, column=MOMENTUM_COL).value
    tier = ws.cell(row=r, column=TIER_COL).value
    total = ws.cell(row=r, column=TOTAL_COL).value
    for name, f in [("Momentum", mom), ("Tier", tier), ("TOTAL", total)]:
        assert isinstance(f, str) and f.startswith("="), f"{name} not a formula: {f!r}"

    # ...and each references the new row, not the row-2 template.
    assert f"AC{r}" in mom, f"Momentum not retargeted to row {r}: {mom!r}"
    assert f"AJ{r}" in tier, f"Tier not retargeted to row {r}: {tier!r}"
    assert f"AD{r}" in total, f"TOTAL does not reference Momentum row {r}: {total!r}"
