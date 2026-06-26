import datetime as dt
from openpyxl import Workbook
import log_forecast as lf

TODAY = dt.date(2026, 6, 26)


def _scoring(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    headers = ["Ticker", "Company", "Layer"] + [None] * 23 + ["Rel Str"]  # col 27 (AA)
    ws.append(headers)
    # 6 Layer-06 names so AVGO gets a real cohort; ratings drive default prob
    data = [("AVGO", 4), ("NVDA", 5), ("MU", 3), ("ALAB", 2), ("SNDK", 1), ("WDC", 4)]
    for t, rating in data:
        row = [t, f"{t} Inc", "06 Silicon"] + [None] * 23 + [rating]
        ws.append(row)
    p = tmp_path / "scoring.xlsx"
    wb.save(p)
    return p


def test_build_proposals_maps_rating_to_default_prob(tmp_path, monkeypatch):
    p = _scoring(tmp_path)
    monkeypatch.setattr(lf, "resolve_targets", lambda *a, **k: ["AVGO", "NVDA"])
    proposals = lf.build_proposals(TODAY, scoring_path=p)
    by_ticker = {x["ticker"]: x for x in proposals}
    assert by_ticker["AVGO"]["probability"] == lf.DEFAULT_PROB[4]   # rating 4 -> 0.65
    assert by_ticker["NVDA"]["probability"] == lf.DEFAULT_PROB[5]   # rating 5 -> 0.80
    a = by_ticker["AVGO"]
    assert a["template"] == "REL_STRENGTH_1Q" and a["dimension"] == "momentum.rel_strength"
    assert a["layer"] == "06" and a["rating_value"] == 4
    assert "AVGO" not in a["resolution_rule"]["constituents"]
    assert dt.date.fromisoformat(a["resolution_date"]) > TODAY   # no look-ahead


def test_missing_rating_defaults_to_base_rate(tmp_path, monkeypatch):
    p = _scoring(tmp_path)
    monkeypatch.setattr(lf, "resolve_targets", lambda *a, **k: ["MU"])
    # blank MU's rating
    from openpyxl import load_workbook
    wb = load_workbook(p); ws = wb["Watchlist"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "MU":
            ws.cell(r, 27).value = None
    wb.save(p)
    proposals = lf.build_proposals(TODAY, scoring_path=p)
    assert proposals[0]["probability"] == lf.BASE_RATE_PROB
    assert proposals[0]["rating_value"] is None
