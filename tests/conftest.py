"""Fixture data for export_site_data tests.

Builds miniature versions of the real workbooks with DISTINCTIVE real-dollar
values (13800, 824.7, 13386.55) that the privacy regression test asserts
never appear in exported output.
"""
import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'scripts'))

REAL_DOLLARS = ['13800', '824.7', '13386.55', '13903.5',
                '900.0', '700.0', '500.0', '71.72', '50.0']   # planted; must never leak


@pytest.fixture
def repo(tmp_path):
    """A miniature repo tree with every input the exporter reads."""
    (tmp_path / '00-master').mkdir()
    (tmp_path / 'tracking').mkdir()
    (tmp_path / 'site' / 'data').mkdir(parents=True)

    # --- portfolio.xlsx: Targets sheet, headers in row 2, data from row 3
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Targets'
    ws.append(['Target Portfolio (test)'])
    ws.append(['Ticker', 'Layer', 'TOTAL', 'Tier', 'Rank', 'Status',
               'Include?', 'Override', 'Target %', 'Notes'])
    ws.append(['NVDA', '06 AI Compute Silicon / GPUs', 83.3, '✓✓', 1,
               'HOLD', 'Y', None, 60.0, None])
    ws.append(['TSM', '05 Fabs & Foundries', 78.1, '✓✓', 2,
               'HOLD', 'Y', None, 40.0, None])
    ws.append(['ARM', '06 Silicon - IP', 66.0, '✓', 3,
               'EXIT', 'N', None, 0.0, None])
    wb.save(tmp_path / '00-master' / 'portfolio.xlsx')

    # --- scoring.xlsx: Watchlist + Rating Audit
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Watchlist'
    ws.append(['Ticker', 'Company', 'Layer'])
    ws.append(['NVDA', 'NVIDIA Corp', '06'])
    ws.append(['TSM', 'TSMC', '05'])
    wa = wb.create_sheet('Rating Audit')
    wa.append(['Date', 'Ticker', 'Dimension', 'Rating', 'Rationale',
               'Source', 'Confidence', 'Type'])
    wa.append(['2026-05-17', 'NVDA', 'D1', 5, 'Initial baseline rating.',
               'src', 'High', 'Initial'])
    wa.append(['2026-06-10', 'NVDA', 'D2', 4, 'Re-rated on directness lens.',
               'src', 'High', 'Re-rate'])
    wa.append(['2026-06-10', 'NVDA', 'D3', 3, 'Second dimension same day.',
               'src', 'Medium', 'Re-rate'])
    wb.save(tmp_path / '00-master' / 'ai_supply_chain_scoring.xlsx')

    # --- performance-config.json (events: inception, then TSM enters / ARM exits)
    cfg = {
        'inception': '2026-05-26', 'capital': 13800.0, 'cash': 71.72,
        'benchmarks': ['SMH', 'QQQ'], 'ew_universe': ['NVDA', 'TSM'],
        'events': [
            {'date': '2026-05-26', 'reason': 'initial deployment',
             'allocations': {'NVDA': 824.7, 'ARM': 500.0}, 'cash': 71.72},
            {'date': '2026-06-10', 'reason': 'rebalance (score refresh)',
             'allocations': {'NVDA': 900.0, 'TSM': 700.0}, 'cash': 50.0},
        ],
    }
    (tmp_path / 'tracking' / 'performance-config.json').write_text(
        json.dumps(cfg))

    # --- performance-series.json (3 days, real dollars incl. planted 13386.55)
    series = {
        'start': '2026-05-26', 'as_of': '2026-05-28',
        'dates': ['2026-05-26', '2026-05-27', '2026-05-28'],
        'model': [13800.0, 13903.5, 13386.55],
        'bench': {'SMH': [1.0, 1.01, 0.98], 'QQQ': [1.0, 1.005, 0.99],
                  'SPY': [1.0, 1.002, 0.95],
                  'EW': [1.0, 1.02, 0.97]},
    }
    (tmp_path / 'tracking' / 'performance-series.json').write_text(
        json.dumps(series))

    # --- thesis files: NVDA populated, TSM still the template
    nv = tmp_path / 'per-stock' / 'NVDA'
    nv.mkdir(parents=True)
    (nv / 'thesis.md').write_text(
        '# NVDA — NVIDIA\n\n## 1. One-line thesis\n\n'
        '> The compute layer toll booth: every training run pays NVDA.\n\n'
        '---\n\n## 2. Position in the AI supply chain\n\ntext\n')
    ts = tmp_path / 'per-stock' / 'TSM'
    ts.mkdir(parents=True)
    (ts / 'thesis.md').write_text(
        '# {TICKER} — {Company Name}\n\n## 1. One-line thesis\n\n'
        '> {Single sentence: why does this stock exist in the portfolio?}\n\n'
        '---\n\n## 2. Position in the AI supply chain\n')

    # --- notion scan links
    (tmp_path / 'tracking' / 'notion-scan-links.json').write_text(json.dumps([
        {'date': '2026-06-05', 'title': 'Weekly News Scan — 2026-06-05',
         'url': 'https://www.notion.so/abc123'},
    ]))
    (tmp_path / 'tracking' / 'weekly-news-scan-2026-06-05.md').write_text('x')
    (tmp_path / 'tracking' / 'weekly-news-scan-2026-05-29.md').write_text('x')

    return tmp_path
