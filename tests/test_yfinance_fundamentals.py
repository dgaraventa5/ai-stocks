"""Tests for the FY Summary builder in yfinance_fundamentals.

The builder writes Excel *formulas* (CLAUDE.md rule 4), so the tests evaluate
them with a small resolver to assert computed behaviour, not formula strings.
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

import pytest

# yfinance_fundamentals imports yfinance at module load; deploy-site.yml CI is
# intentionally yfinance-free (CLAUDE.md), so skip this builder test there rather
# than break pytest collection — which silently froze the site deploy (PR #6/#7).
pytest.importorskip("yfinance")

import yfinance_fundamentals as yf  # noqa: E402


# ---- tiny formula resolver (handles only the forms the builder emits) ----
def _num(v):
    if v in (None, ""):
        raise ValueError("blank")
    return float(v)


def _resolve(wb, sheet, v):
    if not isinstance(v, str) or not v.startswith("="):
        return v
    expr = v[1:]
    m = re.fullmatch(r'IFERROR\((.*),""\)', expr)
    if m:                                  # emulate Excel IFERROR(.., "")
        try:
            return _resolve(wb, sheet, "=" + m.group(1))
        except Exception:
            return ""
    m = re.fullmatch(r"(-?)SUM\('([^']+)'!([A-Z]+)(\d+):([A-Z]+)(\d+)\)", expr)
    if m:
        sign = -1 if m.group(1) else 1
        sh, c1, r, c2 = m.group(2), column_index_from_string(m.group(3)), \
            int(m.group(4)), column_index_from_string(m.group(5))
        return sign * sum(_num(wb[sh].cell(row=r, column=c).value)
                          for c in range(c1, c2 + 1))

    def cross(mo):
        return str(_num(_cell(wb, mo.group(1), mo.group(2))))
    expr = re.sub(r"'([^']+)'!\$?([A-Z]+\$?\d+)", cross, expr)

    def bare(mo):
        return str(_num(_cell(wb, sheet, mo.group(0))))
    expr = re.sub(r"(?<![A-Za-z0-9_$'!])\$?[A-Z]{1,2}\$?\d+", bare, expr)
    return eval(expr)  # noqa: S307 — trusted, test-only arithmetic


def _cell(wb, sheet, coord):
    coord = coord.replace("$", "")
    return _resolve(wb, sheet, wb[sheet][coord].value)


# ---- fixture workbook with statement sheets (round numbers) ----
def _sheet(wb, name, header, rows):
    ws = wb.create_sheet(name)
    ws.append(["Line item"] + header)
    for label, vals in rows:
        ws.append([label] + vals)
    return ws


def _fixture_wb(drop_gross_profit=False):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ai = [
        ("Total Revenue", [1000, 800, 600]),
        ("Gross Profit", [700, 520, 390]),
        ("Operating Income", [400, 240, 180]),
        ("EBITDA", [500, 320, 240]),
    ]
    if drop_gross_profit:
        ai = [r for r in ai if r[0] != "Gross Profit"]
    _sheet(wb, "Income stmt (annual)", ["2025-10-31", "2024-10-31", "2023-10-31"], ai)
    _sheet(wb, "Income stmt (quarterly)",
           ["2026-04-30", "2026-01-31", "2025-10-31", "2025-07-31"],
           [("Total Revenue", [300, 280, 260, 240]),
            ("Gross Profit", [210, 196, 182, 168]),
            ("Operating Income", [120, 112, 104, 96]),
            ("EBITDA", [150, 140, 130, 120])])
    _sheet(wb, "Cash flow (annual)", ["2025-10-31", "2024-10-31", "2023-10-31"],
           [("Free Cash Flow", [300, 200, 150]),
            ("Capital Expenditure", [-50, -40, -30])])
    _sheet(wb, "Cash flow (quarterly)",
           ["2026-04-30", "2026-01-31", "2025-10-31", "2025-07-31"],
           [("Free Cash Flow", [90, 84, 78, 72]),
            ("Capital Expenditure", [-15, -14, -13, -12])])
    _sheet(wb, "Balance sheet (annual)", ["2025-10-31", "2024-10-31", "2023-10-31"],
           [("Net Debt", [250, 300, 200]),
            ("Ordinary Shares Number", [1_000_000_000, 950_000_000, 900_000_000])])
    _sheet(wb, "Balance sheet (quarterly)", ["2026-04-30", "2026-01-31"],
           [("Net Debt", [240, 245]),
            ("Ordinary Shares Number", [1_020_000_000, 1_010_000_000])])
    return wb


def _fy_row(fy, label):
    for r in range(1, fy.max_row + 1):
        if fy.cell(row=r, column=1).value and \
                str(fy.cell(row=r, column=1).value).startswith(label):
            return r
    raise KeyError(label)


def test_fy_summary_computes_margins_and_ttm_from_statements():
    wb = _fixture_wb()
    yf.add_fy_summary(wb, info={}, flags=[])
    fy = wb["FY Summary"]
    # columns: B=FY-3, C=FY-2, D=FY-1, E=TTM
    gm = _fy_row(fy, "Gross margin")
    assert _cell(wb, "FY Summary", f"B{gm}") == 390 / 600    # FY-3
    assert _cell(wb, "FY Summary", f"D{gm}") == 700 / 1000   # FY-1
    assert _cell(wb, "FY Summary", f"E{gm}") == 756 / 1080   # TTM = Σ4 quarters

    rev = _fy_row(fy, "Revenue")
    assert _cell(wb, "FY Summary", f"E{rev}") == 1080        # TTM revenue

    nde = _fy_row(fy, "Net debt / EBITDA")
    assert _cell(wb, "FY Summary", f"E{nde}") == 240 / 540   # MRQ net debt / TTM EBITDA

    cap = _fy_row(fy, "Capex")
    assert _cell(wb, "FY Summary", f"E{cap}") == 54          # shown positive

    yoy = _fy_row(fy, "Share count YoY")
    assert abs(_cell(wb, "FY Summary", f"D{yoy}") - (1000 / 950 - 1)) < 1e-9


def test_fy_summary_degrades_gracefully_when_line_item_missing():
    wb = _fixture_wb(drop_gross_profit=True)
    flags = []
    yf.add_fy_summary(wb, info={}, flags=flags)
    fy = wb["FY Summary"]
    gp = _fy_row(fy, "Gross profit")
    gm = _fy_row(fy, "Gross margin")
    assert fy.cell(row=gp, column=4).value in (None, "")       # raw blank, no crash
    assert _cell(wb, "FY Summary", f"D{gm}") == ""             # IFERROR → blank
    assert any("Gross Profit" in f for f in flags)             # flagged (rule 3)


def test_fy_summary_flags_nongaap_gross_margin_gap():
    wb = _fixture_wb()
    flags = []
    # statement TTM GM is 70%; info says 80% → material non-GAAP gap
    yf.add_fy_summary(wb, info={"grossMargins": 0.80}, flags=flags)
    assert any("non-GAAP" in f or "gross margin" in f.lower() for f in flags)
